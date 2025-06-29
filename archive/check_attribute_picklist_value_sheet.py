#!/usr/bin/env python3
"""
Check why AttributePicklistValue records are not loading.
"""

import pandas as pd
import openpyxl
from pathlib import Path

def check_sheet():
    """Check the AttributePicklistValue sheet contents."""
    print("=" * 60)
    print("CHECKING ATTRIBUTEPICKLISTVALUE SHEET")
    print("=" * 60)
    
    workbook_path = Path('data/Revenue_Cloud_Complete_Upload_Template.xlsx')
    
    # Read the sheet
    try:
        df = pd.read_excel(workbook_path, sheet_name='18_AttributePicklistValue')
        
        print(f"\nSheet found: 18_AttributePicklistValue")
        print(f"Total rows (including header): {len(df) + 1}")
        print(f"Columns: {list(df.columns)}")
        
        # Remove completely empty rows
        df_clean = df.dropna(how='all')
        print(f"\nNon-empty rows: {len(df_clean)}")
        
        if len(df_clean) > 0:
            print("\nFirst few records:")
            print(df_clean.head())
            
            # Check for key fields
            if 'Name*' in df.columns or 'Name' in df.columns:
                name_col = 'Name*' if 'Name*' in df.columns else 'Name'
                non_empty_names = df_clean[df_clean[name_col].notna()]
                print(f"\nRecords with Name values: {len(non_empty_names)}")
                
                if len(non_empty_names) > 0:
                    print("\nSample data:")
                    for idx, row in non_empty_names.head(3).iterrows():
                        print(f"\nRow {idx + 2}:")  # +2 for Excel row number
                        for col in df.columns:
                            if pd.notna(row[col]):
                                print(f"  {col}: {row[col]}")
        else:
            print("\nNo data found in sheet (all rows are empty)")
            
        # Check using openpyxl to see actual cell values
        print("\n" + "-" * 60)
        print("Checking with openpyxl:")
        
        wb = openpyxl.load_workbook(workbook_path)
        ws = wb['18_AttributePicklistValue']
        
        print(f"Sheet dimensions: {ws.max_row} rows x {ws.max_column} columns")
        
        # Check first 5 rows
        print("\nFirst 5 rows:")
        for row_idx in range(1, min(6, ws.max_row + 1)):
            row_data = []
            for col_idx in range(1, min(12, ws.max_column + 1)):  # First 11 columns
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    row_data.append(f"{cell_value}")
            if row_data:
                print(f"Row {row_idx}: {' | '.join(row_data[:5])}...")  # Show first 5 values
                
    except Exception as e:
        print(f"\nError reading sheet: {str(e)}")

def check_existing_data():
    """Check if there are any existing AttributePicklistValue records."""
    print("\n" + "=" * 60)
    print("CHECKING EXISTING ATTRIBUTEPICKLISTVALUE RECORDS")
    print("=" * 60)
    
    import subprocess
    import json
    
    query = "SELECT Id, Name, PicklistId, Value FROM AttributePicklistValue LIMIT 5"
    
    cmd = [
        'sf', 'data', 'query',
        '--query', query,
        '--target-org', 'fortradp2',
        '--json'
    ]
    
    result = subprocess.run(cmd, capture_output=True, text=True)
    
    if result.returncode == 0:
        data = json.loads(result.stdout)
        if 'result' in data:
            total = data['result'].get('totalSize', 0)
            records = data['result'].get('records', [])
            
            print(f"\nTotal AttributePicklistValue records in org: {total}")
            
            if records:
                print("\nSample records:")
                for rec in records:
                    print(f"\n- Name: {rec.get('Name')}")
                    print(f"  Value: {rec.get('Value')}")
                    print(f"  PicklistId: {rec.get('PicklistId')}")
    else:
        print(f"Query failed: {result.stderr}")

def check_required_fields():
    """Check which fields are required for AttributePicklistValue."""
    print("\n" + "=" * 60)
    print("CHECKING REQUIRED FIELDS FOR ATTRIBUTEPICKLISTVALUE")
    print("=" * 60)
    
    import subprocess
    import json
    
    cmd = [
        'sf', 'sobject', 'describe',
        '--sobject', 'AttributePicklistValue',
        '--target-org', 'fortradp2',
        '--json'
    ]
    
    result = subprocess.run(cmd, capture_output=True, text=True)
    
    if result.returncode == 0:
        data = json.loads(result.stdout)
        if 'result' in data and 'fields' in data['result']:
            fields = data['result']['fields']
            
            print("\nRequired fields:")
            for field in fields:
                if not field.get('nillable', True) and field.get('createable', False):
                    print(f"- {field['name']} ({field['type']})")
                    
            print("\nAll createable fields:")
            createable = [f for f in fields if f.get('createable', False)]
            for field in createable[:10]:  # Show first 10
                req = " (REQUIRED)" if not field.get('nillable', True) else ""
                print(f"- {field['name']}: {field['type']}{req}")
    else:
        print(f"Describe failed: {result.stderr}")

def main():
    check_sheet()
    check_existing_data()
    check_required_fields()

if __name__ == '__main__':
    main()