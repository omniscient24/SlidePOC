#\!/usr/bin/env python3
"""
Populate ProductRelatedComponent IDs from org.
"""

import subprocess
import json
import pandas as pd
from openpyxl import load_workbook

def main():
    print("Populating ProductRelatedComponent IDs...")
    
    # Get existing records from org
    query = "SELECT Id, ParentProductId, ChildProductId FROM ProductRelatedComponent"
    
    cmd = [
        'sf', 'data', 'query',
        '--query', query,
        '--target-org', 'fortradp2',
        '--json'
    ]
    
    result = subprocess.run(cmd, capture_output=True, text=True)
    
    id_map = {}
    if result.returncode == 0:
        data = json.loads(result.stdout)
        if 'result' in data and 'records' in data['result']:
            for rec in data['result']['records']:
                key = f"{rec['ParentProductId']}|{rec['ChildProductId']}"
                id_map[key] = rec['Id']
    
    print(f"Found {len(id_map)} existing component records")
    
    # Update workbook
    wb = load_workbook('data/Revenue_Cloud_Complete_Upload_Template.xlsx')
    ws = wb['25_ProductRelatedComponent']
    
    # Find column positions
    headers = [cell.value for cell in ws[1] if cell.value]
    id_col = None
    parent_col = None
    child_col = None
    
    for i, header in enumerate(headers, 1):
        clean_header = str(header).replace('*', '').strip()
        if clean_header == 'Id':
            id_col = i
        elif clean_header == 'ParentProductId':
            parent_col = i
        elif clean_header == 'ChildProductId':
            child_col = i
    
    # Populate IDs
    mapped = 0
    for row in range(2, ws.max_row + 1):
        if parent_col and child_col:
            parent_id = ws.cell(row=row, column=parent_col).value
            child_id = ws.cell(row=row, column=child_col).value
            
            if parent_id and child_id:
                key = f"{parent_id}|{child_id}"
                if key in id_map and id_col:
                    ws.cell(row=row, column=id_col).value = id_map[key]
                    mapped += 1
    
    print(f"Mapped {mapped} IDs")
    
    # Save workbook
    wb.save('data/Revenue_Cloud_Complete_Upload_Template.xlsx')
    print("✓ Updated workbook")

if __name__ == '__main__':
    main()
