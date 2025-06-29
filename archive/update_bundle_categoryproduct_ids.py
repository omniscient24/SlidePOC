#!/usr/bin/env python3
"""
Update bundle ProductCategoryProduct IDs.
"""

from openpyxl import load_workbook

def main():
    print("Updating bundle ProductCategoryProduct IDs...")
    
    # Known mappings from query
    bundle_ids = {
        '01tdp000006JEGkAAO': '0ZRdp0000000ALxGAM',  # DCS Elite
        '01tdp000006JEGlAAO': '0ZRdp0000000ALyGAM',  # DCS Essentials
        '01tdp000006JEGjAAO': '0ZRdp0000000ALzGAM',  # DCS Advanced
    }
    
    # Load workbook
    wb = load_workbook('data/Revenue_Cloud_Complete_Upload_Template.xlsx')
    ws = wb['26_ProductCategoryProduct']
    
    # Find column positions
    id_col = None
    product_col = None
    
    for col, cell in enumerate(ws[1], 1):
        if cell.value:
            clean_value = str(cell.value).replace('*', '').strip()
            if clean_value == 'Id':
                id_col = col
            elif clean_value == 'ProductId':
                product_col = col
    
    # Update IDs for bundle products
    updated = 0
    for row in range(2, ws.max_row + 1):
        if product_col and id_col:
            product_id = ws.cell(row=row, column=product_col).value
            if product_id in bundle_ids:
                current_id = ws.cell(row=row, column=id_col).value
                if not current_id:
                    ws.cell(row=row, column=id_col).value = bundle_ids[product_id]
                    updated += 1
                    print(f"  Updated row {row}: {product_id} -> {bundle_ids[product_id]}")
    
    print(f"\nUpdated {updated} bundle product IDs")
    
    # Save workbook
    wb.save('data/Revenue_Cloud_Complete_Upload_Template.xlsx')
    print("✓ Updated workbook")

if __name__ == '__main__':
    main()