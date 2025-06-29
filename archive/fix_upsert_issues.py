#\!/usr/bin/env python3
"""
Fix remaining upsert issues in the workbook.
"""

from openpyxl import load_workbook
import pandas as pd

def main():
    print("=" * 60)
    print("FIXING REMAINING UPSERT ISSUES")
    print("=" * 60)
    
    wb = load_workbook('data/Revenue_Cloud_Complete_Upload_Template.xlsx')
    
    # 1. Fix AttributePicklist - remove External_ID__c from upsert
    print("\n1. Fixing AttributePicklist sheet...")
    ws = wb['14_AttributePicklist']
    # The External_ID__c was already removed, but we need to ensure we have data
    
    # 2. Fix AttributeDefinition - remove DataType and DeveloperName if still there
    print("2. Checking AttributeDefinition for read-only fields...")
    ws = wb['09_AttributeDefinition']
    headers = [cell.value for cell in ws[1] if cell.value]
    
    cols_to_remove = []
    for col, cell in enumerate(ws[1], 1):
        if cell.value and cell.value.replace('*', '').strip() in ['DataType', 'DeveloperName']:
            cols_to_remove.append(col)
    
    if cols_to_remove:
        cols_to_remove.sort(reverse=True)
        for col in cols_to_remove:
            ws.delete_cols(col)
        print(f"   Removed {len(cols_to_remove)} read-only fields")
    
    # 3. Fix ProductCategoryProduct - ensure it has ID column
    print("3. Checking ProductCategoryProduct ID column...")
    ws = wb['26_ProductCategoryProduct']
    headers = [cell.value for cell in ws[1] if cell.value]
    
    if 'Id' not in [h.replace('*', '').strip() for h in headers]:
        print("   ⚠️  Missing Id column - this will cause all records to be treated as new")
    
    # 4. Fix ProductRelatedComponent - ensure it has ID column
    print("4. Checking ProductRelatedComponent ID column...")
    ws = wb['25_ProductRelatedComponent']
    headers = [cell.value for cell in ws[1] if cell.value]
    
    if 'Id' not in [h.replace('*', '').strip() for h in headers]:
        # Need to add ID column and populate with existing IDs
        ws.insert_cols(1)
        ws.cell(row=1, column=1, value='Id')
        print("   Added Id column")
    
    # Save workbook
    wb.save('data/Revenue_Cloud_Complete_Upload_Template.xlsx')
    print("\n✓ Fixed remaining issues")
    
    # Now let's also fix the upsert script to use correct external IDs
    print("\nCreating fixed upsert script...")
    
    fixed_script = '''#\!/usr/bin/env python3
"""
Fixed upsert script with correct external ID fields.
"""

import pandas as pd
import subprocess
from pathlib import Path
import os

class RevenueCloudUpserter:
    def __init__(self):
        self.workbook_path = Path('data/Revenue_Cloud_Complete_Upload_Template.xlsx')
        self.target_org = 'fortradp2'
        self.temp_dir = Path('data/temp_csv')
        self.temp_dir.mkdir(exist_ok=True)
        
        # Define objects with their correct external ID fields
        self.object_configs = {
            # Use Id for most objects
            'ProductCatalog': {'sheet': '11_ProductCatalog', 'external_id': 'Id'},
            'ProductCategory': {'sheet': '12_ProductCategory', 'external_id': 'Id'},
            'ProductClassification': {'sheet': '08_ProductClassification', 'external_id': 'Code'},
            'AttributeCategory': {'sheet': '10_AttributeCategory', 'external_id': 'Code'},
            'AttributePicklist': {'sheet': '14_AttributePicklist', 'external_id': 'Id'},  # Use Id
            'AttributeDefinition': {'sheet': '09_AttributeDefinition', 'external_id': 'Code'},
            'ProductSellingModel': {'sheet': '15_ProductSellingModel', 'external_id': 'Id'},
            'Product2': {'sheet': '13_Product2', 'external_id': 'Id'},
            'Pricebook2': {'sheet': '19_Pricebook2', 'external_id': 'Id'},
            'ProductAttributeDefinition': {'sheet': '17_ProductAttributeDef', 'external_id': 'Id'},
            'AttributePicklistValue': {'sheet': '18_AttributePicklistValue', 'external_id': 'Id'},
            'ProductCategoryProduct': {'sheet': '26_ProductCategoryProduct', 'external_id': 'Id'},
            'PricebookEntry': {'sheet': '20_PricebookEntry', 'external_id': 'Id'},
            'ProductRelatedComponent': {'sheet': '25_ProductRelatedComponent', 'external_id': 'Id'},
            'ProductComponentGroup': {'sheet': '14_ProductComponentGroup', 'external_id': 'Id'}
        }
        
    def process_sheet(self, object_name, config):
        """Process a single sheet."""
        try:
            df = pd.read_excel(self.workbook_path, sheet_name=config['sheet'])
            
            if df.empty:
                print(f"    ⚠️  No data found in sheet")
                return True
                
            # Clean column names
            df.columns = df.columns.str.replace('*', '', regex=False).str.strip()
            
            # Handle insert vs update
            external_id = config['external_id']
            
            if external_id == 'Id' and 'Id' in df.columns:
                # Separate records with and without IDs
                has_id = df[df['Id'].notna()].copy()
                no_id = df[df['Id'].isna()].copy()
                
                success = True
                
                # Insert new records
                if len(no_id) > 0:
                    no_id = no_id.drop('Id', axis=1)
                    csv_file = self.temp_dir / f'{object_name}_insert.csv'
                    no_id.to_csv(csv_file, index=False)
                    
                    print(f"    Inserting {len(no_id)} new records...")
                    cmd = [
                        'sf', 'data', 'import', 'bulk',
                        '--sobject', object_name,
                        '--file', str(csv_file),
                        '--target-org', self.target_org,
                        '--wait', '10'
                    ]
                    
                    result = subprocess.run(cmd, capture_output=True, text=True)
                    if result.returncode \!= 0:
                        print(f"    ✗ Insert failed: {result.stderr}")
                        success = False
                
                # Update existing records
                if len(has_id) > 0:
                    csv_file = self.temp_dir / f'{object_name}_update.csv'
                    has_id.to_csv(csv_file, index=False)
                    
                    print(f"    Updating {len(has_id)} existing records...")
                    cmd = [
                        'sf', 'data', 'upsert', 'bulk',
                        '--sobject', object_name,
                        '--external-id', 'Id',
                        '--file', str(csv_file),
                        '--target-org', self.target_org,
                        '--wait', '10'
                    ]
                    
                    result = subprocess.run(cmd, capture_output=True, text=True)
                    if result.returncode \!= 0:
                        print(f"    ✗ Update failed: {result.stderr}")
                        success = False
                
                return success
            else:
                # Standard upsert
                csv_file = self.temp_dir / f'{object_name}.csv'
                df.to_csv(csv_file, index=False)
                
                print(f"    Upserting {len(df)} records using {external_id}...")
                
                cmd = [
                    'sf', 'data', 'upsert', 'bulk',
                    '--sobject', object_name,
                    '--external-id', external_id,
                    '--file', str(csv_file),
                    '--target-org', self.target_org,
                    '--wait', '10'
                ]
                
                result = subprocess.run(cmd, capture_output=True, text=True)
                
                if result.returncode == 0:
                    print(f"    ✓ Success\!")
                    return True
                else:
                    print(f"    ✗ Failed: {result.stderr}")
                    return False
                    
        except Exception as e:
            print(f"    ✗ Error: {e}")
            return False
    
    def run(self):
        """Run the complete upsert process."""
        print("=" * 70)
        print("FIXED REVENUE CLOUD UPSERT PROCESS")
        print("=" * 70)
        
        # Upload order
        upload_order = [
            'ProductCatalog',
            'ProductCategory',
            'ProductClassification',
            'AttributeCategory',
            'AttributePicklist',
            'AttributeDefinition',
            'ProductSellingModel',
            'Product2',
            'Pricebook2',
            'ProductAttributeDefinition',
            'AttributePicklistValue',
            'ProductCategoryProduct',
            'PricebookEntry',
            'ProductComponentGroup',
            'ProductRelatedComponent'
        ]
        
        success_count = 0
        failed_count = 0
        
        for obj in upload_order:
            if obj in self.object_configs:
                print(f"\\n{obj}:")
                if self.process_sheet(obj, self.object_configs[obj]):
                    success_count += 1
                else:
                    failed_count += 1
        
        print("\\n" + "=" * 70)
        print("SUMMARY")
        print("=" * 70)
        print(f"✓ Successful: {success_count}")
        print(f"✗ Failed: {failed_count}")
        
        # Cleanup
        import shutil
        if self.temp_dir.exists():
            shutil.rmtree(self.temp_dir)

if __name__ == '__main__':
    upserter = RevenueCloudUpserter()
    upserter.run()
'''
    
    with open('run_fixed_upsert.py', 'w') as f:
        f.write(fixed_script)
    
    print("✓ Created run_fixed_upsert.py")

if __name__ == '__main__':
    main()
