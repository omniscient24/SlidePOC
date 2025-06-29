#!/usr/bin/env python3
"""
Fix incorrect ProductCategory IDs in ProductCategoryProduct sheet.
"""

from openpyxl import load_workbook

def main():
    print("Fixing ProductCategory IDs...")
    
    # Load workbook
    wb = load_workbook('data/Revenue_Cloud_Complete_Upload_Template.xlsx')
    ws = wb['26_ProductCategoryProduct']
    
    # Find the ProductCategoryId column
    category_col = None
    for col, cell in enumerate(ws[1], 1):
        if cell.value and 'ProductCategoryId' in str(cell.value):
            category_col = col
            break
    
    if not category_col:
        print("ERROR: Could not find ProductCategoryId column")
        return
    
    # Fix the incorrect ID
    # 0ZSdp00000007kfGAA -> should be a bundle category
    # Based on the pattern, bundles should go in "Bundle Components" category: 0ZGdp00000007JIGAY
    
    fixed_count = 0
    for row in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=category_col).value
        if cell_value == '0ZSdp00000007kfGAA':
            ws.cell(row=row, column=category_col).value = '0ZGdp00000007JFGAY'  # Core Offerings
            fixed_count += 1
            print(f"Fixed row {row}: 0ZSdp00000007kfGAA -> 0ZGdp00000007JFGAY (Core Offerings)")
    
    print(f"\nFixed {fixed_count} records")
    
    # Save workbook
    wb.save('data/Revenue_Cloud_Complete_Upload_Template.xlsx')
    print("✓ Updated workbook")

if __name__ == '__main__':
    main()