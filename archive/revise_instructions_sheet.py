#\!/usr/bin/env python3
"""
Revise the instructions sheet with clearer descriptions of object purposes and relationships.
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def main():
    print("=" * 60)
    print("REVISING INSTRUCTIONS SHEET")
    print("=" * 60)
    
    # Load workbook
    wb = load_workbook('data/Revenue_Cloud_Complete_Upload_Template.xlsx')
    ws = wb['Instructions']
    
    # Clear existing content
    for row in ws.iter_rows():
        for cell in row:
            cell.value = None
    
    # Define column headers
    headers = ['Step', 'Object', 'Purpose in Revenue Cloud', 'Key Relationships', 'Upload Order Dependency']
    
    # Set headers with formatting
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Define object information with clear purposes and relationships
    objects = [
        {
            'step': 1,
            'object': 'Setup',
            'purpose': 'Initial configuration and preparation step',
            'relationships': 'Foundation for all other objects',
            'dependency': 'Must be completed first'
        },
        {
            'step': 2,
            'object': 'ProductClassification',
            'purpose': 'Categorizes products into hierarchical classifications for organizing and grouping similar products',
            'relationships': 'Links to: Product2 (via ProductClassificationProduct), AttributeDefinition (classification-specific attributes)',
            'dependency': 'Required before products that use classifications'
        },
        {
            'step': 3,
            'object': 'AttributeCategory',
            'purpose': 'Groups related attributes together for better organization and management of product characteristics',
            'relationships': 'Links to: AttributeDefinition (groups attributes), ProductAttributeDefinition (assigns grouped attributes to products)',
            'dependency': 'Required before AttributeDefinition'
        },
        {
            'step': 4,
            'object': 'AttributePicklist',
            'purpose': 'Defines picklist data sources for attributes that have predefined selectable values',
            'relationships': 'Links to: AttributePicklistValue (contains the actual picklist values), AttributeDefinition (uses picklist as data source)',
            'dependency': 'Required before picklist-type AttributeDefinitions'
        },
        {
            'step': 5,
            'object': 'AttributePicklistValue',
            'purpose': 'Contains the actual selectable values for picklist attributes (e.g., colors, sizes, options)',
            'relationships': 'Links to: AttributePicklist (parent picklist), Used by: AttributeDefinition of picklist type',
            'dependency': 'Required after AttributePicklist, before using in products'
        },
        {
            'step': 6,
            'object': 'AttributeDefinition',
            'purpose': 'Defines reusable product characteristics and configuration options that can be applied to multiple products',
            'relationships': 'Links to: AttributeCategory (for grouping), AttributePicklist (for picklist types), ProductAttributeDefinition (assigns to products)',
            'dependency': 'Required before ProductAttributeDefinition'
        },
        {
            'step': 7,
            'object': 'ProductCatalog',
            'purpose': 'Top-level container that organizes products and categories for different markets, channels, or business units',
            'relationships': 'Links to: ProductCategory (contains categories), Product2 (contains products via ProductCatalogProduct)',
            'dependency': 'Required before ProductCategory'
        },
        {
            'step': 8,
            'object': 'ProductCategory',
            'purpose': 'Organizes products into browsable categories within catalogs (e.g., Hardware, Software, Services)',
            'relationships': 'Links to: ProductCatalog (parent catalog), Product2 (via ProductCategoryProduct), can have parent/child category relationships',
            'dependency': 'Required before ProductCategoryProduct'
        },
        {
            'step': 9,
            'object': 'ProductSellingModel',
            'purpose': 'Defines how products are sold and priced over time (One-time, Subscription-based, Usage-based, etc.)',
            'relationships': 'Links to: Product2 (via ProductSellingModelOption), PricebookEntry (pricing per model), defines billing and revenue recognition patterns',
            'dependency': 'Required before ProductSellingModelOption'
        },
        {
            'step': 10,
            'object': 'Product2',
            'purpose': 'Core product definition containing all product details, SKUs, and configuration options',
            'relationships': 'Central object linking to: ProductCategory, ProductSellingModel, ProductAttributeDefinition, PricebookEntry, ProductRelatedComponent (for bundles)',
            'dependency': 'Required before most product-related configurations'
        },
        {
            'step': 11,
            'object': 'ProductSellingModelOption',
            'purpose': 'Links products to their available selling models, enabling products to be sold in multiple ways',
            'relationships': 'Links: Product2 to ProductSellingModel, Required for: PricebookEntry creation in Revenue Cloud',
            'dependency': 'Required after Product2 and ProductSellingModel'
        },
        {
            'step': 12,
            'object': 'ProductAttributeDefinition',
            'purpose': 'Assigns specific attributes to products and defines how they behave for that product (default values, constraints)',
            'relationships': 'Links: Product2 to AttributeDefinition, Can reference: AttributeCategory for grouping on product',
            'dependency': 'Required after Product2 and AttributeDefinition'
        },
        {
            'step': 13,
            'object': 'Pricebook2',
            'purpose': 'Container for product prices, supporting multiple price lists for different markets, currencies, or customer segments',
            'relationships': 'Links to: PricebookEntry (contains actual prices), can have standard and custom pricebooks',
            'dependency': 'Required before PricebookEntry'
        },
        {
            'step': 14,
            'object': 'PricebookEntry',
            'purpose': 'Defines the actual price for a product within a specific pricebook and selling model combination',
            'relationships': 'Links: Product2 + Pricebook2 + ProductSellingModel, Requires: ProductSellingModelOption to exist first',
            'dependency': 'Required after Pricebook2 and ProductSellingModelOption'
        },
        {
            'step': 15,
            'object': 'ProductCategoryProduct',
            'purpose': 'Junction object that assigns products to categories, allowing products to appear in multiple categories',
            'relationships': 'Links: Product2 to ProductCategory, enables product browsing and organization',
            'dependency': 'Required after Product2 and ProductCategory'
        },
        {
            'step': 16,
            'object': 'ProductComponentGroup',
            'purpose': 'Groups bundle components into logical sections for configurable products (e.g., "Core Components", "Optional Add-ons")',
            'relationships': 'Links to: Product2 (parent bundle), ProductRelatedComponent (groups components), required for configurable bundles',
            'dependency': 'Required before ProductRelatedComponent for configurable bundles'
        },
        {
            'step': 17,
            'object': 'ProductRelatedComponent',
            'purpose': 'Defines parent-child relationships between products for bundles, kits, and product options',
            'relationships': 'Links: Parent Product2 to Child Product2, Uses: ProductComponentGroup (for configurable bundles), ProductRelationshipType',
            'dependency': 'Last step - requires all products and groups to exist'
        }
    ]
    
    # Write data rows
    row_num = 2
    for obj in objects:
        ws.cell(row=row_num, column=1, value=obj['step'])
        ws.cell(row=row_num, column=2, value=obj['object'])
        ws.cell(row=row_num, column=3, value=obj['purpose'])
        ws.cell(row=row_num, column=4, value=obj['relationships'])
        ws.cell(row=row_num, column=5, value=obj['dependency'])
        row_num += 1
    
    # Format data cells
    data_alignment = Alignment(vertical="top", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in range(2, row_num):
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            cell.alignment = data_alignment
            cell.border = border
    
    # Adjust column widths
    column_widths = [8, 25, 60, 60, 35]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Add summary section at the bottom
    summary_row = row_num + 2
    ws.cell(row=summary_row, column=1, value="KEY CONCEPTS:")
    ws.cell(row=summary_row, column=1).font = Font(bold=True)
    
    concepts = [
        "• Upload Order: Follow the step sequence to ensure all dependencies are met",
        "• Relationships: Most objects reference other objects via ID fields (e.g., Product2Id, AttributeDefinitionId)",
        "• Revenue Cloud Specific: ProductSellingModel and ProductSellingModelOption are required for Revenue Cloud pricing",
        "• Bundles: Configurable bundles require ProductComponentGroup; static bundles can use ProductRelatedComponent directly",
        "• Attributes: Create reusable attributes once, then assign to multiple products with specific configurations"
    ]
    
    for i, concept in enumerate(concepts):
        ws.cell(row=summary_row + i + 1, column=2, value=concept)
        ws.merge_cells(f'B{summary_row + i + 1}:E{summary_row + i + 1}')
    
    # Save workbook
    wb.save('data/Revenue_Cloud_Complete_Upload_Template.xlsx')
    
    print("✓ Instructions sheet revised with clear purposes and relationships")
    print(f"✓ Added {len(objects)} object descriptions")
    print("✓ Included upload order dependencies")
    print("✓ Added key concepts section")

if __name__ == '__main__':
    main()
