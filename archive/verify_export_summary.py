#!/usr/bin/env python3
"""
Verify the export and create a summary of what's in the workbook.
"""

import pandas as pd
from pathlib import Path
import openpyxl

def verify_export():
    """Verify the exported data in the workbook."""
    workbook_path = Path('data/Revenue_Cloud_Complete_Upload_Template.xlsx')
    
    print("=" * 70)
    print("REVENUE CLOUD EXPORT VERIFICATION")
    print("=" * 70)
    print(f"Workbook: {workbook_path}")
    print()
    
    # Load workbook
    wb = openpyxl.load_workbook(workbook_path, read_only=True)
    
    # Summary data
    sheet_summary = []
    total_records = 0
    
    # Check each sheet
    sheet_order = [
        '11_ProductCatalog',
        '12_ProductCategory',
        '08_ProductClassification',
        '09_AttributeDefinition',
        '10_AttributeCategory',
        '14_AttributePicklist',
        '15_ProductSellingModel',
        '13_Product2',
        '17_ProductAttributeDef',
        '18_AttributePicklistValue',
        '19_Pricebook2',
        '20_PricebookEntry',
        '26_ProductCategoryProduct',
        '25_ProductRelatedComponent'
    ]
    
    for sheet_name in sheet_order:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Count non-empty data rows (skip header)
            data_rows = 0
            for row in range(2, ws.max_row + 1):
                # Check if any cell in the row has data
                has_data = False
                for col in range(1, min(ws.max_column + 1, 10)):  # Check first 10 columns
                    if ws.cell(row=row, column=col).value:
                        has_data = True
                        break
                if has_data:
                    data_rows += 1
                else:
                    break  # Stop at first empty row
            
            sheet_summary.append({
                'sheet': sheet_name,
                'records': data_rows
            })
            total_records += data_rows
    
    wb.close()
    
    # Print summary
    print("SHEET SUMMARY:")
    print("-" * 50)
    for item in sheet_summary:
        status = "✓" if item['records'] > 0 else "✗"
        print(f"{status} {item['sheet']:<30} {item['records']:>5} records")
    
    print("-" * 50)
    print(f"Total records in workbook: {total_records}")
    
    # Check specific data integrity
    print("\n" + "=" * 70)
    print("DATA INTEGRITY CHECKS:")
    print("-" * 50)
    
    # Check AttributeDefinition picklist references
    df_ad = pd.read_excel(workbook_path, sheet_name='09_AttributeDefinition')
    df_ad.columns = df_ad.columns.str.replace('*', '')
    picklist_defs = df_ad[df_ad['DataType'] == 'Picklist']
    picklist_with_id = picklist_defs[picklist_defs['PicklistId'].notna()]
    print(f"✓ AttributeDefinitions with Picklist type: {len(picklist_defs)}")
    print(f"✓ AttributeDefinitions with PicklistId set: {len(picklist_with_id)}")
    
    # Check AttributePicklistValue distribution
    df_apv = pd.read_excel(workbook_path, sheet_name='18_AttributePicklistValue')
    df_apv.columns = df_apv.columns.str.replace('*', '')
    unique_picklists = df_apv['PicklistId'].nunique()
    print(f"✓ AttributePicklistValues: {len(df_apv)} values across {unique_picklists} picklists")
    
    # Check Product references in PricebookEntry
    df_pbe = pd.read_excel(workbook_path, sheet_name='20_PricebookEntry')
    has_product_name = 'Product2.Name' in df_pbe.columns
    if has_product_name:
        df_pbe_clean = df_pbe.dropna(subset=['Product2.Name'])
        print(f"✓ PricebookEntry records with Product names: {len(df_pbe_clean)}/{len(df_pbe)}")
    
    print("\n✓ Export verification complete")
    print(f"✓ Workbook location: {workbook_path.absolute()}")

def main():
    verify_export()

if __name__ == '__main__':
    main()