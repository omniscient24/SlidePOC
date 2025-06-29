#\!/usr/bin/env python3
"""
Create ProductRelatedComponent records for bundle products.
"""

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook

def main():
    print("=" * 60)
    print("CREATING PRODUCTRELATEDCOMPONENT RECORDS")
    print("=" * 60)
    
    # Define bundle-component relationships
    # Based on typical bundle structures:
    # - Essentials: Basic components
    # - Advanced: Essentials + more components
    # - Elite: All components
    
    bundle_configs = {
        'DCS Essentials': {
            'id': '01tdp000006JEGlAAO',  # V2 bundle
            'components': [
                {'id': '01tdp000006HfpoAAC', 'name': 'DCS for Windows', 'required': True, 'qty': 1},
                {'id': '01tdp000006HfpkAAC', 'name': 'Data Detection Engine', 'required': True, 'qty': 1},
                {'id': '01tdp000006HfprAAC', 'name': 'DCS Getting Started Package', 'required': True, 'qty': 1}
            ]
        },
        'DCS Advanced': {
            'id': '01tdp000006JEGjAAO',  # V2 bundle
            'components': [
                {'id': '01tdp000006HfpoAAC', 'name': 'DCS for Windows', 'required': True, 'qty': 1},
                {'id': '01tdp000006HfpkAAC', 'name': 'Data Detection Engine', 'required': True, 'qty': 1},
                {'id': '01tdp000006HfpnAAC', 'name': 'DCS Admin Console', 'required': True, 'qty': 1},
                {'id': '01tdp000006HfpmAAC', 'name': 'DCS Analysis Collector', 'required': True, 'qty': 1},
                {'id': '01tdp000006HfppAAC', 'name': 'DCS for OWA', 'required': False, 'qty': 1}
            ]
        },
        'DCS Elite': {
            'id': '01tdp000006JEGkAAO',  # V2 bundle
            'components': [
                {'id': '01tdp000006HfpoAAC', 'name': 'DCS for Windows', 'required': True, 'qty': 1},
                {'id': '01tdp000006HfpkAAC', 'name': 'Data Detection Engine', 'required': True, 'qty': 1},
                {'id': '01tdp000006HfpnAAC', 'name': 'DCS Admin Console', 'required': True, 'qty': 1},
                {'id': '01tdp000006HfpmAAC', 'name': 'DCS Analysis Collector', 'required': True, 'qty': 1},
                {'id': '01tdp000006HfppAAC', 'name': 'DCS for OWA', 'required': True, 'qty': 1},
                {'id': '01tdp000006HfplAAC', 'name': 'Unlimited Classification', 'required': True, 'qty': 1},
                {'id': '01tdp000006HfpqAAC', 'name': 'Software Development Kit', 'required': False, 'qty': 1}
            ]
        }
    }
    
    # Create records
    records = []
    sequence = 10
    
    for bundle_name, bundle_config in bundle_configs.items():
        print(f"\n{bundle_name}:")
        bundle_id = bundle_config['id']
        
        for component in bundle_config['components']:
            record = {
                'Name': f"{bundle_name} - {component['name']}",
                'ParentProductId': bundle_id,
                'ChildProductId': component['id'],
                'Quantity': component['qty'],
                'MinQuantity': 1 if component['required'] else 0,
                'MaxQuantity': 999,  # Unlimited
                'IsComponentRequired': component['required'],
                'Sequence': sequence,
                'DoesBundlePriceIncludeChild': True  # Bundle pricing includes components
            }
            records.append(record)
            
            status = "Required" if component['required'] else "Optional"
            print(f"  - {component['name']} ({status})")
            
            sequence += 10
    
    print(f"\nTotal ProductRelatedComponent records to create: {len(records)}")
    
    # Load workbook and update sheet
    wb = load_workbook('data/Revenue_Cloud_Complete_Upload_Template.xlsx')
    ws = wb['25_ProductRelatedComponent']
    
    # Get column headers
    headers = []
    header_map = {}
    for col, cell in enumerate(ws[1], 1):
        if cell.value:
            header_name = cell.value.replace('*', '').strip()
            headers.append(header_name)
            header_map[header_name] = col
    
    print(f"\nSheet headers: {headers}")
    
    # Clear existing data (if any)
    for row in range(2, ws.max_row + 1):
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).value = None
    
    # Write new data
    row_num = 2
    for record in records:
        for field, value in record.items():
            if field in header_map:
                ws.cell(row=row_num, column=header_map[field]).value = value
        row_num += 1
    
    # Save workbook
    wb.save('data/Revenue_Cloud_Complete_Upload_Template.xlsx')
    print(f"\n✓ Updated workbook with {len(records)} ProductRelatedComponent records")
    
    # Create CSV for import
    df = pd.DataFrame(records)
    csv_file = Path('data/product_related_components.csv')
    
    # Only include fields that exist in Salesforce
    import_fields = ['Name', 'ParentProductId', 'ChildProductId', 'Quantity', 
                     'MinQuantity', 'MaxQuantity', 'IsComponentRequired', 
                     'Sequence', 'DoesBundlePriceIncludeChild']
    
    import_df = df[import_fields]
    import_df.to_csv(csv_file, index=False)
    
    print("\n✓ Created import file: data/product_related_components.csv")
    print("\nReady to import ProductRelatedComponent records")

if __name__ == '__main__':
    main()
