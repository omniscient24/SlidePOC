#!/usr/bin/env python3
"""
Verify and fix AttributeDefinition sheet with proper PicklistId values.
"""

import pandas as pd
import openpyxl
from pathlib import Path
import subprocess
import json

class AttributeDefinitionFixer:
    def __init__(self):
        self.workbook_path = Path('data/Revenue_Cloud_Complete_Upload_Template.xlsx')
        self.target_org = 'fortradp2'
        
    def verify_current_state(self):
        """Check current state of AttributeDefinition sheet."""
        print("=" * 60)
        print("VERIFYING ATTRIBUTEDEFINITION SHEET")
        print("=" * 60)
        
        # Read the sheet
        df = pd.read_excel(self.workbook_path, sheet_name='09_AttributeDefinition')
        df.columns = df.columns.str.replace('*', '', regex=False)
        
        # Show all records
        print(f"\nTotal AttributeDefinition records: {len(df)}")
        
        # Check picklist records
        picklist_records = df[df['DataType'] == 'Picklist']
        print(f"\nPicklist-type AttributeDefinitions: {len(picklist_records)}")
        
        for idx, row in picklist_records.iterrows():
            print(f"\n{row['Name']}:")
            print(f"  Code: {row['Code']}")
            print(f"  PicklistId: {row.get('PicklistId', 'NOT SET')}")
            
        # Check if we have the right AttributePicklist IDs
        self.query_attribute_picklists()
        
    def query_attribute_picklists(self):
        """Query current AttributePicklist records."""
        print("\n" + "=" * 60)
        print("QUERYING ATTRIBUTEPICKLIST RECORDS")
        print("=" * 60)
        
        query = "SELECT Id, Name FROM AttributePicklist ORDER BY Name"
        cmd = [
            'sf', 'data', 'query',
            '--query', query,
            '--target-org', self.target_org,
            '--json'
        ]
        
        result = subprocess.run(cmd, capture_output=True, text=True)
        
        if result.returncode == 0:
            data = json.loads(result.stdout)
            if 'result' in data and 'records' in data['result']:
                records = data['result']['records']
                
                print(f"\nFound {len(records)} AttributePicklist records:")
                self.picklist_map = {}
                for rec in records:
                    print(f"  {rec['Name']}: {rec['Id']}")
                    self.picklist_map[rec['Name']] = rec['Id']
                
                # Now update the sheet
                self.update_sheet_with_correct_ids()
        else:
            print(f"Failed to query: {result.stderr}")
            
    def update_sheet_with_correct_ids(self):
        """Update AttributeDefinition sheet with correct PicklistId values."""
        print("\n" + "=" * 60)
        print("UPDATING ATTRIBUTEDEFINITION WITH CORRECT PICKLISTIDS")
        print("=" * 60)
        
        # Load workbook
        wb = openpyxl.load_workbook(self.workbook_path)
        ws = wb['09_AttributeDefinition']
        
        # Find column indices
        headers = [cell.value for cell in ws[1]]
        picklist_id_col = None
        code_col = None
        datatype_col = None
        name_col = None
        
        for idx, header in enumerate(headers, 1):
            if header:
                if 'PicklistId' in str(header):
                    picklist_id_col = idx
                elif header == 'Code':
                    code_col = idx
                elif 'DataType' in str(header):
                    datatype_col = idx
                elif 'Name' in str(header):
                    name_col = idx
        
        print(f"\nColumn indices found:")
        print(f"  Name: {name_col}")
        print(f"  Code: {code_col}")
        print(f"  DataType: {datatype_col}")
        print(f"  PicklistId: {picklist_id_col}")
        
        if all([picklist_id_col, code_col, datatype_col, name_col]):
            # Map AttributeDefinition codes to picklist names
            code_to_picklist = {
                'PT': 'Pricing Tier Options',
                'LT': 'License Type Options',
                'UT': 'Unit Type Options',
                'ST ': 'Server Type Options',  # Note the space
                'TA': 'Agent Type Options',
                'KT': 'Key Type Options',
                'SLT': 'Server Location Type Options',
                'PG': 'PGroup Options',
                'OS': 'Operating System Options'
            }
            
            updates = 0
            print("\nUpdating records:")
            
            for row in range(2, ws.max_row + 1):
                datatype_cell = ws.cell(row=row, column=datatype_col)
                if datatype_cell.value == 'Picklist':
                    name_cell = ws.cell(row=row, column=name_col)
                    code_cell = ws.cell(row=row, column=code_col)
                    picklist_cell = ws.cell(row=row, column=picklist_id_col)
                    
                    code = code_cell.value
                    if code in code_to_picklist:
                        picklist_name = code_to_picklist[code]
                        if picklist_name in self.picklist_map:
                            new_id = self.picklist_map[picklist_name]
                            old_value = picklist_cell.value
                            picklist_cell.value = new_id
                            updates += 1
                            print(f"  {name_cell.value} ({code}): {old_value} -> {new_id}")
                        else:
                            print(f"  WARNING: No picklist found for {picklist_name}")
                    else:
                        print(f"  WARNING: No mapping for code {code}")
            
            wb.save(self.workbook_path)
            print(f"\n✓ Updated {updates} AttributeDefinition records")
            
            # Now try importing just AttributeDefinition
            self.import_attribute_definitions()
        else:
            print("\n✗ Could not find all required columns")
            
    def import_attribute_definitions(self):
        """Import only AttributeDefinition records."""
        print("\n" + "=" * 60)
        print("IMPORTING ATTRIBUTEDEFINITION RECORDS")
        print("=" * 60)
        
        # Read and prepare data
        df = pd.read_excel(self.workbook_path, sheet_name='09_AttributeDefinition')
        df.columns = df.columns.str.replace('*', '', regex=False)
        
        # Remove empty rows
        df = df[df['Name'].notna()]
        
        print(f"\nPreparing to upsert {len(df)} AttributeDefinition records")
        
        # Save to CSV
        csv_file = Path('data/temp_attribute_definition.csv')
        df.to_csv(csv_file, index=False)
        
        # Upsert
        cmd = [
            'sf', 'data', 'upsert', 'bulk',
            '--sobject', 'AttributeDefinition',
            '--file', str(csv_file),
            '--external-id', 'Code',
            '--target-org', self.target_org,
            '--wait', '10'
        ]
        
        result = subprocess.run(cmd, capture_output=True, text=True)
        
        if result.returncode == 0:
            print("✓ Successfully imported AttributeDefinition records")
        else:
            print(f"✗ Failed: {result.stderr}")
            
            # Get detailed errors
            if 'job-id' in result.stderr:
                import re
                match = re.search(r'--job-id (\w+)', result.stderr)
                if match:
                    job_id = match.group(1)
                    self.get_error_details(job_id)
        
        # Clean up
        csv_file.unlink()
        
    def get_error_details(self, job_id):
        """Get detailed error information."""
        print(f"\nGetting error details for job {job_id}...")
        
        cmd = [
            'sf', 'data', 'bulk', 'results',
            '--job-id', job_id,
            '--target-org', self.target_org
        ]
        
        subprocess.run(cmd)
        
        # Read failed records
        failed_file = Path(f"{job_id}-failed-records.csv")
        if failed_file.exists():
            df = pd.read_csv(failed_file)
            print(f"\nFirst 5 failures:")
            for idx, row in df.head().iterrows():
                print(f"\n{row.get('Name', 'Unknown')}:")
                print(f"  Error: {row.get('sf__Error', 'Unknown')}")
            
            # Clean up
            failed_file.unlink()
            success_file = Path(f"{job_id}-success-records.csv")
            if success_file.exists():
                success_file.unlink()

def main():
    fixer = AttributeDefinitionFixer()
    fixer.verify_current_state()

if __name__ == '__main__':
    main()