#!/usr/bin/env python3
"""
Remove system-managed fields from ProductCategory sheet.
"""

from openpyxl import load_workbook
import subprocess
import json

def get_system_fields():
    """Get list of system-managed fields for ProductCategory."""
    cmd = ['sf', 'sobject', 'describe', '--sobject', 'ProductCategory', '--target-org', 'fortradp2', '--json']
    result = subprocess.run(cmd, capture_output=True, text=True)
    
    system_fields = []
    if result.returncode == 0:
        data = json.loads(result.stdout)
        if 'result' in data:
            for field in data['result']['fields']:
                # System-managed fields are neither createable nor updateable
                if not field.get('createable', True) and not field.get('updateable', True):
                    system_fields.append(field['name'])
    
    return system_fields

def main():
    print("Removing system-managed fields from ProductCategory...")
    
    # Get system fields
    system_fields = get_system_fields()
    print(f"Found {len(system_fields)} system-managed fields: {system_fields}")
    
    # Load workbook
    wb = load_workbook('data/Revenue_Cloud_Complete_Upload_Template.xlsx')
    ws = wb['12_ProductCategory']
    
    # Find columns to remove
    cols_to_remove = []
    for col, cell in enumerate(ws[1], 1):
        if cell.value:
            field_name = str(cell.value).replace('*', '').strip()
            if field_name in system_fields:
                cols_to_remove.append((col, field_name))
    
    print(f"\nColumns to remove: {[name for _, name in cols_to_remove]}")
    
    # Remove columns (in reverse order to maintain indices)
    removed = []
    for col, name in sorted(cols_to_remove, reverse=True):
        ws.delete_cols(col)
        removed.append(name)
    
    print(f"Removed {len(removed)} columns: {removed}")
    
    # Save workbook
    wb.save('data/Revenue_Cloud_Complete_Upload_Template.xlsx')
    print("\n✓ Updated workbook")

if __name__ == '__main__':
    main()