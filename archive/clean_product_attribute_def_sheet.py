#\!/usr/bin/env python3
"""
Clean ProductAttributeDef sheet by removing invalid columns.
"""

import pandas as pd
from openpyxl import load_workbook

def main():
    print("=" * 60)
    print("CLEANING PRODUCTATTRIBUTEDEF SHEET")
    print("=" * 60)
    
    # Valid fields for ProductAttributeDefinition (creatable/updateable)
    valid_fields = [
        'Id',  # For upserts
        'Name',
        'Product2Id',
        'AttributeDefinitionId',
        'AttributeCategoryId',
        'Sequence',
        'IsRequired',
        'IsHidden',
        'IsReadOnly',
        'IsPriceImpacting',
        'DefaultValue',
        'HelpText',
        'MinimumValue',
        'MaximumValue',
        'MaximumCharacterCount',
        'MinimumCharacterCount',
        'DisplayType',
        'Status',
        'AttributeNameOverride',
        'Description',
        'StepValue',
        'ValueDescription',
        'OverriddenProductAttributeDefinitionId',
        'OverrideContextId',
        'ProductClassificationAttributeId',
        'OwnerId'
    ]
    
    # Load workbook
    wb = load_workbook('data/Revenue_Cloud_Complete_Upload_Template.xlsx')
    ws = wb['17_ProductAttributeDef']
    
    # Get current headers and their positions
    headers = []
    header_positions = {}
    for col, cell in enumerate(ws[1], 1):
        if cell.value:
            header_name = cell.value.replace('*', '').strip()
            headers.append(header_name)
            header_positions[header_name] = col
    
    print(f"Current headers: {len(headers)}")
    
    # Identify invalid columns (relationship fields and non-existent fields)
    invalid_fields = []
    for header in headers:
        # Remove relationship fields (contain dots)
        if '.' in header:
            invalid_fields.append(header)
        # Check if field is not in valid list
        elif header not in valid_fields:
            invalid_fields.append(header)
    
    print(f"\nRemoving {len(invalid_fields)} invalid columns:")
    for field in invalid_fields:
        print(f"  - {field}")
    
    # Delete columns in reverse order
    cols_to_delete = []
    for field in invalid_fields:
        if field in header_positions:
            cols_to_delete.append(header_positions[field])
    
    cols_to_delete.sort(reverse=True)
    
    # Delete columns
    for col_idx in cols_to_delete:
        ws.delete_cols(col_idx)
    
    # Update required field markers
    required_fields = ['Product2Id', 'AttributeDefinitionId']
    
    # Re-read headers after deletion and update required markers
    for col, cell in enumerate(ws[1], 1):
        if cell.value:
            header_name = cell.value.replace('*', '').strip()
            if header_name in required_fields and '*' not in cell.value:
                cell.value = f"{header_name}*"
    
    # Save workbook
    wb.save('data/Revenue_Cloud_Complete_Upload_Template.xlsx')
    
    print("\n✓ Sheet cleaned successfully")
    
    # Verify the changes
    print("\nRemaining valid columns:")
    remaining_headers = []
    for cell in ws[1]:
        if cell.value:
            remaining_headers.append(cell.value)
    
    for i, header in enumerate(remaining_headers, 1):
        print(f"  {i}. {header}")
    
    print(f"\nTotal valid columns: {len(remaining_headers)}")

if __name__ == '__main__':
    main()
