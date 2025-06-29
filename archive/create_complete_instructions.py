#\!/usr/bin/env python3
"""
Create complete instructions sheet with all objects.
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def main():
    print("=" * 60)
    print("CREATING COMPLETE INSTRUCTIONS SHEET")
    print("=" * 60)
    
    # Load workbook
    wb = load_workbook('data/Revenue_Cloud_Complete_Upload_Template.xlsx')
    
    # Remove old instructions sheet
    if 'Instructions' in wb.sheetnames:
        wb.remove(wb['Instructions'])
    
    # Create new instructions sheet
    ws = wb.create_sheet('Instructions', 0)  # Insert at beginning
    
    # Define column headers
    headers = ['Step', 'Object', 'Purpose in Revenue Cloud', 'Key Relationships', 'Upload Order Dependency']
    
    # Set headers with formatting
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Define ALL objects in proper order
    objects = [
        # Core Setup
        {
            'step': 1, 'object': 'Setup',
            'purpose': 'Initial configuration and preparation step',
            'relationships': 'Foundation for all other objects',
            'dependency': 'Must be completed first'
        },
        
        # Financial and Compliance Objects (Optional)
        {
            'step': 2, 'object': 'CostBook',
            'purpose': 'Defines cost structures for products to track profitability and margins',
            'relationships': 'Links to: CostBookEntry (contains actual costs per product)',
            'dependency': 'Optional - Required only if tracking product costs'
        },
        {
            'step': 3, 'object': 'LegalEntity',
            'purpose': 'Represents legal business entities for financial reporting and compliance',
            'relationships': 'Used in: Order and billing transactions for legal compliance',
            'dependency': 'Optional - Required for multi-entity organizations'
        },
        {
            'step': 4, 'object': 'TaxEngine',
            'purpose': 'Defines the tax calculation engine to be used (e.g., Avalara, Vertex)',
            'relationships': 'Links to: TaxPolicy (uses engine for calculations)',
            'dependency': 'Optional - Required only if using tax calculations'
        },
        {
            'step': 5, 'object': 'TaxPolicy',
            'purpose': 'Defines tax calculation rules and policies for different jurisdictions',
            'relationships': 'Links to: TaxEngine (calculation method), TaxTreatment (specific tax rules)',
            'dependency': 'Optional - After TaxEngine if using tax features'
        },
        {
            'step': 6, 'object': 'TaxTreatment',
            'purpose': 'Specifies how specific products or transactions are treated for tax purposes',
            'relationships': 'Links to: TaxPolicy (applies treatment rules), Product2 (product-specific tax treatment)',
            'dependency': 'Optional - After TaxPolicy if using tax features'
        },
        {
            'step': 7, 'object': 'BillingPolicy',
            'purpose': 'Defines billing rules, cycles, and payment terms for customer accounts',
            'relationships': 'Links to: BillingTreatment (specific billing rules), Order (applies billing rules)',
            'dependency': 'Optional - Required for billing automation'
        },
        {
            'step': 8, 'object': 'BillingTreatment',
            'purpose': 'Specifies detailed billing treatment rules for specific scenarios or products',
            'relationships': 'Links to: BillingPolicy (extends policy rules), Product2 (product-specific billing)',
            'dependency': 'Optional - After BillingPolicy if using billing features'
        },
        
        # Product Foundation Objects
        {
            'step': 9, 'object': 'ProductClassification',
            'purpose': 'Categorizes products into hierarchical classifications for organizing and grouping similar products',
            'relationships': 'Links to: Product2 (via ProductClassificationProduct), AttributeDefinition (classification-specific attributes)',
            'dependency': 'Required before products that use classifications'
        },
        {
            'step': 10, 'object': 'AttributeCategory',
            'purpose': 'Groups related attributes together for better organization and management of product characteristics',
            'relationships': 'Links to: AttributeDefinition (groups attributes), ProductAttributeDefinition (assigns grouped attributes to products)',
            'dependency': 'Required before AttributeDefinition'
        },
        {
            'step': 11, 'object': 'AttributePicklist',
            'purpose': 'Defines picklist data sources for attributes that have predefined selectable values',
            'relationships': 'Links to: AttributePicklistValue (contains the actual picklist values), AttributeDefinition (uses picklist as data source)',
            'dependency': 'Required before picklist-type AttributeDefinitions'
        },
        {
            'step': 12, 'object': 'AttributePicklistValue',
            'purpose': 'Contains the actual selectable values for picklist attributes (e.g., colors, sizes, options)',
            'relationships': 'Links to: AttributePicklist (parent picklist), Used by: AttributeDefinition of picklist type',
            'dependency': 'Required after AttributePicklist, before using in products'
        },
        {
            'step': 13, 'object': 'AttributeDefinition',
            'purpose': 'Defines reusable product characteristics and configuration options that can be applied to multiple products',
            'relationships': 'Links to: AttributeCategory (for grouping), AttributePicklist (for picklist types), ProductAttributeDefinition (assigns to products)',
            'dependency': 'Required before ProductAttributeDefinition'
        },
        
        # Catalog Structure
        {
            'step': 14, 'object': 'ProductCatalog',
            'purpose': 'Top-level container that organizes products and categories for different markets, channels, or business units',
            'relationships': 'Links to: ProductCategory (contains categories), Product2 (contains products via ProductCatalogProduct)',
            'dependency': 'Required before ProductCategory'
        },
        {
            'step': 15, 'object': 'ProductCategory',
            'purpose': 'Organizes products into browsable categories within catalogs (e.g., Hardware, Software, Services)',
            'relationships': 'Links to: ProductCatalog (parent catalog), Product2 (via ProductCategoryProduct), can have parent/child category relationships',
            'dependency': 'Required before ProductCategoryProduct'
        },
        
        # Selling Models and Products
        {
            'step': 16, 'object': 'ProductSellingModel',
            'purpose': 'Defines how products are sold and priced over time (One-time, Subscription-based, Usage-based, etc.)',
            'relationships': 'Links to: Product2 (via ProductSellingModelOption), PricebookEntry (pricing per model), defines billing and revenue recognition patterns',
            'dependency': 'Required before ProductSellingModelOption'
        },
        {
            'step': 17, 'object': 'Product2',
            'purpose': 'Core product definition containing all product details, SKUs, and configuration options',
            'relationships': 'Central object linking to: ProductCategory, ProductSellingModel, ProductAttributeDefinition, PricebookEntry, ProductRelatedComponent (for bundles)',
            'dependency': 'Required before most product-related configurations'
        },
        {
            'step': 18, 'object': 'ProductSellingModelOption',
            'purpose': 'Links products to their available selling models, enabling products to be sold in multiple ways',
            'relationships': 'Junction between: Product2 and ProductSellingModel, Required for: Revenue Cloud PricebookEntry',
            'dependency': 'Required after Product2 and ProductSellingModel, before PricebookEntry'
        },
        
        # Cost Tracking (Optional)
        {
            'step': 19, 'object': 'CostBookEntry',
            'purpose': 'Contains the actual cost values for products within specific cost books',
            'relationships': 'Links: Product2 to CostBook with cost amount, similar to PricebookEntry for costs',
            'dependency': 'Optional - After CostBook and Product2 if tracking costs'
        },
        
        # Product Attributes
        {
            'step': 20, 'object': 'ProductAttributeDefinition',
            'purpose': 'Assigns specific attributes to products and defines how they behave for that product (default values, constraints)',
            'relationships': 'Links: Product2 to AttributeDefinition, Can reference: AttributeCategory for grouping on product',
            'dependency': 'Required after Product2 and AttributeDefinition'
        },
        
        # Pricing
        {
            'step': 21, 'object': 'Pricebook2',
            'purpose': 'Container for product prices, supporting multiple price lists for different markets, currencies, or customer segments',
            'relationships': 'Links to: PricebookEntry (contains actual prices), can have standard and custom pricebooks',
            'dependency': 'Required before PricebookEntry'
        },
        {
            'step': 22, 'object': 'PricebookEntry',
            'purpose': 'Defines the actual price for a product within a specific pricebook and selling model combination',
            'relationships': 'Links: Product2 + Pricebook2 + ProductSellingModel, Requires: ProductSellingModelOption to exist first',
            'dependency': 'Required after Pricebook2 and ProductSellingModelOption'
        },
        
        # Product Organization
        {
            'step': 23, 'object': 'ProductCategoryProduct',
            'purpose': 'Junction object that assigns products to categories, allowing products to appear in multiple categories',
            'relationships': 'Links: Product2 to ProductCategory, enables product browsing and organization',
            'dependency': 'Required after Product2 and ProductCategory'
        },
        
        # Advanced Pricing (Optional)
        {
            'step': 24, 'object': 'PriceAdjustmentSchedule',
            'purpose': 'Defines time-based or quantity-based pricing adjustment schedules (e.g., promotional discounts)',
            'relationships': 'Links to: PriceAdjustmentTier (contains tier details), PricebookEntry (applies adjustments)',
            'dependency': 'Optional - For dynamic pricing strategies'
        },
        {
            'step': 25, 'object': 'PriceAdjustmentTier',
            'purpose': 'Defines specific tiers within price adjustment schedules (e.g., volume discount tiers)',
            'relationships': 'Links to: PriceAdjustmentSchedule (parent schedule), defines discount percentages or amounts',
            'dependency': 'Optional - After PriceAdjustmentSchedule'
        },
        {
            'step': 26, 'object': 'AttributeBasedAdjRule',
            'purpose': 'Defines rules for price adjustments based on product attribute selections',
            'relationships': 'Links to: AttributeBasedAdj (specific adjustments), ProductAttributeDefinition (triggers rules)',
            'dependency': 'Optional - For attribute-based pricing'
        },
        {
            'step': 27, 'object': 'AttributeBasedAdj',
            'purpose': 'Specifies the actual price adjustments when specific attribute values are selected',
            'relationships': 'Links to: AttributeBasedAdjRule (parent rule), AttributeDefinition (attribute that triggers adjustment)',
            'dependency': 'Optional - After AttributeBasedAdjRule'
        },
        
        # Bundle Configuration
        {
            'step': 28, 'object': 'ProductComponentGroup',
            'purpose': 'Groups bundle components into logical sections for configurable products (e.g., "Core Components", "Optional Add-ons")',
            'relationships': 'Links to: Product2 (parent bundle), ProductRelatedComponent (groups components), required for configurable bundles',
            'dependency': 'Required before ProductRelatedComponent for configurable bundles'
        },
        {
            'step': 29, 'object': 'ProductRelatedComponent',
            'purpose': 'Defines parent-child relationships between products for bundles, kits, and product options',
            'relationships': 'Links: Parent Product2 to Child Product2, Uses: ProductComponentGroup (for configurable bundles), ProductRelationshipType',
            'dependency': 'Last step - requires all products and groups to exist'
        }
    ]
    
    # Write data rows
    data_alignment = Alignment(vertical="top", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for i, obj in enumerate(objects, 2):
        ws.cell(row=i, column=1, value=obj['step'])
        ws.cell(row=i, column=2, value=obj['object'])
        ws.cell(row=i, column=3, value=obj['purpose'])
        ws.cell(row=i, column=4, value=obj['relationships'])
        ws.cell(row=i, column=5, value=obj['dependency'])
        
        # Apply formatting
        for col in range(1, 6):
            cell = ws.cell(row=i, column=col)
            cell.alignment = data_alignment
            cell.border = border
    
    # Adjust column widths
    column_widths = [8, 30, 60, 60, 40]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Add summary section
    summary_row = len(objects) + 4
    cell = ws.cell(row=summary_row, column=1, value="KEY CONCEPTS:")
    cell.font = Font(bold=True)
    
    concepts = [
        "• Upload Order: Follow the step sequence to ensure all dependencies are met",
        "• Optional Objects: Steps 2-8 and 19, 24-27 are optional based on your requirements",
        "• Core Requirements: Steps 9-23 and 28-29 are typically required for Revenue Cloud implementations",
        "• Relationships: Most objects reference other objects via ID fields (e.g., Product2Id, AttributeDefinitionId)",
        "• Revenue Cloud Specific: ProductSellingModel and ProductSellingModelOption are required for Revenue Cloud pricing",
        "• Bundles: Configurable bundles require ProductComponentGroup; static bundles can use ProductRelatedComponent directly",
        "• Attributes: Create reusable attributes once, then assign to multiple products with specific configurations"
    ]
    
    for i, concept in enumerate(concepts):
        ws.cell(row=summary_row + i + 1, column=2, value=concept)
        ws.merge_cells(f'B{summary_row + i + 1}:E{summary_row + i + 1}')
    
    # Save workbook
    wb.save('data/Revenue_Cloud_Complete_Upload_Template.xlsx')
    
    print(f"✓ Created complete instructions with {len(objects)} objects")
    print("✓ Included all optional and required objects")
    print("✓ Added clear dependency information")

if __name__ == '__main__':
    main()
