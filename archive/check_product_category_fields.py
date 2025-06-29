#\!/usr/bin/env python3
"""
Check and fix ProductCategory system-managed fields.
"""

import subprocess
import json
from openpyxl import load_workbook

def main():
    print("=" * 60)
    print("CHECKING PRODUCTCATEGORY FIELDS")
    print("=" * 60)
    
    # Get field info for ProductCategory
    cmd = [
        'sf', 'sobject', 'describe',
        '--sobject', 'ProductCategory',
        '--target-org', 'fortradp2',
        '--json'
    ]
    
    result = subprocess.run(cmd, capture_output=True, text=True)
    
    system_fields = []
    if result.returncode == 0:
        data = json.loads(result.stdout)
        if 'result' in data and 'fields' in data['result']:
            for field in data['result']['fields']:
                # Check for system-managed fields (not createable and not updateable)
                if field['name'] \!= 'Id':
                    if not field.get('createable', False) and not field.get('updateable', False):
                        system_fields.append(field['name'])
    
    print(f"System-managed fields found: {system_fields}")
    
    # Load workbook
    wb = load_workbook('data/Revenue_Cloud_Complete_Upload_Template.xlsx')
    ws = wb['12_ProductCategory']
    
    # Get current headers
    headers = []
    header_positions = {}
    for col, cell in enumerate(ws[1], 1):
        if cell.value:
            header_name = str(cell.value).replace('*', '').strip()
            headers.append(header_name)
            header_positions[header_name] = col
    
    print(f"\nCurrent headers in sheet: {headers}")
    
    # Identify which system fields are still in the sheet
    fields_to_remove = []
    for field in system_fields:
        if field in headers:
            fields_to_remove.append(field)
    
    # Also check for any other problematic fields
    additional_problem_fields = ['RootCategoryId', 'Path', 'Level', 'NumberOfProducts']
    for field in additional_problem_fields:
        if field in headers and field not in fields_to_remove:
            fields_to_remove.append(field)
    
    if fields_to_remove:
        print(f"\nRemoving {len(fields_to_remove)} system-managed fields:")
        for field in fields_to_remove:
            print(f"  - {field}")
        
        # Delete columns in reverse order
        cols_to_delete = []
        for field in fields_to_remove:
            if field in header_positions:
                cols_to_delete.append(header_positions[field])
        
        cols_to_delete.sort(reverse=True)
        
        for col_idx in cols_to_delete:
            ws.delete_cols(col_idx)
        
        # Save workbook
        wb.save('data/Revenue_Cloud_Complete_Upload_Template.xlsx')
        print("\n✓ Removed system-managed fields from ProductCategory")
    else:
        print("\n✓ No system-managed fields found in sheet")
    
    # Show remaining headers
    print("\nRemaining headers after cleanup:")
    remaining_headers = []
    for cell in ws[1]:
        if cell.value:
            remaining_headers.append(cell.value)
    print(remaining_headers)

if __name__ == '__main__':
    main()
