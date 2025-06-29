#!/usr/bin/env python3
"""
Fix AttributePicklist DataType values and insert records.
"""

import openpyxl
from pathlib import Path
import subprocess
import pandas as pd
import json

class AttributePicklistFixer:
    def __init__(self):
        self.workbook_path = Path('data/Revenue_Cloud_Complete_Upload_Template.xlsx')
        self.target_org = 'fortradp2'
        
    def fix_datatype(self):
        """Fix DataType values in AttributePicklist sheet."""
        print("=" * 60)
        print("FIXING ATTRIBUTEPICKLIST DATATYPE VALUES")
        print("=" * 60)
        
        # Load workbook
        wb = openpyxl.load_workbook(self.workbook_path)
        ws = wb['14_AttributePicklist']
        
        # Find DataType column
        headers = [cell.value for cell in ws[1]]
        datatype_col = None
        
        for idx, header in enumerate(headers, 1):
            if header and 'DataType' in str(header):
                datatype_col = idx
                break
        
        if datatype_col:
            # Update all DataType values to 'Text'
            updates = 0
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=datatype_col)
                if cell.value == 'Picklist':
                    cell.value = 'Text'
                    updates += 1
            
            wb.save(self.workbook_path)
            print(f"✓ Updated {updates} DataType values from 'Picklist' to 'Text'")
        else:
            print("✗ Could not find DataType column")
            
    def insert_picklists(self):
        """Insert AttributePicklist records."""
        print("\n" + "=" * 60)
        print("INSERTING ATTRIBUTEPICKLIST RECORDS")
        print("=" * 60)
        
        # Read AttributePicklist sheet
        df = pd.read_excel(self.workbook_path, sheet_name='14_AttributePicklist')
        df = df.dropna(how='all')
        
        # Clean column names
        df.columns = df.columns.str.replace('*', '', regex=False)
        
        # Remove Id column for insert
        if 'Id' in df.columns:
            df = df.drop(columns=['Id'])
        
        # Save External_ID__c mapping for later use
        if 'External_ID__c' in df.columns:
            self.code_mapping = dict(zip(df['External_ID__c'], df['Name']))
            df = df.drop(columns=['External_ID__c'])
        
        # Remove empty rows
        df = df[df['Name'].notna()]
        
        print(f"\nInserting {len(df)} AttributePicklist records...")
        print("\nData preview:")
        print(df[['Name', 'DataType', 'Status']].head())
        
        # Save to CSV
        csv_file = Path('data/temp_attribute_picklist.csv')
        df.to_csv(csv_file, index=False)
        
        # Insert records
        cmd = [
            'sf', 'data', 'import', 'bulk',
            '--sobject', 'AttributePicklist',
            '--file', str(csv_file),
            '--target-org', self.target_org,
            '--wait', '10'
        ]
        
        result = subprocess.run(cmd, capture_output=True, text=True)
        
        if result.returncode == 0:
            print("\n✓ Successfully inserted AttributePicklist records")
            
            # Query back to get IDs
            self.query_picklist_ids()
        else:
            print(f"\n✗ Failed to insert: {result.stderr}")
        
        # Clean up
        csv_file.unlink()
        
    def query_picklist_ids(self):
        """Query inserted AttributePicklist records to get their IDs."""
        print("\nQuerying AttributePicklist IDs...")
        
        query = "SELECT Id, Name FROM AttributePicklist ORDER BY Name"
        cmd = [
            'sf', 'data', 'query',
            '--query', query,
            '--target-org', self.target_org,
            '--json'
        ]
        
        result = subprocess.run(cmd, capture_output=True, text=True)
        
        if result.returncode == 0:
            data = json.loads(result.stdout)
            if 'result' in data and 'records' in data['result']:
                records = data['result']['records']
                
                # Create ID mapping
                self.picklist_id_map = {rec['Name']: rec['Id'] for rec in records}
                
                print(f"\n✓ Found {len(records)} AttributePicklist records:")
                for name, id in self.picklist_id_map.items():
                    print(f"  - {name}: {id}")
                
                # Update AttributeDefinition sheet with actual IDs
                self.update_attribute_definitions()
        else:
            print(f"\n✗ Failed to query: {result.stderr}")
    
    def update_attribute_definitions(self):
        """Update AttributeDefinition sheet with actual PicklistId values."""
        print("\n" + "=" * 60)
        print("UPDATING ATTRIBUTEDEFINITION WITH PICKLIST IDS")
        print("=" * 60)
        
        # Load workbook
        wb = openpyxl.load_workbook(self.workbook_path)
        ws = wb['09_AttributeDefinition']
        
        # Find column indices
        headers = [cell.value for cell in ws[1]]
        picklist_id_col = None
        code_col = None
        datatype_col = None
        
        for idx, header in enumerate(headers, 1):
            if header:
                if 'PicklistId' in str(header):
                    picklist_id_col = idx
                elif header == 'Code':
                    code_col = idx
                elif 'DataType' in str(header):
                    datatype_col = idx
        
        if picklist_id_col and code_col and datatype_col:
            # Map AttributeDefinition codes to picklist names
            code_to_picklist = {
                'PT': 'Pricing Tier Options',
                'LT': 'License Type Options',
                'UT': 'Unit Type Options',
                'ST ': 'Server Type Options',  # Note the space
                'TA': 'Agent Type Options',
                'KT': 'Key Type Options',
                'SLT': 'Server Location Type Options',
                'PG': 'PGroup Options',
                'OS': 'Operating System Options'
            }
            
            updates = 0
            for row in range(2, ws.max_row + 1):
                datatype_cell = ws.cell(row=row, column=datatype_col)
                if datatype_cell.value == 'Picklist':
                    code_cell = ws.cell(row=row, column=code_col)
                    if code_cell.value in code_to_picklist:
                        picklist_name = code_to_picklist[code_cell.value]
                        if picklist_name in self.picklist_id_map:
                            picklist_cell = ws.cell(row=row, column=picklist_id_col)
                            picklist_cell.value = self.picklist_id_map[picklist_name]
                            updates += 1
                            print(f"  Updated {code_cell.value} -> {picklist_name} -> {self.picklist_id_map[picklist_name]}")
            
            wb.save(self.workbook_path)
            print(f"\n✓ Updated {updates} AttributeDefinition records with PicklistId values")
        else:
            print("\n✗ Could not find required columns")

def main():
    fixer = AttributePicklistFixer()
    
    # First fix the DataType values
    fixer.fix_datatype()
    
    # Then insert the records
    fixer.insert_picklists()

if __name__ == '__main__':
    main()