#!/usr/bin/env python3
"""
Add AttributePicklistValue sheet to the Revenue Cloud workbook.
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from pathlib import Path

class AttributePicklistValueSheetCreator:
    def __init__(self):
        self.workbook_path = Path('data/Revenue_Cloud_Complete_Upload_Template.xlsx')
        
    def create_sheet(self):
        """Create AttributePicklistValue sheet with proper formatting."""
        print("=" * 60)
        print("ADDING ATTRIBUTEPICKLISTVALUE SHEET")
        print("=" * 60)
        
        # Load workbook
        wb = openpyxl.load_workbook(self.workbook_path)
        
        # Check if sheet already exists
        sheet_name = '18_AttributePicklistValue'
        if sheet_name in wb.sheetnames:
            print(f"Sheet {sheet_name} already exists. Removing and recreating...")
            wb.remove(wb[sheet_name])
        
        # Find position to insert (after ProductAttributeDef)
        sheet_names = wb.sheetnames
        insert_position = None
        if '17_ProductAttributeDef' in sheet_names:
            insert_position = sheet_names.index('17_ProductAttributeDef') + 1
        else:
            # Insert before Pricebook2 if ProductAttributeDef not found
            if '19_Pricebook2' in sheet_names:
                insert_position = sheet_names.index('19_Pricebook2')
        
        # Create new sheet
        ws = wb.create_sheet(sheet_name, insert_position)
        print(f"\nCreated sheet: {sheet_name}")
        if insert_position:
            print(f"  Inserted at position: {insert_position}")
        
        # Define headers based on AttributePicklistValue fields
        headers = [
            'Id',
            'Name*',
            'PicklistId*',
            'Value*',
            'DisplayValue',
            'Code',
            'Sequence',
            'IsDefault',
            'Status',
            'Abbreviation',
            'External_ID__c'
        ]
        
        # Add headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            
            # Apply header formatting (matching other sheets)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Add borders
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            cell.border = thin_border
        
        # Set column widths (matching style of other sheets)
        column_widths = {
            'A': 18,  # Id
            'B': 30,  # Name
            'C': 18,  # PicklistId
            'D': 20,  # Value
            'E': 20,  # DisplayValue
            'F': 15,  # Code
            'G': 10,  # Sequence
            'H': 10,  # IsDefault
            'I': 10,  # Status
            'J': 15,  # Abbreviation
            'K': 20,  # External_ID__c
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # Add some sample rows to demonstrate structure (empty values)
        # This helps users understand the format
        for row in range(2, 5):
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row, column=col)
                # Add borders to data cells too
                cell.border = thin_border
        
        # Add a comment to the first data row explaining the purpose
        comment_cell = ws.cell(row=2, column=2)
        comment_cell.comment = openpyxl.comments.Comment(
            "AttributePicklistValue stores the available values for AttributeDefinition records where DataType='Picklist'",
            "Revenue Cloud"
        )
        
        # Save workbook
        wb.save(self.workbook_path)
        wb.close()
        
        print("\n✓ Sheet created successfully")
        print(f"✓ Saved to: {self.workbook_path}")
        
        # Show field descriptions
        print("\n" + "=" * 60)
        print("ATTRIBUTEPICKLISTVALUE FIELD DESCRIPTIONS")
        print("=" * 60)
        print("Id: Salesforce record ID")
        print("Name*: Name of the picklist value")
        print("PicklistId*: Reference to AttributePicklist parent")
        print("Value*: The actual value stored")
        print("DisplayValue: Value shown to users")
        print("Code: Unique code for the value")
        print("Sequence: Display order")
        print("IsDefault: Whether this is the default value")
        print("Status: Active/Inactive")
        print("Abbreviation: Short form of the value")
        print("External_ID__c: External identifier for upserts")
        print("\n* = Required field")
        
        # Query and populate existing data if any
        self.populate_existing_data(wb)
        
    def populate_existing_data(self, wb=None):
        """Query and populate any existing AttributePicklistValue data."""
        import subprocess
        import json
        
        print("\n" + "-" * 60)
        print("Checking for existing AttributePicklistValue data...")
        
        # Query for existing records
        query = """SELECT Id, Name, PicklistId, Value, DisplayValue, Code, 
                  Sequence, IsDefault, Status, Abbreviation 
                  FROM AttributePicklistValue ORDER BY PicklistId, Sequence"""
        
        cmd = [
            'sf', 'data', 'query',
            '--query', query,
            '--target-org', 'fortradp2',
            '--json'
        ]
        
        result = subprocess.run(cmd, capture_output=True, text=True)
        
        if result.returncode == 0:
            data = json.loads(result.stdout)
            if 'result' in data and 'records' in data['result']:
                records = data['result']['records']
                if records:
                    print(f"Found {len(records)} existing AttributePicklistValue records")
                    
                    # Reload workbook if not provided
                    if wb is None:
                        wb = openpyxl.load_workbook(self.workbook_path)
                    
                    ws = wb['18_AttributePicklistValue']
                    
                    # Populate data starting from row 2
                    for row_idx, record in enumerate(records, 2):
                        ws.cell(row=row_idx, column=1).value = record.get('Id')
                        ws.cell(row=row_idx, column=2).value = record.get('Name')
                        ws.cell(row=row_idx, column=3).value = record.get('PicklistId')
                        ws.cell(row=row_idx, column=4).value = record.get('Value')
                        ws.cell(row=row_idx, column=5).value = record.get('DisplayValue')
                        ws.cell(row=row_idx, column=6).value = record.get('Code')
                        ws.cell(row=row_idx, column=7).value = record.get('Sequence')
                        ws.cell(row=row_idx, column=8).value = record.get('IsDefault')
                        ws.cell(row=row_idx, column=9).value = record.get('Status')
                        ws.cell(row=row_idx, column=10).value = record.get('Abbreviation')
                        
                        # Apply borders
                        thin_border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )
                        for col in range(1, 12):
                            ws.cell(row=row_idx, column=col).border = thin_border
                    
                    wb.save(self.workbook_path)
                    print(f"✓ Populated {len(records)} records")
                else:
                    print("No AttributePicklistValue records found in org")
        else:
            print("Could not query AttributePicklistValue data")

def main():
    creator = AttributePicklistValueSheetCreator()
    creator.create_sheet()

if __name__ == '__main__':
    main()