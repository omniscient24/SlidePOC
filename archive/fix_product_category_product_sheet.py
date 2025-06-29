#\!/usr/bin/env python3
"""
Fix ProductCategoryProduct sheet by adding ID column and mapping existing records.
"""

import pandas as pd
import subprocess
import json
from openpyxl import load_workbook

def main():
    print("=" * 60)
    print("FIXING PRODUCTCATEGORYPRODUCT SHEET")
    print("=" * 60)
    
    # Get existing records from org
    query = "SELECT Id, ProductCategoryId, ProductId FROM ProductCategoryProduct"
    
    cmd = [
        'sf', 'data', 'query',
        '--query', query,
        '--target-org', 'fortradp2',
        '--json'
    ]
    
    result = subprocess.run(cmd, capture_output=True, text=True)
    
    existing_records = {}
    if result.returncode == 0:
        data = json.loads(result.stdout)
        if 'result' in data and 'records' in data['result']:
            for rec in data['result']['records']:
                key = f"{rec['ProductCategoryId']}|{rec['ProductId']}"
                existing_records[key] = rec['Id']
    
    print(f"Found {len(existing_records)} existing records in org")
    
    # Load workbook
    wb = load_workbook('data/Revenue_Cloud_Complete_Upload_Template.xlsx')
    ws = wb['26_ProductCategoryProduct']
    
    # Get current headers
    headers = [cell.value for cell in ws[1]]
    print(f"Current headers: {headers}")
    
    # Add Id column if not present
    if 'Id' not in headers:
        # Insert new column at beginning
        ws.insert_cols(1)
        ws.cell(row=1, column=1, value='Id')
        
        # Shift all data one column to the right is already done by insert_cols
        print("Added Id column")
    
    # Now map IDs to existing records
    id_col = 1  # Id is now first column
    cat_col = headers.index('ProductCategoryId*') + 2 if 'Id' not in headers else headers.index('ProductCategoryId*') + 1
    prod_col = headers.index('ProductId*') + 2 if 'Id' not in headers else headers.index('ProductId*') + 1
    
    mapped = 0
    for row in range(2, ws.max_row + 1):
        cat_id = ws.cell(row=row, column=cat_col).value
        prod_id = ws.cell(row=row, column=prod_col).value
        
        if cat_id and prod_id:
            key = f"{cat_id}|{prod_id}"
            if key in existing_records:
                ws.cell(row=row, column=id_col, value=existing_records[key])
                mapped += 1
    
    print(f"Mapped {mapped} IDs to existing records")
    
    # Save workbook
    wb.save('data/Revenue_Cloud_Complete_Upload_Template.xlsx')
    print("✓ Updated workbook saved")
    
    # Verify the fix
    print("\nVerifying updated sheet:")
    df = pd.read_excel('data/Revenue_Cloud_Complete_Upload_Template.xlsx', sheet_name='26_ProductCategoryProduct')
    print(f"Total rows: {len(df)}")
    print(f"Columns: {list(df.columns)}")
    
    if 'Id' in df.columns:
        has_id = df[df['Id'].notna()]
        no_id = df[df['Id'].isna()]
        print(f"Records with IDs: {len(has_id)}")
        print(f"Records without IDs (new): {len(no_id)}")
        
        if len(no_id) > 0:
            print("\nNew records to be inserted:")
            for idx, row in no_id.iterrows():
                print(f"  Category: {row.get('ProductCategoryId*', 'N/A')}, Product: {row.get('ProductId*', 'N/A')}")

if __name__ == '__main__':
    main()
