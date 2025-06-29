#!/usr/bin/env python3
"""
Update PricebookEntry sheet with Name and ProductCode from Product2 sheet.
"""

import openpyxl
import pandas as pd
from pathlib import Path

class PricebookEntryUpdater:
    def __init__(self):
        self.workbook_path = Path('data/Revenue_Cloud_Complete_Upload_Template.xlsx')
        
    def update_pricebook_entries(self):
        """Update PricebookEntry sheet with Product2 data."""
        print("=" * 60)
        print("UPDATING PRICEBOOKENTRY WITH PRODUCT NAMES AND CODES")
        print("=" * 60)
        
        # Load workbook
        wb = openpyxl.load_workbook(self.workbook_path)
        
        # Read Product2 data
        print("\nReading Product2 sheet...")
        product_sheet = wb['13_Product2']
        
        # Get Product2 headers
        product_headers = []
        for cell in product_sheet[1]:
            if cell.value:
                product_headers.append(cell.value)
            else:
                break
        
        # Find column indices for Product2
        try:
            product_id_col = product_headers.index('Id') + 1
            product_name_col = product_headers.index('Name') + 1
            product_code_col = product_headers.index('ProductCode') + 1
        except ValueError as e:
            print(f"Error: Could not find required column in Product2: {e}")
            return
        
        # Build Product2 lookup dictionary
        product_lookup = {}
        row_num = 2
        while True:
            id_cell = product_sheet.cell(row=row_num, column=product_id_col)
            if not id_cell.value:
                break
            
            product_id = id_cell.value
            product_name = product_sheet.cell(row=row_num, column=product_name_col).value
            product_code = product_sheet.cell(row=row_num, column=product_code_col).value
            
            product_lookup[product_id] = {
                'Name': product_name,
                'ProductCode': product_code
            }
            row_num += 1
        
        print(f"  Found {len(product_lookup)} products")
        
        # Update PricebookEntry sheet
        print("\nUpdating PricebookEntry sheet...")
        pbe_sheet = wb['20_PricebookEntry']
        
        # Get PricebookEntry headers
        pbe_headers = []
        for cell in pbe_sheet[1]:
            if cell.value:
                pbe_headers.append(cell.value)
            else:
                break
        
        # Find column indices for PricebookEntry
        try:
            pbe_product_id_col = pbe_headers.index('Product2Id') + 1
            pbe_name_col = pbe_headers.index('Name') + 1
            pbe_product_code_col = pbe_headers.index('ProductCode') + 1
        except ValueError as e:
            print(f"Error: Could not find required column in PricebookEntry: {e}")
            return
        
        # Update each PricebookEntry row
        updated_count = 0
        row_num = 2
        while True:
            product_id_cell = pbe_sheet.cell(row=row_num, column=pbe_product_id_col)
            if not product_id_cell.value:
                break
            
            product_id = product_id_cell.value
            
            # Look up product info
            if product_id in product_lookup:
                product_info = product_lookup[product_id]
                
                # Update Name
                name_cell = pbe_sheet.cell(row=row_num, column=pbe_name_col)
                name_cell.value = product_info['Name']
                
                # Update ProductCode
                code_cell = pbe_sheet.cell(row=row_num, column=pbe_product_code_col)
                code_cell.value = product_info['ProductCode']
                
                updated_count += 1
                
                if updated_count <= 5:  # Show first 5 updates
                    print(f"  Row {row_num}: {product_info['Name']} ({product_info['ProductCode']})")
            else:
                print(f"  ⚠️  Row {row_num}: Product ID {product_id} not found in Product2 sheet")
            
            row_num += 1
        
        # Save workbook
        wb.save(self.workbook_path)
        wb.close()
        
        print(f"\n✓ Updated {updated_count} PricebookEntry records")
        print(f"✓ Saved to: {self.workbook_path}")
        
        # Show summary
        print("\n" + "=" * 60)
        print("UPDATE COMPLETE")
        print("=" * 60)
        print(f"The PricebookEntry sheet now contains:")
        print(f"- Name: Product names from Product2 sheet")
        print(f"- ProductCode: Product codes from Product2 sheet")
        print(f"- All matched by Product2Id")

def main():
    updater = PricebookEntryUpdater()
    updater.update_pricebook_entries()

if __name__ == '__main__':
    main()