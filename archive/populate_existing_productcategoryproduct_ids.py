#!/usr/bin/env python3
"""
Populate existing ProductCategoryProduct IDs from org.
"""

import subprocess
import json
from openpyxl import load_workbook

def main():
    print("Populating existing ProductCategoryProduct IDs...")
    
    # Query all ProductCategoryProduct records
    query = "SELECT Id, ProductCategoryId, ProductId FROM ProductCategoryProduct"
    
    cmd = [
        'sf', 'data', 'query',
        '--query', query,
        '--target-org', 'fortradp2',
        '--json'
    ]
    
    result = subprocess.run(cmd, capture_output=True, text=True)
    
    id_map = {}
    if result.returncode == 0:
        data = json.loads(result.stdout)
        if 'result' in data and 'records' in data['result']:
            for rec in data['result']['records']:
                key = f"{rec['ProductCategoryId']}|{rec['ProductId']}"
                id_map[key] = rec['Id']
    
    print(f"Found {len(id_map)} existing records in org")
    
    # Load workbook
    wb = load_workbook('data/Revenue_Cloud_Complete_Upload_Template.xlsx')
    ws = wb['26_ProductCategoryProduct']
    
    # Find column positions
    id_col = None
    category_col = None
    product_col = None
    
    for col, cell in enumerate(ws[1], 1):
        if cell.value:
            clean_value = str(cell.value).replace('*', '').strip()
            if clean_value == 'Id':
                id_col = col
            elif clean_value == 'ProductCategoryId':
                category_col = col
            elif clean_value == 'ProductId':
                product_col = col
    
    # Populate IDs
    populated = 0
    for row in range(2, ws.max_row + 1):
        if category_col and product_col:
            category_id = ws.cell(row=row, column=category_col).value
            product_id = ws.cell(row=row, column=product_col).value
            
            if category_id and product_id:
                key = f"{category_id}|{product_id}"
                if key in id_map and id_col:
                    current_id = ws.cell(row=row, column=id_col).value
                    if not current_id:  # Only populate if empty
                        ws.cell(row=row, column=id_col).value = id_map[key]
                        populated += 1
                        print(f"  Populated row {row}: {id_map[key]}")
    
    print(f"\nPopulated {populated} IDs")
    
    # Save workbook
    wb.save('data/Revenue_Cloud_Complete_Upload_Template.xlsx')
    print("✓ Updated workbook")

if __name__ == '__main__':
    main()