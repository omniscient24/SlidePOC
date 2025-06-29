#!/usr/bin/env python3
"""
Remove TransferRecordMode field from Product2 sheet.
"""

from openpyxl import load_workbook

def main():
    print("Removing TransferRecordMode from Product2...")
    
    # Load workbook
    wb = load_workbook('data/Revenue_Cloud_Complete_Upload_Template.xlsx')
    ws = wb['13_Product2']
    
    # Find and remove TransferRecordMode column
    col_to_remove = None
    for col, cell in enumerate(ws[1], 1):
        if cell.value and 'TransferRecordMode' in str(cell.value):
            col_to_remove = col
            break
    
    if col_to_remove:
        ws.delete_cols(col_to_remove)
        print(f"Removed TransferRecordMode column at position {col_to_remove}")
    else:
        print("TransferRecordMode column not found")
    
    # Save workbook
    wb.save('data/Revenue_Cloud_Complete_Upload_Template.xlsx')
    print("✓ Updated workbook")

if __name__ == '__main__':
    main()