#!/usr/bin/env python3
"""
Assign AttributeCategories to ProductAttributeDefinition records based on attribute types.
"""

import pandas as pd
from openpyxl import load_workbook

def main():
    print("Assigning AttributeCategories to ProductAttributeDefinitions...")
    
    # Read sheets
    wb = load_workbook('data/Revenue_Cloud_Complete_Upload_Template.xlsx')
    
    # Get AttributeCategory mappings
    df_cat = pd.read_excel('data/Revenue_Cloud_Complete_Upload_Template.xlsx', sheet_name='10_AttributeCategory')
    df_cat.columns = df_cat.columns.str.replace('*', '', regex=False).str.strip()
    
    # Get AttributeDefinition mappings
    df_attr = pd.read_excel('data/Revenue_Cloud_Complete_Upload_Template.xlsx', sheet_name='09_AttributeDefinition')
    df_attr.columns = df_attr.columns.str.replace('*', '', regex=False).str.strip()
    
    # Create category mapping based on attribute names/codes
    category_map = {
        # Deployment Option Attributes (DOA)
        'DO': '0v3dp00000000BJAAY',  # Deployment Option
        'MS': '0v3dp00000000BJAAY',  # Managed Service
        'AG': '0v3dp00000000BJAAY',  # Agents
        'TA': '0v3dp00000000BJAAY',  # Type of Agents
        
        # Service Attributes (SA)
        'OS': '0v3dp00000000BLAAY',  # Operating System
        'ST': '0v3dp00000000BLAAY',  # Server Type
        'SLT': '0v3dp00000000BLAAY', # Server Location Type
        'CR': '0v3dp00000000BLAAY',  # Cluster Ready
        
        # Storage (STO)
        'STO': '0v3dp00000000BNAAY', # Studios
        'UN': '0v3dp00000000BNAAY',  # Units
        'UT': '0v3dp00000000BNAAY',  # Unit Type
        
        # Support (SP001)
        'MT': '0v3dp00000000BMAAY',  # Maintenance Type
        'VER': '0v3dp00000000BMAAY', # Version
        'LT': '0v3dp00000000BMAAY',  # License Type
        'KT': '0v3dp00000000BMAAY',  # Key Type
        
        # Training (T001)
        'Term': '0v3dp00000000BOAAY', # Term
        'USR': '0v3dp00000000BOAAY',  # Users
        'MYCAP': '0v3dp00000000BOAAY', # MYCAP
        'PG': '0v3dp00000000BOAAY',   # PGroup
        'PT': '0v3dp00000000BOAAY',   # Pricing Tier
    }
    
    # Create attribute code to ID mapping
    attr_code_map = dict(zip(df_attr['Code'], df_attr['Id']))
    
    # Load ProductAttributeDef sheet
    ws = wb['17_ProductAttributeDef']
    
    # Find column positions
    attr_def_col = None
    category_col = None
    
    for col, cell in enumerate(ws[1], 1):
        if cell.value:
            clean_value = str(cell.value).replace('*', '').strip()
            if clean_value == 'AttributeDefinitionId':
                attr_def_col = col
            elif clean_value == 'AttributeCategoryId':
                category_col = col
    
    if not attr_def_col or not category_col:
        print("ERROR: Could not find required columns")
        return
    
    # Assign categories
    assigned = 0
    for row in range(2, ws.max_row + 1):
        attr_def_id = ws.cell(row=row, column=attr_def_col).value
        
        if attr_def_id:
            # Find the attribute code
            attr_code = None
            for code, id_val in attr_code_map.items():
                if id_val == attr_def_id:
                    attr_code = code
                    break
            
            if attr_code and attr_code in category_map:
                ws.cell(row=row, column=category_col).value = category_map[attr_code]
                assigned += 1
                print(f"  Assigned {attr_code} -> {category_map[attr_code]}")
    
    print(f"\nAssigned categories to {assigned} records")
    
    # Save workbook
    wb.save('data/Revenue_Cloud_Complete_Upload_Template.xlsx')
    print("✓ Updated workbook")
    
    # Show summary
    print("\n\nCategory Assignment Summary:")
    print("=" * 60)
    
    # Read updated data to show summary
    df_updated = pd.read_excel('data/Revenue_Cloud_Complete_Upload_Template.xlsx', sheet_name='17_ProductAttributeDef')
    df_updated.columns = df_updated.columns.str.replace('*', '', regex=False).str.strip()
    
    category_usage = df_updated['AttributeCategoryId'].value_counts()
    
    for cat_id, count in category_usage.items():
        if pd.notna(cat_id):
            cat_name = df_cat[df_cat['Id'] == cat_id]['Name'].iloc[0] if not df_cat[df_cat['Id'] == cat_id].empty else 'Unknown'
            print(f"{cat_name}: {count} attributes")

if __name__ == '__main__':
    main()