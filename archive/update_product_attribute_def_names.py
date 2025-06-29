#!/usr/bin/env python3
"""
Update ProductAttributeDef sheet with Product2.Name from Product2 sheet.
"""

import openpyxl
from pathlib import Path

class ProductAttributeDefUpdater:
    def __init__(self):
        self.workbook_path = Path('data/Revenue_Cloud_Complete_Upload_Template.xlsx')
        
    def update_product_attribute_def(self):
        """Update ProductAttributeDef sheet with Product2 names."""
        print("=" * 60)
        print("UPDATING PRODUCTATTRIBUTEDEF WITH PRODUCT NAMES")
        print("=" * 60)
        
        # Load workbook
        wb = openpyxl.load_workbook(self.workbook_path)
        
        # Read Product2 data
        print("\nReading Product2 sheet...")
        product_sheet = wb['13_Product2']
        
        # Get Product2 headers
        product_headers = []
        for cell in product_sheet[1]:
            if cell.value:
                product_headers.append(cell.value)
            else:
                break
        
        # Find column indices for Product2
        try:
            product_id_col = product_headers.index('Id') + 1
            product_name_col = product_headers.index('Name') + 1
        except ValueError as e:
            print(f"Error: Could not find required column in Product2: {e}")
            return
        
        # Build Product2 lookup dictionary
        product_lookup = {}
        row_num = 2
        while True:
            id_cell = product_sheet.cell(row=row_num, column=product_id_col)
            if not id_cell.value:
                break
            
            product_id = id_cell.value
            product_name = product_sheet.cell(row=row_num, column=product_name_col).value
            
            product_lookup[product_id] = product_name
            row_num += 1
        
        print(f"  Found {len(product_lookup)} products")
        
        # Update ProductAttributeDef sheet
        print("\nUpdating ProductAttributeDef sheet...")
        pad_sheet = wb['17_ProductAttributeDef']
        
        # Get ProductAttributeDef headers
        pad_headers = []
        for cell in pad_sheet[1]:
            if cell.value:
                pad_headers.append(str(cell.value))
            else:
                break
        
        # Find column indices for ProductAttributeDef
        try:
            pad_product_id_col = pad_headers.index('Product2Id') + 1
            # Look for Product2.Name column
            pad_product_name_col = None
            for idx, header in enumerate(pad_headers):
                if header == 'Product2.Name':
                    pad_product_name_col = idx + 1
                    break
            
            if pad_product_name_col is None:
                print("Error: Could not find 'Product2.Name' column in ProductAttributeDef")
                return
                
        except ValueError as e:
            print(f"Error: Could not find required column in ProductAttributeDef: {e}")
            return
        
        print(f"  Product2Id column: {pad_product_id_col}")
        print(f"  Product2.Name column: {pad_product_name_col}")
        
        # Update each ProductAttributeDef row
        updated_count = 0
        row_num = 2
        while True:
            product_id_cell = pad_sheet.cell(row=row_num, column=pad_product_id_col)
            if not product_id_cell.value:
                break
            
            product_id = product_id_cell.value
            
            # Look up product name
            if product_id in product_lookup:
                product_name = product_lookup[product_id]
                
                # Update Product2.Name
                name_cell = pad_sheet.cell(row=row_num, column=pad_product_name_col)
                name_cell.value = product_name
                
                updated_count += 1
                
                if updated_count <= 5:  # Show first 5 updates
                    print(f"  Row {row_num}: Product2Id {product_id} → {product_name}")
            else:
                print(f"  ⚠️  Row {row_num}: Product ID {product_id} not found in Product2 sheet")
            
            row_num += 1
        
        # Save workbook
        wb.save(self.workbook_path)
        wb.close()
        
        print(f"\n✓ Updated {updated_count} ProductAttributeDef records")
        print(f"✓ Saved to: {self.workbook_path}")
        
        # Show summary
        print("\n" + "=" * 60)
        print("UPDATE COMPLETE")
        print("=" * 60)
        print(f"The ProductAttributeDef sheet now contains:")
        print(f"- Product2.Name: Product names from Product2 sheet")
        print(f"- All matched by Product2Id")

def main():
    updater = ProductAttributeDefUpdater()
    updater.update_product_attribute_def()

if __name__ == '__main__':
    main()