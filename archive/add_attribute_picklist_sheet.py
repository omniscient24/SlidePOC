#!/usr/bin/env python3
"""
Add AttributePicklist sheet to the Revenue Cloud workbook.
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from pathlib import Path

class AttributePicklistSheetCreator:
    def __init__(self):
        self.workbook_path = Path('data/Revenue_Cloud_Complete_Upload_Template.xlsx')
        
    def create_sheet(self):
        """Create AttributePicklist sheet with proper formatting."""
        print("=" * 60)
        print("ADDING ATTRIBUTEPICKLIST SHEET")
        print("=" * 60)
        
        # Load workbook
        wb = openpyxl.load_workbook(self.workbook_path)
        
        # Check if sheet already exists
        sheet_name = '14_AttributePicklist'
        if sheet_name in wb.sheetnames:
            print(f"Sheet {sheet_name} already exists. Removing and recreating...")
            wb.remove(wb[sheet_name])
        
        # Find position to insert (after Product2)
        sheet_names = wb.sheetnames
        insert_position = None
        if '13_Product2' in sheet_names:
            insert_position = sheet_names.index('13_Product2') + 1
        else:
            # Insert before ProductSellingModel if Product2 not found
            if '15_ProductSellingModel' in sheet_names:
                insert_position = sheet_names.index('15_ProductSellingModel')
        
        # Create new sheet
        ws = wb.create_sheet(sheet_name, insert_position)
        print(f"\nCreated sheet: {sheet_name}")
        if insert_position:
            print(f"  Inserted at position: {insert_position}")
        
        # Define headers based on AttributePicklist fields
        headers = [
            'Id',
            'Name*',
            'Description',
            'Status',
            'DataType*',
            'UnitOfMeasureId',
            'OwnerId',
            'External_ID__c'
        ]
        
        # Add headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            
            # Apply header formatting (matching other sheets)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Add borders
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            cell.border = thin_border
        
        # Set column widths (matching style of other sheets)
        column_widths = {
            'A': 18,  # Id
            'B': 30,  # Name
            'C': 40,  # Description
            'D': 10,  # Status
            'E': 15,  # DataType
            'F': 18,  # UnitOfMeasureId
            'G': 18,  # OwnerId
            'H': 20,  # External_ID__c
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # Add sample data based on the AttributeDefinitions that need picklists
        picklist_data = [
            {'Name': 'Pricing Tier Options', 'Description': 'Available pricing tiers', 'DataType': 'Picklist', 'Status': 'Active', 'Code': 'PT_PICKLIST'},
            {'Name': 'License Type Options', 'Description': 'Types of licenses', 'DataType': 'Picklist', 'Status': 'Active', 'Code': 'LT_PICKLIST'},
            {'Name': 'Unit Type Options', 'Description': 'Types of units', 'DataType': 'Picklist', 'Status': 'Active', 'Code': 'UT_PICKLIST'},
            {'Name': 'Server Type Options', 'Description': 'Types of servers', 'DataType': 'Picklist', 'Status': 'Active', 'Code': 'ST_PICKLIST'},
            {'Name': 'Agent Type Options', 'Description': 'Types of agents', 'DataType': 'Picklist', 'Status': 'Active', 'Code': 'TA_PICKLIST'},
            {'Name': 'Key Type Options', 'Description': 'Types of keys', 'DataType': 'Picklist', 'Status': 'Active', 'Code': 'KT_PICKLIST'},
            {'Name': 'Server Location Type Options', 'Description': 'Server location types', 'DataType': 'Picklist', 'Status': 'Active', 'Code': 'SLT_PICKLIST'},
            {'Name': 'PGroup Options', 'Description': 'Product group options', 'DataType': 'Picklist', 'Status': 'Active', 'Code': 'PG_PICKLIST'},
            {'Name': 'Operating System Options', 'Description': 'Operating system options', 'DataType': 'Picklist', 'Status': 'Active', 'Code': 'OS_PICKLIST'},
        ]
        
        # Add the sample data rows
        for row_idx, data in enumerate(picklist_data, 2):
            ws.cell(row=row_idx, column=1).value = None  # Id (empty for new)
            ws.cell(row=row_idx, column=2).value = data['Name']
            ws.cell(row=row_idx, column=3).value = data['Description']
            ws.cell(row=row_idx, column=4).value = data['Status']
            ws.cell(row=row_idx, column=5).value = data['DataType']
            ws.cell(row=row_idx, column=6).value = None  # UnitOfMeasureId
            ws.cell(row=row_idx, column=7).value = None  # OwnerId
            ws.cell(row=row_idx, column=8).value = data['Code']  # External_ID__c
            
            # Apply borders to data cells
            for col in range(1, 9):
                ws.cell(row=row_idx, column=col).border = thin_border
        
        # Add a comment to the first data row
        comment_cell = ws.cell(row=2, column=2)
        comment_cell.comment = openpyxl.comments.Comment(
            "AttributePicklist defines the picklist that AttributeDefinition records reference when DataType='Picklist'",
            "Revenue Cloud"
        )
        
        # Save workbook
        wb.save(self.workbook_path)
        wb.close()
        
        print("\n✓ Sheet created successfully")
        print(f"✓ Saved to: {self.workbook_path}")
        print(f"✓ Added {len(picklist_data)} sample AttributePicklist records")
        
        # Show field descriptions
        print("\n" + "=" * 60)
        print("ATTRIBUTEPICKLIST FIELD DESCRIPTIONS")
        print("=" * 60)
        print("Id: Salesforce record ID")
        print("Name*: Name of the picklist")
        print("Description: Description of the picklist")
        print("Status: Active/Inactive")
        print("DataType*: Type of data (usually 'Picklist')")
        print("UnitOfMeasureId: Reference to unit of measure (if applicable)")
        print("OwnerId: Record owner")
        print("External_ID__c: External identifier for upserts")
        print("\n* = Required field")
        
        # Update AttributeDefinition sheet to reference these picklists
        self.update_attribute_definition_sheet(wb)
        
    def update_attribute_definition_sheet(self, wb=None):
        """Update AttributeDefinition sheet to reference the AttributePicklists."""
        print("\n" + "-" * 60)
        print("Updating AttributeDefinition sheet with PicklistId references...")
        
        if wb is None:
            wb = openpyxl.load_workbook(self.workbook_path)
        
        # Get AttributeDefinition sheet
        if '09_AttributeDefinition' in wb.sheetnames:
            ad_sheet = wb['09_AttributeDefinition']
            
            # Find column indices
            headers = []
            picklist_id_col = None
            datatype_col = None
            code_col = None
            
            for col_idx, cell in enumerate(ad_sheet[1], 1):
                if cell.value:
                    headers.append(cell.value)
                    if 'PicklistId' in str(cell.value):
                        picklist_id_col = col_idx
                    elif 'DataType' in str(cell.value):
                        datatype_col = col_idx
                    elif cell.value == 'Code':
                        code_col = col_idx
            
            if picklist_id_col and datatype_col and code_col:
                # Map AttributeDefinition codes to AttributePicklist codes
                picklist_mapping = {
                    'PT': 'PT_PICKLIST',
                    'LT': 'LT_PICKLIST',
                    'UT': 'UT_PICKLIST',
                    'ST ': 'ST_PICKLIST',  # Note the space in the code
                    'TA': 'TA_PICKLIST',
                    'KT': 'KT_PICKLIST',
                    'SLT': 'SLT_PICKLIST',
                    'PG': 'PG_PICKLIST',
                    'OS': 'OS_PICKLIST'
                }
                
                # Update rows with DataType='Picklist'
                updates = 0
                for row in range(2, ad_sheet.max_row + 1):
                    datatype_cell = ad_sheet.cell(row=row, column=datatype_col)
                    if datatype_cell.value == 'Picklist':
                        code_cell = ad_sheet.cell(row=row, column=code_col)
                        if code_cell.value in picklist_mapping:
                            picklist_cell = ad_sheet.cell(row=row, column=picklist_id_col)
                            picklist_cell.value = f"REF_{picklist_mapping[code_cell.value]}"
                            updates += 1
                
                print(f"  ✓ Updated {updates} AttributeDefinition records with PicklistId references")
                
                wb.save(self.workbook_path)
                print(f"  ✓ Saved updates to workbook")
            else:
                print("  ⚠️  Could not find required columns in AttributeDefinition sheet")

def main():
    creator = AttributePicklistSheetCreator()
    creator.create_sheet()

if __name__ == '__main__':
    main()