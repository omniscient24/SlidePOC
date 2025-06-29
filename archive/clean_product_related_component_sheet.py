#\!/usr/bin/env python3
"""
Clean ProductRelatedComponent sheet by removing invalid columns.
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def main():
    print("=" * 60)
    print("CLEANING PRODUCTRELATEDCOMPONENT SHEET")
    print("=" * 60)
    
    # Valid fields for ProductRelatedComponent (creatable/updateable)
    valid_fields = [
        'ParentProductId',
        'ChildProductId',
        'ChildProductClassificationId',
        'ChildSellingModelId',
        'ParentSellingModelId',
        'ProductComponentGroupId',
        'ProductRelationshipTypeId',
        'Quantity',
        'MinQuantity',
        'MaxQuantity',
        'IsQuantityEditable',
        'IsComponentRequired',
        'IsDefaultComponent',
        'Sequence',
        'DoesBundlePriceIncludeChild',
        'QuantityScaleMethod',
        'QuoteVisibility'
    ]
    
    # Fields to remove (read-only or non-existent)
    invalid_fields = ['Name', 'External_ID__c']
    
    # Load workbook
    wb = load_workbook('data/Revenue_Cloud_Complete_Upload_Template.xlsx')
    ws = wb['25_ProductRelatedComponent']
    
    # Get current headers and their positions
    headers = []
    header_positions = {}
    for col, cell in enumerate(ws[1], 1):
        if cell.value:
            header_name = cell.value.replace('*', '').strip()
            headers.append(header_name)
            header_positions[header_name] = col
    
    print(f"Current headers: {headers}")
    print(f"\nRemoving invalid columns: {invalid_fields}")
    
    # Identify columns to delete (in reverse order to avoid index issues)
    cols_to_delete = []
    for field in invalid_fields:
        if field in header_positions:
            cols_to_delete.append(header_positions[field])
    
    # Sort in reverse order
    cols_to_delete.sort(reverse=True)
    
    # Delete columns
    for col_idx in cols_to_delete:
        ws.delete_cols(col_idx)
        print(f"  - Removed column {col_idx} ({field})")
    
    # Update headers to show which are required
    required_fields = ['ParentProductId', 'ChildProductId']
    
    # Re-read headers after deletion
    for col, cell in enumerate(ws[1], 1):
        if cell.value:
            header_name = cell.value.replace('*', '').strip()
            if header_name in required_fields:
                cell.value = f"{header_name}*"
    
    # Save workbook
    wb.save('data/Revenue_Cloud_Complete_Upload_Template.xlsx')
    
    print("\n✓ Sheet cleaned successfully")
    
    # Verify the changes
    print("\nRemaining valid columns:")
    remaining_headers = []
    for cell in ws[1]:
        if cell.value:
            remaining_headers.append(cell.value)
    
    for i, header in enumerate(remaining_headers, 1):
        print(f"  {i}. {header}")
    
    print(f"\nTotal valid columns: {len(remaining_headers)}")

if __name__ == '__main__':
    main()
