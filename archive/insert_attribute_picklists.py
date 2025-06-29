#!/usr/bin/env python3
"""
Insert AttributePicklist records and update AttributeDefinition references.
"""

import subprocess
import pandas as pd
import json
from pathlib import Path

class AttributePicklistImporter:
    def __init__(self):
        self.excel_file = Path('data/Revenue_Cloud_Complete_Upload_Template.xlsx')
        self.target_org = 'fortradp2'
        
    def insert_picklists(self):
        """Insert AttributePicklist records."""
        print("=" * 60)
        print("INSERTING ATTRIBUTEPICKLIST RECORDS")
        print("=" * 60)
        
        # Read AttributePicklist sheet
        df = pd.read_excel(self.excel_file, sheet_name='14_AttributePicklist')
        df = df.dropna(how='all')
        
        # Clean column names
        df.columns = df.columns.str.replace('*', '', regex=False)
        
        # Remove Id column for insert
        if 'Id' in df.columns:
            df = df.drop(columns=['Id'])
        
        # Remove External_ID__c since it's not a real field
        if 'External_ID__c' in df.columns:
            # Save mapping for later use
            self.code_mapping = dict(zip(df['External_ID__c'], df['Name']))
            df = df.drop(columns=['External_ID__c'])
        
        # Remove empty rows
        df = df[df['Name'].notna()]
        
        print(f"\nInserting {len(df)} AttributePicklist records...")
        
        # Save to CSV
        csv_file = Path('data/temp_attribute_picklist.csv')
        df.to_csv(csv_file, index=False)
        
        # Insert records
        cmd = [
            'sf', 'data', 'import', 'bulk',
            '--sobject', 'AttributePicklist',
            '--file', str(csv_file),
            '--target-org', self.target_org,
            '--wait', '10'
        ]
        
        result = subprocess.run(cmd, capture_output=True, text=True)
        
        if result.returncode == 0:
            print("✓ Successfully inserted AttributePicklist records")
            
            # Query back to get IDs
            self.query_picklist_ids()
        else:
            print(f"✗ Failed to insert: {result.stderr}")
        
        # Clean up
        csv_file.unlink()
        
    def query_picklist_ids(self):
        """Query inserted AttributePicklist records to get their IDs."""
        print("\nQuerying AttributePicklist IDs...")
        
        query = "SELECT Id, Name FROM AttributePicklist"
        cmd = [
            'sf', 'data', 'query',
            '--query', query,
            '--target-org', self.target_org,
            '--json'
        ]
        
        result = subprocess.run(cmd, capture_output=True, text=True)
        
        if result.returncode == 0:
            data = json.loads(result.stdout)
            if 'result' in data and 'records' in data['result']:
                records = data['result']['records']
                
                # Create ID mapping
                self.picklist_id_map = {rec['Name']: rec['Id'] for rec in records}
                
                print(f"✓ Found {len(records)} AttributePicklist records")
                for name, id in self.picklist_id_map.items():
                    print(f"  - {name}: {id}")
                
                # Update AttributeDefinition sheet
                self.update_attribute_definitions()
        else:
            print(f"✗ Failed to query: {result.stderr}")
    
    def update_attribute_definitions(self):
        """Update AttributeDefinition sheet with actual PicklistId values."""
        print("\nUpdating AttributeDefinition sheet with PicklistId values...")
        
        # Load workbook
        import openpyxl
        wb = openpyxl.load_workbook(self.excel_file)
        ws = wb['09_AttributeDefinition']
        
        # Find column indices
        headers = [cell.value for cell in ws[1]]
        picklist_id_col = None
        code_col = None
        
        for idx, header in enumerate(headers, 1):
            if header and 'PicklistId' in str(header):
                picklist_id_col = idx
            elif header == 'Code':
                code_col = idx
        
        if picklist_id_col and code_col:
            # Map AttributeDefinition codes to picklist names
            code_to_picklist = {
                'PT': 'Pricing Tier Options',
                'LT': 'License Type Options',
                'UT': 'Unit Type Options',
                'ST ': 'Server Type Options',
                'TA': 'Agent Type Options',
                'KT': 'Key Type Options',
                'SLT': 'Server Location Type Options',
                'PG': 'PGroup Options',
                'OS': 'Operating System Options'
            }
            
            updates = 0
            for row in range(2, ws.max_row + 1):
                code_cell = ws.cell(row=row, column=code_col)
                if code_cell.value in code_to_picklist:
                    picklist_name = code_to_picklist[code_cell.value]
                    if picklist_name in self.picklist_id_map:
                        picklist_cell = ws.cell(row=row, column=picklist_id_col)
                        picklist_cell.value = self.picklist_id_map[picklist_name]
                        updates += 1
            
            wb.save(self.excel_file)
            print(f"✓ Updated {updates} AttributeDefinition records with PicklistId values")
            
            # Now run AttributeDefinition import
            self.import_attribute_definitions()
        else:
            print("✗ Could not find required columns")
    
    def import_attribute_definitions(self):
        """Import AttributeDefinition records with updated PicklistId values."""
        print("\n" + "=" * 60)
        print("IMPORTING ATTRIBUTEDEFINITION RECORDS")
        print("=" * 60)
        
        # Read AttributeDefinition sheet
        df = pd.read_excel(self.excel_file, sheet_name='09_AttributeDefinition')
        df = df.dropna(how='all')
        
        # Clean column names
        df.columns = df.columns.str.replace('*', '', regex=False)
        
        # Remove empty rows
        df = df[df['Name'].notna()]
        
        print(f"\nUpserting {len(df)} AttributeDefinition records...")
        
        # Save to CSV
        csv_file = Path('data/temp_attribute_definition.csv')
        df.to_csv(csv_file, index=False)
        
        # Upsert records
        cmd = [
            'sf', 'data', 'upsert', 'bulk',
            '--sobject', 'AttributeDefinition',
            '--file', str(csv_file),
            '--external-id', 'Code',
            '--target-org', self.target_org,
            '--wait', '10'
        ]
        
        result = subprocess.run(cmd, capture_output=True, text=True)
        
        if result.returncode == 0:
            print("✓ Successfully imported AttributeDefinition records")
        else:
            print(f"✗ Failed to import: {result.stderr}")
        
        # Clean up
        csv_file.unlink()

def main():
    importer = AttributePicklistImporter()
    importer.insert_picklists()

if __name__ == '__main__':
    main()