#!/usr/bin/env python3
"""
Comprehensive field cleanup - Remove ALL read-only, system-managed, and non-updateable fields
from ALL sheets in the Revenue Cloud upload template.
"""

import pandas as pd
import subprocess
import json
from openpyxl import load_workbook
from pathlib import Path
from datetime import datetime

class ComprehensiveFieldCleanup:
    def __init__(self):
        self.workbook_path = Path('data/Revenue_Cloud_Complete_Upload_Template.xlsx')
        self.target_org = 'fortradp2'
        self.cleanup_report = []
        
        # Define objects and their sheets
        self.object_sheet_mapping = {
            'ProductCatalog': '11_ProductCatalog',
            'ProductCategory': '12_ProductCategory',
            'ProductClassification': '08_ProductClassification',
            'AttributeCategory': '10_AttributeCategory',
            'AttributePicklist': '14_AttributePicklist',
            'AttributeDefinition': '09_AttributeDefinition',
            'ProductSellingModel': '15_ProductSellingModel',
            'Product2': '13_Product2',
            'Pricebook2': '19_Pricebook2',
            'ProductAttributeDefinition': '17_ProductAttributeDef',
            'AttributePicklistValue': '18_AttributePicklistValue',
            'ProductCategoryProduct': '26_ProductCategoryProduct',
            'PricebookEntry': '20_PricebookEntry',
            'ProductRelatedComponent': '25_ProductRelatedComponent',
            'ProductComponentGroup': '14_ProductComponentGroup'
        }
        
        # Fields that should ALWAYS be removed (even if marked as updateable)
        self.always_remove_fields = {
            'LastModifiedDate', 'CreatedDate', 'SystemModstamp', 
            'LastModifiedById', 'CreatedById', 'IsDeleted',
            'LastViewedDate', 'LastReferencedDate', 'LastActivityDate',
            'CurrencyIsoCode', 'RecordTypeId', 'OwnerId'
        }
        
        # Special handling for specific objects
        self.object_specific_rules = {
            'Product2': {
                'remove_if_exists': ['Type', 'TransferRecordMode'],  # Type cannot be updated after creation
                'keep_for_insert_only': ['Type']  # But needed for new records
            },
            'ProductAttributeDefinition': {
                'remove_if_exists': ['AttributeCategoryId'],  # Can't be updated after creation
                'keep_for_insert_only': ['AttributeCategoryId', 'ProductClassificationAttributeId']
            },
            'ProductCategoryProduct': {
                'keep_only': ['Id', 'ProductCategoryId', 'ProductId']  # Junction object - very limited
            },
            'ProductComponentGroup': {
                'required_for_upsert': ['Code']  # External ID field
            }
        }
    
    def get_field_metadata(self, object_name):
        """Get comprehensive field metadata for an object."""
        cmd = [
            'sf', 'sobject', 'describe',
            '--sobject', object_name,
            '--target-org', self.target_org,
            '--json'
        ]
        
        result = subprocess.run(cmd, capture_output=True, text=True)
        
        if result.returncode == 0:
            data = json.loads(result.stdout)
            if 'result' in data and 'fields' in data['result']:
                return data['result']['fields']
        
        return []
    
    def analyze_field(self, field):
        """Analyze if a field should be kept or removed."""
        field_name = field['name']
        
        # Always remove certain fields (except Id which might be used as external ID)
        if field_name in self.always_remove_fields and field_name != 'Id':
            return False, "System field - always removed"
        
        # Check if field is read-only
        createable = field.get('createable', False)
        updateable = field.get('updateable', False)
        
        # Special case: Keep Id field for upsert operations
        if field_name == 'Id':
            return True, "Id field kept for upsert operations"
        
        # If field is neither createable nor updateable, remove it
        if not createable and not updateable:
            return False, "Not createable or updateable"
        
        # If field is calculated or auto-generated
        if field.get('calculated', False) or field.get('autoNumber', False):
            return False, "Calculated or auto-number field"
        
        # If field is a system field based on naming patterns
        if field_name.endswith('__pc') or field_name.startswith('System'):
            return False, "System field pattern"
        
        # Check data type - some types are always read-only
        if field['type'] in ['address', 'location']:
            if not updateable:
                return False, f"{field['type']} field not updateable"
        
        # Default: keep the field
        return True, "Field is valid for updates"
    
    def clean_sheet(self, object_name, sheet_name):
        """Clean a single sheet by removing non-updateable fields."""
        print(f"\nProcessing {object_name} ({sheet_name})...")
        self.cleanup_report.append(f"\n{object_name} ({sheet_name})")
        self.cleanup_report.append("-" * 60)
        
        # Get field metadata
        fields = self.get_field_metadata(object_name)
        if not fields:
            print(f"  ⚠️  Could not get field metadata")
            self.cleanup_report.append("  ERROR: Could not get field metadata")
            return
        
        # Create field analysis
        field_analysis = {}
        for field in fields:
            keep, reason = self.analyze_field(field)
            field_analysis[field['name']] = {
                'keep': keep,
                'reason': reason,
                'createable': field.get('createable', False),
                'updateable': field.get('updateable', False)
            }
        
        # Apply object-specific rules
        if object_name in self.object_specific_rules:
            rules = self.object_specific_rules[object_name]
            
            # Remove specific fields
            for field_name in rules.get('remove_if_exists', []):
                if field_name in field_analysis:
                    field_analysis[field_name]['keep'] = False
                    field_analysis[field_name]['reason'] = "Object-specific rule: remove"
            
            # Handle keep_only rules
            if 'keep_only' in rules:
                keep_only_fields = set(rules['keep_only'])
                for field_name in field_analysis:
                    if field_name not in keep_only_fields:
                        field_analysis[field_name]['keep'] = False
                        field_analysis[field_name]['reason'] = "Object-specific rule: not in keep_only list"
        
        # Load workbook
        wb = load_workbook(self.workbook_path)
        if sheet_name not in wb.sheetnames:
            print(f"  ⚠️  Sheet not found")
            self.cleanup_report.append("  ERROR: Sheet not found")
            return
        
        ws = wb[sheet_name]
        
        # Get current columns
        headers = []
        header_positions = {}
        for col, cell in enumerate(ws[1], 1):
            if cell.value:
                clean_header = str(cell.value).replace('*', '').strip()
                headers.append(clean_header)
                header_positions[clean_header] = col
        
        # Identify columns to remove
        columns_to_remove = []
        for header in headers:
            if header not in field_analysis:
                # Field doesn't exist in object
                columns_to_remove.append((header, "Field not found in object"))
            elif not field_analysis[header]['keep']:
                # Field should be removed
                columns_to_remove.append((header, field_analysis[header]['reason']))
        
        # Remove columns (in reverse order to maintain indices)
        removed_count = 0
        for header, reason in columns_to_remove:
            if header in header_positions:
                col_index = header_positions[header]
                # Adjust index for previously removed columns
                adjusted_index = col_index - removed_count
                
                # Only remove if column still exists
                if adjusted_index > 0 and adjusted_index <= ws.max_column:
                    ws.delete_cols(adjusted_index)
                    removed_count += 1
                    
                    print(f"  ✗ Removed: {header} ({reason})")
                    self.cleanup_report.append(f"  - Removed: {header} ({reason})")
                    
                    # Update positions of remaining columns
                    for h, pos in header_positions.items():
                        if pos > col_index:
                            header_positions[h] = pos - 1
        
        # Report kept fields
        kept_fields = [h for h in headers if h not in [c[0] for c in columns_to_remove]]
        print(f"  ✓ Kept {len(kept_fields)} fields: {', '.join(kept_fields[:5])}{'...' if len(kept_fields) > 5 else ''}")
        self.cleanup_report.append(f"  ✓ Kept {len(kept_fields)} fields")
        
        # Save workbook
        wb.save(self.workbook_path)
    
    def create_backup(self):
        """Create a backup of the current workbook."""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_path = self.workbook_path.parent / f"{self.workbook_path.stem}_pre_cleanup_{timestamp}.xlsx"
        
        import shutil
        shutil.copy2(self.workbook_path, backup_path)
        
        print(f"✓ Created backup: {backup_path}")
        return backup_path
    
    def run_cleanup(self):
        """Run the comprehensive cleanup process."""
        print("=" * 80)
        print("COMPREHENSIVE FIELD CLEANUP")
        print("=" * 80)
        
        # Create backup
        backup_path = self.create_backup()
        self.cleanup_report.append(f"Backup created: {backup_path}")
        
        # Process each object
        for object_name, sheet_name in self.object_sheet_mapping.items():
            self.clean_sheet(object_name, sheet_name)
        
        # Save cleanup report
        report_path = 'field_cleanup_report.txt'
        with open(report_path, 'w') as f:
            f.write("FIELD CLEANUP REPORT\n")
            f.write("=" * 80 + "\n")
            f.write(f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write("\n".join(self.cleanup_report))
        
        print("\n" + "=" * 80)
        print("CLEANUP COMPLETE")
        print("=" * 80)
        print(f"✓ Backup saved: {backup_path}")
        print(f"✓ Report saved: {report_path}")
        print("✓ Workbook cleaned: data/Revenue_Cloud_Complete_Upload_Template.xlsx")
        
        # Run export to ensure synchronization
        print("\nRunning export to synchronize with Salesforce...")
        subprocess.run(['python3', 'export_to_same_template.py'], capture_output=True)
        print("✓ Export complete")

def main():
    cleanup = ComprehensiveFieldCleanup()
    cleanup.run_cleanup()

if __name__ == '__main__':
    main()