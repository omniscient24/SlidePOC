#\!/usr/bin/env python3
"""
Add missing important fields to ProductRelatedComponent sheet.
"""

from openpyxl import load_workbook

def main():
    print("=" * 60)
    print("ADDING MISSING FIELDS TO PRODUCTRELATEDCOMPONENT")
    print("=" * 60)
    
    # Load workbook
    wb = load_workbook('data/Revenue_Cloud_Complete_Upload_Template.xlsx')
    ws = wb['25_ProductRelatedComponent']
    
    # Get current headers
    current_headers = []
    for cell in ws[1]:
        if cell.value:
            current_headers.append(cell.value.replace('*', '').strip())
    
    print(f"Current headers: {len(current_headers)}")
    
    # Important fields that should be added if missing
    fields_to_add = [
        'ProductRelationshipTypeId',  # Required for insert
        'IsDefaultComponent',         # Important for configurable bundles
        'IsQuantityEditable'          # Controls quantity changes
    ]
    
    # Add missing fields
    next_col = len(current_headers) + 1
    added_fields = []
    
    for field in fields_to_add:
        if field not in current_headers:
            ws.cell(row=1, column=next_col, value=field)
            added_fields.append(field)
            next_col += 1
    
    if added_fields:
        print(f"\nAdded {len(added_fields)} missing fields:")
        for field in added_fields:
            print(f"  - {field}")
    else:
        print("\nNo missing fields to add")
    
    # Save workbook
    wb.save('data/Revenue_Cloud_Complete_Upload_Template.xlsx')
    
    # Show final headers
    print("\nFinal column structure:")
    final_headers = []
    for i, cell in enumerate(ws[1], 1):
        if cell.value:
            final_headers.append(cell.value)
            print(f"  {i}. {cell.value}")
    
    print(f"\nTotal columns: {len(final_headers)}")
    print("\n✓ ProductRelatedComponent sheet is now properly configured")

if __name__ == '__main__':
    main()
