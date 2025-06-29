#\!/usr/bin/env python3
"""
Remove all read-only fields from all object sheets in the workbook.
"""

import subprocess
import json
from openpyxl import load_workbook
from pathlib import Path

def get_field_info(object_name):
    """Get field information for an object."""
    cmd = [
        'sf', 'sobject', 'describe',
        '--sobject', object_name,
        '--target-org', 'fortradp2',
        '--json'
    ]
    
    result = subprocess.run(cmd, capture_output=True, text=True)
    
    if result.returncode == 0:
        data = json.loads(result.stdout)
        fields = {}
        if 'result' in data and 'fields' in data['result']:
            for field in data['result']['fields']:
                fields[field['name']] = {
                    'createable': field.get('createable', False),
                    'updateable': field.get('updateable', False)
                }
        return fields
    return {}

def main():
    print("=" * 60)
    print("REMOVING ALL READ-ONLY FIELDS FROM ALL SHEETS")
    print("=" * 60)
    
    # Object name mappings (sheet name to Salesforce object name)
    object_mappings = {
        'ProductAttributeDef': 'ProductAttributeDefinition',
        'AttributeBasedAdj': 'AttributeBasedAdjustment',
        'AttributeBasedAdjRule': 'AttributeBasedAdjustmentRule'
    }
    
    # Load workbook
    wb = load_workbook('data/Revenue_Cloud_Complete_Upload_Template.xlsx')
    
    # Process each object sheet
    total_removed = 0
    sheets_processed = 0
    
    for sheet_name in wb.sheetnames:
        # Skip non-object sheets
        if not (len(sheet_name) > 2 and sheet_name[0:2].isdigit() and '_' in sheet_name):
            continue
            
        # Extract object name
        parts = sheet_name.split('_', 1)
        if len(parts) < 2:
            continue
            
        sheet_object = parts[1]
        
        # Get actual Salesforce object name
        sf_object = object_mappings.get(sheet_object, sheet_object)
        
        print(f"\nProcessing {sheet_name} ({sf_object})...")
        
        # Get field information
        field_info = get_field_info(sf_object)
        
        if not field_info:
            print(f"  ⚠️  Could not get field info for {sf_object}")
            continue
        
        # Get worksheet
        ws = wb[sheet_name]
        
        # Get headers and identify columns to remove
        headers = []
        header_positions = {}
        for col, cell in enumerate(ws[1], 1):
            if cell.value:
                header_name = str(cell.value).replace('*', '').strip()
                headers.append(header_name)
                header_positions[header_name] = col
        
        # Identify read-only fields
        cols_to_remove = []
        fields_to_remove = []
        
        for header in headers:
            # Special handling for Id field - keep it for upserts
            if header == 'Id':
                continue
                
            # Remove relationship fields (contain dots)
            if '.' in header:
                cols_to_remove.append(header_positions[header])
                fields_to_remove.append(header)
                continue
            
            # Check if field exists and is writable
            if header in field_info:
                field = field_info[header]
                # Remove if not createable AND not updateable
                if not field['createable'] and not field['updateable']:
                    cols_to_remove.append(header_positions[header])
                    fields_to_remove.append(header)
            else:
                # Field doesn't exist in object
                cols_to_remove.append(header_positions[header])
                fields_to_remove.append(header)
        
        # Delete columns in reverse order
        if cols_to_remove:
            cols_to_remove.sort(reverse=True)
            
            print(f"  Removing {len(cols_to_remove)} read-only/invalid fields:")
            for field in fields_to_remove:
                print(f"    - {field}")
            
            for col_idx in cols_to_remove:
                ws.delete_cols(col_idx)
            
            total_removed += len(cols_to_remove)
        else:
            print(f"  ✓ No read-only fields to remove")
        
        sheets_processed += 1
    
    # Save workbook
    wb.save('data/Revenue_Cloud_Complete_Upload_Template.xlsx')
    
    print("\n" + "=" * 60)
    print("SUMMARY")
    print("=" * 60)
    print(f"✓ Processed {sheets_processed} object sheets")
    print(f"✓ Removed {total_removed} read-only/invalid fields total")
    print("✓ Workbook saved with all read-only fields removed")

if __name__ == '__main__':
    main()
