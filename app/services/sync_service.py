"""
Salesforce Data Sync Service
Handles downloading data from Salesforce and updating the workbook
"""
import subprocess
import json
import pandas as pd
import os
from datetime import datetime
from typing import Dict, List, Tuple, Optional
from pathlib import Path

from config.settings.app_config import CLI_COMMAND, DATA_ROOT

class SyncService:
    """Service for syncing data between Salesforce and local workbook"""
    
    def __init__(self):
        self.workbook_path = DATA_ROOT / 'Revenue_Cloud_Complete_Upload_Template_FINAL.xlsx'
        self.sync_log_dir = DATA_ROOT / 'sync-logs'
        self.sync_log_dir.mkdir(exist_ok=True)
        
        # Object to sheet mapping
        self.object_sheet_mapping = {
            'ProductCatalog': '11_ProductCatalog',
            'ProductCategory': '12_ProductCategory',
            'Product2': '13_Product2',
            'ProductClassification': '08_ProductClassification',
            'AttributeDefinition': '09_AttributeDefinition',
            'AttributeCategory': '10_AttributeCategory',
            'AttributePicklist': '14_AttributePicklist',
            'AttributePicklistValue': '18_AttributePicklistValue',
            'ProductAttributeDefinition': '17_ProductAttributeDef',
            'Pricebook2': '19_Pricebook2',
            'PricebookEntry': '20_PricebookEntry',
            'CostBook': '01_CostBook',
            'CostBookEntry': '15_CostBookEntry',
            'PriceAdjustmentSchedule': '21_PriceAdjustmentSchedule',
            'PriceAdjustmentTier': '22_PriceAdjustmentTier',
            'LegalEntity': '02_LegalEntity',
            'TaxEngine': '03_TaxEngine',
            'TaxPolicy': '04_TaxPolicy',
            'TaxTreatment': '05_TaxTreatment',
            'BillingPolicy': '06_BillingPolicy',
            'BillingTreatment': '07_BillingTreatment',
            'ProductSellingModel': '15_ProductSellingModel',
            'ProductSellingModelOption': '16_ProductSellingModelOption',
            'ProductComponentGroup': '14_ProductComponentGroup',
            'ProductRelatedComponent': '25_ProductRelatedComponent',
            'ProductCategoryProduct': '26_ProductCategoryProduct',
            'AttributeBasedAdjRule': '23_AttributeBasedAdjRule',
            'AttributeBasedAdjustment': '24_AttributeBasedAdj',
            # Transaction objects
            'Order': '27_Order',
            'OrderItem': '28_OrderItem',
            'Asset': '29_Asset',
            'AssetAction': '30_AssetAction',
            'AssetActionSource': '31_AssetActionSource',
            'Contract': '32_Contract'
        }
    
    def sync_objects(self, object_names: List[str], connection_alias: str) -> Tuple[bool, Dict]:
        """
        Sync multiple objects from Salesforce to workbook
        
        Args:
            object_names: List of object API names to sync
            connection_alias: Salesforce CLI alias for the connection
            
        Returns:
            Tuple of (success, result_dict)
        """
        sync_results = {
            'success': True,
            'synced': [],
            'failed': [],
            'messages': []
        }
        
        for object_name in object_names:
            try:
                print(f"[SYNC] Starting sync for {object_name}")
                success, result = self.sync_single_object(object_name, connection_alias)
                
                if success:
                    sync_results['synced'].append({
                        'object': object_name,
                        'recordCount': result.get('recordCount', 0),
                        'message': result.get('message', 'Sync completed')
                    })
                else:
                    sync_results['failed'].append({
                        'object': object_name,
                        'error': result.get('error', 'Unknown error')
                    })
                    
            except Exception as e:
                print(f"[ERROR] Failed to sync {object_name}: {e}")
                sync_results['failed'].append({
                    'object': object_name,
                    'error': str(e)
                })
        
        # Update overall success status
        sync_results['success'] = len(sync_results['failed']) == 0
        
        # Save sync log
        self._save_sync_log(sync_results)
        
        return sync_results['success'], sync_results
    
    def sync_single_object(self, object_name: str, connection_alias: str) -> Tuple[bool, Dict]:
        """
        Sync a single object from Salesforce to workbook
        
        Args:
            object_name: Object API name
            connection_alias: Salesforce CLI alias
            
        Returns:
            Tuple of (success, result_dict)
        """
        try:
            # Check if object has a corresponding sheet
            sheet_name = self.object_sheet_mapping.get(object_name)
            if not sheet_name:
                return False, {'error': f'No sheet mapping found for {object_name}'}
            
            # Download data from Salesforce
            print(f"[SYNC] Downloading {object_name} records from Salesforce...")
            records = self._download_from_salesforce(object_name, connection_alias)
            
            if records is None:
                return False, {'error': 'Failed to download records from Salesforce'}
            
            print(f"[SYNC] Downloaded {len(records)} records")
            
            # Update workbook with the records
            print(f"[SYNC] Updating workbook sheet {sheet_name}...")
            success = self._update_workbook_sheet(sheet_name, records, object_name)
            
            if success:
                return True, {
                    'recordCount': len(records),
                    'message': f'Successfully synced {len(records)} {object_name} records'
                }
            else:
                return False, {'error': 'Failed to update workbook'}
                
        except Exception as e:
            print(f"[ERROR] Exception in sync_single_object: {e}")
            import traceback
            traceback.print_exc()
            return False, {'error': str(e)}
    
    def _download_from_salesforce(self, object_name: str, connection_alias: str) -> Optional[List[Dict]]:
        """
        Download records from Salesforce using CLI
        
        Args:
            object_name: Salesforce object API name
            connection_alias: CLI connection alias
            
        Returns:
            List of records or None if failed
        """
        try:
            # First, get the actual fields that exist in the org
            existing_fields = self._get_existing_fields(object_name, connection_alias)
            if not existing_fields:
                print(f"[ERROR] Could not retrieve fields for {object_name}")
                return None
            
            # Build the query - always include Id as first field
            fields = self._get_object_fields(object_name)
            if not fields:
                # If we can't get fields, at least get Id and Name
                fields = ['Id', 'Name']
            elif 'Id' not in fields:
                fields = ['Id'] + fields
            
            # Filter to only include fields that exist in the org
            valid_fields = []
            for field in fields:
                if field in existing_fields:
                    valid_fields.append(field)
                else:
                    print(f"[WARN] Field '{field}' does not exist in {object_name}, skipping")
            
            # Ensure Id is first
            if 'Id' in valid_fields and valid_fields[0] != 'Id':
                valid_fields.remove('Id')
                valid_fields = ['Id'] + valid_fields
            elif 'Id' not in valid_fields:
                valid_fields = ['Id'] + valid_fields
            
            query = f"SELECT {', '.join(valid_fields)} FROM {object_name} LIMIT 1000"
            
            # Execute query using CLI
            cmd = [
                CLI_COMMAND, 'data', 'query',
                '--query', query,
                '--target-org', connection_alias,
                '--json'
            ]
            
            print(f"[SYNC] Executing query: {query}")
            print(f"[SYNC] CLI command: {' '.join(cmd)}")
            
            result = subprocess.run(cmd, capture_output=True, text=True, timeout=300)
            
            if result.returncode != 0:
                print(f"[ERROR] Query failed with return code {result.returncode}")
                print(f"[ERROR] STDERR: {result.stderr}")
                print(f"[ERROR] STDOUT: {result.stdout}")
                
                # Log to file for debugging
                import datetime
                error_log = f"""
[{datetime.datetime.now()}] Salesforce CLI Error
Command: {' '.join(cmd)}
Return Code: {result.returncode}
STDERR: {result.stderr}
STDOUT: {result.stdout}
"""
                with open('sync_error.log', 'a') as f:
                    f.write(error_log)
                    
                return None
            
            # Parse the JSON response
            if not result.stdout:
                print(f"[ERROR] No output from CLI command")
                return None
                
            try:
                data = json.loads(result.stdout)
            except json.JSONDecodeError as e:
                print(f"[ERROR] Failed to parse JSON response: {e}")
                print(f"[ERROR] Raw output: {result.stdout}")
                print(f"[ERROR] Raw error: {result.stderr}")
                return None
                
            if data.get('status') != 0:
                print(f"[ERROR] Query returned error status: {data}")
                return None
            
            # Extract records
            records = data.get('result', {}).get('records', [])
            
            # Clean up records (remove attributes field)
            cleaned_records = []
            for record in records:
                cleaned_record = {k: v for k, v in record.items() if k != 'attributes'}
                cleaned_records.append(cleaned_record)
            
            return cleaned_records
            
        except subprocess.TimeoutExpired:
            print(f"[ERROR] Query timeout for {object_name}")
            return None
        except Exception as e:
            print(f"[ERROR] Failed to download from Salesforce: {e}")
            return None
    
    def _get_existing_fields(self, object_name: str, connection_alias: str) -> List[str]:
        """
        Get the actual fields that exist in the Salesforce org for an object
        
        Args:
            object_name: Salesforce object API name
            connection_alias: CLI connection alias
            
        Returns:
            List of field names that exist in the org
        """
        try:
            # Use sf sobject describe to get actual fields
            cmd = [
                CLI_COMMAND, 'sobject', 'describe',
                '--sobject', object_name,
                '--target-org', connection_alias,
                '--json'
            ]
            
            result = subprocess.run(cmd, capture_output=True, text=True, timeout=60)
            
            if result.returncode != 0:
                print(f"[ERROR] Failed to describe {object_name}: {result.stderr}")
                return []
            
            try:
                data = json.loads(result.stdout)
                if data.get('status') == 0:
                    fields = data.get('result', {}).get('fields', [])
                    return [field['name'] for field in fields]
            except json.JSONDecodeError:
                print(f"[ERROR] Failed to parse describe response")
                
            return []
            
        except Exception as e:
            print(f"[ERROR] Failed to get existing fields: {e}")
            return []
    
    def _get_object_fields(self, object_name: str) -> List[str]:
        """
        Get relevant fields for an object based on workbook columns
        
        Args:
            object_name: Salesforce object API name
            
        Returns:
            List of field names
        """
        try:
            sheet_name = self.object_sheet_mapping.get(object_name)
            if not sheet_name or not self.workbook_path.exists():
                return []
            
            # Read the sheet to get column names
            df = pd.read_excel(self.workbook_path, sheet_name=sheet_name, nrows=0)
            fields = df.columns.tolist()
            
            # Ensure Id is included
            if 'Id' not in fields:
                fields = ['Id'] + fields
            
            # Filter out any non-API field names and clean field names
            cleaned_fields = []
            for field in fields:
                # Skip empty or whitespace-only fields
                if field and str(field).strip():
                    # Remove any trailing/leading whitespace
                    cleaned_field = str(field).strip()
                    # Only include valid Salesforce field names (alphanumeric + underscore)
                    if cleaned_field.replace('_', '').replace('.', '').isalnum():
                        cleaned_fields.append(cleaned_field)
            
            # Ensure we have at least Id
            if not cleaned_fields or (len(cleaned_fields) == 1 and cleaned_fields[0] == 'Id'):
                # Default to common Product2 fields if no valid fields found
                if object_name == 'Product2':
                    return ['Id', 'Name', 'ProductCode', 'IsActive', 'Description']
                else:
                    return ['Id', 'Name']
            
            return cleaned_fields
            
        except Exception as e:
            print(f"[ERROR] Failed to get fields for {object_name}: {e}")
            return []
    
    def _update_workbook_sheet(self, sheet_name: str, records: List[Dict], object_name: str) -> bool:
        """
        Update a specific sheet in the workbook with new records
        
        Args:
            sheet_name: Name of the sheet to update
            records: List of records to write
            object_name: Object API name for logging
            
        Returns:
            True if successful, False otherwise
        """
        try:
            if not records:
                print(f"[SYNC] No records to update for {sheet_name}")
                return True
            
            # Convert records to DataFrame
            df_new = pd.DataFrame(records)
            
            # Ensure Id is the first column
            if 'Id' in df_new.columns and df_new.columns[0] != 'Id':
                cols = ['Id'] + [col for col in df_new.columns if col != 'Id']
                df_new = df_new[cols]
            
            # Read all sheets from workbook
            with pd.ExcelFile(self.workbook_path) as xl_file:
                sheet_dict = {}
                for sheet in xl_file.sheet_names:
                    sheet_dict[sheet] = pd.read_excel(xl_file, sheet_name=sheet)
            
            # Update the specific sheet
            sheet_dict[sheet_name] = df_new
            
            # Write back all sheets
            with pd.ExcelWriter(self.workbook_path, engine='openpyxl') as writer:
                for sheet, df in sheet_dict.items():
                    df.to_excel(writer, sheet_name=sheet, index=False)
            
            # Apply formatting cleanup
            self._cleanup_workbook_formatting()
            
            print(f"[SYNC] Successfully updated {sheet_name} with {len(records)} records")
            return True
            
        except Exception as e:
            print(f"[ERROR] Failed to update workbook sheet {sheet_name}: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def _cleanup_workbook_formatting(self):
        """
        Clean up workbook formatting after sync:
        - Apply consistent formatting from header row to all data rows
        - Remove completely empty rows
        - Ensure table formatting is consistent
        """
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
            from openpyxl.utils import get_column_letter
            
            print("[SYNC] Cleaning up workbook formatting...")
            
            # Load workbook with openpyxl for formatting
            wb = load_workbook(self.workbook_path)
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                
                # Turn off gridlines for all sheets
                sheet.sheet_view.showGridLines = False
                
                # Apply formatting even to sheets with only headers
                if sheet.max_row >= 1:
                    # Apply black background with white text to header row and left alignment to all
                    black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
                    white_font = Font(color="FFFFFF", bold=True)
                    left_align = Alignment(horizontal='left', vertical='center')
                    
                    # Format header row
                    for col in range(1, sheet.max_column + 1):
                        header_cell = sheet.cell(row=1, column=col)
                        header_cell.fill = black_fill
                        header_cell.font = white_font
                        header_cell.alignment = left_align
                
                # Skip further processing for sheets with no data rows
                if sheet.max_row <= 1:
                    continue
                
                # The variables are already defined above, just need to apply to data rows
                # Apply left alignment to all data cells
                for row in range(2, sheet.max_row + 1):
                    for col in range(1, sheet.max_column + 1):
                        cell = sheet.cell(row=row, column=col)
                        cell.alignment = left_align
                
                # Get header row formatting (row 1) and column formatting from row 2
                header_formats = {}
                column_formats = {}
                
                for col in range(1, sheet.max_column + 1):
                    header_cell = sheet.cell(row=1, column=col)
                    data_cell = sheet.cell(row=2, column=col) if sheet.max_row > 1 else None
                    
                    header_formats[col] = {
                        'font': white_font,  # Use the white font we defined
                        'fill': black_fill,  # Use the black fill we defined
                        'border': Border(
                            left=Side(style=header_cell.border.left.style) if header_cell.border.left.style else None,
                            right=Side(style=header_cell.border.right.style) if header_cell.border.right.style else None,
                            top=Side(style=header_cell.border.top.style) if header_cell.border.top.style else None,
                            bottom=Side(style=header_cell.border.bottom.style) if header_cell.border.bottom.style else None
                        ) if header_cell.border else None,
                        'alignment': left_align  # Use consistent left alignment
                    }
                    
                    # Get number format from data row
                    if data_cell:
                        column_formats[col] = {
                            'number_format': data_cell.number_format
                        }
                
                # Identify and remove empty rows (from bottom up to avoid index issues)
                rows_to_delete = []
                for row in range(sheet.max_row, 1, -1):  # Start from bottom, skip header
                    is_empty = True
                    for col in range(1, sheet.max_column + 1):
                        if sheet.cell(row=row, column=col).value is not None:
                            is_empty = False
                            break
                    if is_empty:
                        rows_to_delete.append(row)
                
                # Delete empty rows
                for row in rows_to_delete:
                    sheet.delete_rows(row)
                    print(f"[SYNC] Removed empty row {row} from {sheet_name}")
                
                # Apply consistent formatting to all data rows
                if sheet.max_row > 1:
                    # Determine if there's a pattern (alternating row colors, etc.)
                    row2_fill = sheet.cell(row=2, column=1).fill if sheet.max_row > 1 else None
                    row3_fill = sheet.cell(row=3, column=1).fill if sheet.max_row > 2 else None
                    
                    has_alternating_rows = (row2_fill and row3_fill and 
                                          row2_fill.start_color != row3_fill.start_color)
                    
                    # Check if this sheet has table formatting
                    has_table_format = False
                    for table in sheet.tables:
                        has_table_format = True
                        # Extend table range if needed
                        table_range = table.ref
                        table_start, table_end = table_range.split(':')
                        table_end_col = ''.join(c for c in table_end if c.isalpha())
                        new_range = f"{table_start}:{table_end_col}{sheet.max_row}"
                        if new_range != table.ref:
                            table.ref = new_range
                            print(f"[SYNC] Extended table '{table.name}' range to {new_range}")
                    
                    # Apply formatting to all data rows
                    for row in range(2, sheet.max_row + 1):
                        for col in range(1, sheet.max_column + 1):
                            cell = sheet.cell(row=row, column=col)
                            
                            # Apply border from header
                            if header_formats[col]['border']:
                                cell.border = header_formats[col]['border']
                            
                            # Apply left alignment (already done above, but ensure it's consistent)
                            cell.alignment = left_align
                            
                            # Apply number format from column
                            if col in column_formats and column_formats[col]['number_format']:
                                cell.number_format = column_formats[col]['number_format']
                            
                            # Apply alternating row colors if pattern exists and not in a table
                            if not has_table_format:
                                if has_alternating_rows:
                                    if row % 2 == 0:
                                        cell.fill = row2_fill
                                    else:
                                        cell.fill = row3_fill
                                elif row2_fill and row2_fill.fill_type:
                                    # Apply consistent fill from row 2
                                    cell.fill = row2_fill
                
                # Auto-adjust column widths based on content
                for col in range(1, sheet.max_column + 1):
                    max_length = 0
                    column_letter = get_column_letter(col)
                    
                    for row in range(1, min(sheet.max_row + 1, 100)):  # Check first 100 rows
                        cell = sheet.cell(row=row, column=col)
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    
                    # Set width with some padding
                    adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                    sheet.column_dimensions[column_letter].width = adjusted_width
            
            # Save the formatted workbook
            wb.save(self.workbook_path)
            print("[SYNC] Workbook formatting cleanup completed")
            
        except Exception as e:
            print(f"[ERROR] Failed to clean up workbook formatting: {e}")
            # Don't fail the sync if formatting cleanup fails
            
    def _save_sync_log(self, sync_results: Dict):
        """Save sync results to a log file"""
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            log_file = self.sync_log_dir / f'sync_log_{timestamp}.json'
            
            with open(log_file, 'w') as f:
                json.dump({
                    'timestamp': datetime.now().isoformat(),
                    'results': sync_results
                }, f, indent=2)
                
            print(f"[SYNC] Sync log saved to {log_file}")
            
        except Exception as e:
            print(f"[ERROR] Failed to save sync log: {e}")

# Singleton instance
sync_service = SyncService()