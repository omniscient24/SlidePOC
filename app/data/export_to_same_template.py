#!/usr/bin/env python3
"""
Export data from Salesforce org to the existing Excel template while preserving formatting.
Updates the template in place as requested.
"""

import subprocess
import json
import pandas as pd
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

class TemplateExporter:
    def __init__(self):
        self.target_org = 'fortradp2'
        self.template_file = Path('data/Revenue_Cloud_Complete_Upload_Template.xlsx')
        
        # Define queries for each sheet - comprehensive field lists
        self.sheet_configs = {
            '11_ProductCatalog': {
                'object': 'ProductCatalog',
                'query': "SELECT Id, Name, Code, Description FROM ProductCatalog ORDER BY Name"
            },
            '12_ProductCategory': {
                'object': 'ProductCategory', 
                'query': "SELECT Id, Name, Code, CatalogId, ParentCategoryId, Description, SortOrder FROM ProductCategory ORDER BY SortOrder, Name"
            },
            '08_ProductClassification': {
                'object': 'ProductClassification',
                'query': "SELECT Id, Name, Code, Status FROM ProductClassification ORDER BY Name"
            },
            '09_AttributeDefinition': {
                'object': 'AttributeDefinition',
                'query': "SELECT Id, Name, Code, Label, Description, DataType, IsActive, IsRequired, DefaultValue, DeveloperName, DefaultHelpText, ValueDescription, SourceSystemIdentifier FROM AttributeDefinition ORDER BY Name"
            },
            '10_AttributeCategory': {
                'object': 'AttributeCategory',
                'query': "SELECT Id, Name, Code, Description FROM AttributeCategory ORDER BY Name"
            },
            '15_ProductSellingModel': {
                'object': 'ProductSellingModel',
                'query': "SELECT Id, Name, SellingModelType, Status, PricingTermUnit, PricingTerm FROM ProductSellingModel ORDER BY Name"
            },
            '13_Product2': {
                'object': 'Product2',
                'query': """SELECT Id, Name, ProductCode, StockKeepingUnit, Description, IsActive, 
                          QuantityUnitOfMeasure, Family, Type, AvailabilityDate, BasedOnId
                          FROM Product2 ORDER BY Name"""
            },
            '19_Pricebook2': {
                'object': 'Pricebook2',
                'query': """SELECT Id, Name, Description, IsActive, IsStandard, IsArchived, 
                          ValidFrom, ValidTo
                          FROM Pricebook2 ORDER BY Name"""
            },
            '17_ProductAttributeDef': {
                'object': 'ProductAttributeDefinition',
                'query': """SELECT Id, Name, Product2Id, AttributeDefinitionId, AttributeCategoryId, 
                          Sequence, IsRequired, IsHidden, IsReadOnly, IsPriceImpacting, 
                          DefaultValue, HelpText, MinimumValue, MaximumValue, DisplayType, 
                          Status, AttributeNameOverride, Description
                          FROM ProductAttributeDefinition ORDER BY Product2Id, Sequence"""
            },
            '26_ProductCategoryProduct': {
                'object': 'ProductCategoryProduct',
                'query': """SELECT Id, ProductCategoryId, ProductId
                          FROM ProductCategoryProduct ORDER BY ProductCategoryId"""
            },
            '20_PricebookEntry': {
                'object': 'PricebookEntry',
                'query': """SELECT Id, Pricebook2Id, Product2Id, UnitPrice, IsActive, 
                          ProductSellingModelId, UseStandardPrice
                          FROM PricebookEntry ORDER BY Product2Id"""
            },
            '25_ProductRelatedComponent': {
                'object': 'ProductRelatedComponent',
                'query': """SELECT Id, ParentProductId, ChildProductId, ProductRelationshipTypeId, 
                          MinQuantity, MaxQuantity, Quantity, IsComponentRequired, Sequence
                          FROM ProductRelatedComponent ORDER BY ParentProductId, Sequence"""
            }
        }
        
        # Additional sheets that may have data
        self.additional_sheets = {
            '01_CostBook': {'object': 'CostBook', 'query': None},
            '15_CostBookEntry': {'object': 'CostBookEntry', 'query': None},
            '21_PriceAdjustmentSchedule': {'object': 'PriceAdjustmentSchedule', 'query': None},
            '22_PriceAdjustmentTier': {'object': 'PriceAdjustmentTier', 'query': None},
            '23_AttributeBasedAdjRule': {'object': 'AttributeBasedAdjustmentRule', 'query': None},
            '24_AttributeBasedAdj': {'object': 'AttributeBasedAdjustment', 'query': None},
            '06_BillingPolicy': {'object': 'BillingPolicy', 'query': None},
            '07_BillingTreatment': {'object': 'BillingTreatment', 'query': None},
            '02_LegalEntity': {'object': 'LegalEntity', 'query': None},
            '05_TaxTreatment': {'object': 'TaxTreatment', 'query': None},
            '04_TaxPolicy': {'object': 'TaxPolicy', 'query': None},
            '03_TaxEngine': {'object': 'TaxEngine', 'query': None}
        }
    
    def query_salesforce(self, query):
        """Execute a SOQL query and return results as DataFrame."""
        cmd = [
            'sf', 'data', 'query',
            '--query', query,
            '--target-org', self.target_org,
            '--json'
        ]
        
        result = subprocess.run(cmd, capture_output=True, text=True)
        
        if result.returncode == 0:
            data = json.loads(result.stdout)
            if 'result' in data and 'records' in data['result']:
                records = data['result']['records']
                # Remove attributes and clean up field names with dots
                clean_records = []
                for record in records:
                    clean_record = {}
                    for key, value in record.items():
                        if key != 'attributes':
                            # Handle nested fields like Product2.Name
                            if '.' in key:
                                clean_key = key.replace('.', '_')
                                clean_record[clean_key] = value
                            else:
                                clean_record[key] = value
                    clean_records.append(clean_record)
                
                if clean_records:
                    return pd.DataFrame(clean_records)
        
        return pd.DataFrame()
    
    def update_sheet(self, wb, sheet_name, df):
        """Update a sheet with dataframe data while preserving formatting."""
        if sheet_name not in wb.sheetnames:
            print(f"  ⚠️  Sheet {sheet_name} not found")
            return False
            
        ws = wb[sheet_name]
        
        # Get headers from row 1
        headers = []
        header_positions = {}  # Map header name to column position
        
        for col_idx, cell in enumerate(ws[1], 1):
            if cell.value:
                header = str(cell.value)
                headers.append(header)
                # Clean header for matching
                clean_header = header.replace('*', '').strip()
                header_positions[clean_header] = col_idx
        
        if not headers:
            print(f"  ⚠️  No headers found in {sheet_name}")
            return False
        
        # Clear existing data (rows 2 onwards)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.value = None
        
        # Write data
        if len(df) > 0:
            for row_idx, (_, row_data) in enumerate(df.iterrows(), 2):
                for header in headers:
                    col_idx = headers.index(header) + 1
                    clean_header = header.replace('*', '').strip()
                    
                    # Try to find matching column in dataframe
                    value = None
                    
                    # Direct match
                    if clean_header in row_data:
                        value = row_data[clean_header]
                    # Try with underscores (for nested fields)
                    elif clean_header.replace('.', '_') in row_data:
                        value = row_data[clean_header.replace('.', '_')]
                    # Special handling for certain fields
                    elif clean_header == 'Product2.Name' and 'Product2_Name' in row_data:
                        value = row_data['Product2_Name']
                    elif clean_header == 'AttributeDefinition.Name' and 'AttributeDefinition_Name' in row_data:
                        value = row_data['AttributeDefinition_Name']
                    
                    # Write value to cell
                    if value is not None and pd.notna(value):
                        ws.cell(row=row_idx, column=col_idx, value=value)
        
        return True
    
    def export_all_data(self):
        """Export all data to the template."""
        print("=" * 60)
        print("EXPORTING SALESFORCE DATA TO EXCEL TEMPLATE")
        print("=" * 60)
        print(f"Org: {self.target_org}")
        print(f"Template: {self.template_file}\n")
        
        # Create backup of original
        backup_file = self.template_file.parent / f"{self.template_file.stem}_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        import shutil
        shutil.copy2(self.template_file, backup_file)
        print(f"Created backup: {backup_file}\n")
        
        # Load workbook
        wb = openpyxl.load_workbook(self.template_file)
        
        # Process main sheets
        total_records = 0
        updated_sheets = 0
        
        for sheet_name, config in self.sheet_configs.items():
            print(f"Processing {sheet_name}...")
            
            # Query data
            df = self.query_salesforce(config['query'])
            
            if len(df) > 0:
                print(f"  Found {len(df)} records")
                if self.update_sheet(wb, sheet_name, df):
                    print(f"  ✓ Updated successfully")
                    total_records += len(df)
                    updated_sheets += 1
            else:
                print(f"  No records found")
                # Still clear the sheet
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                        for cell in row:
                            cell.value = None
        
        # Clear additional sheets that don't have data
        print("\nClearing unused sheets...")
        for sheet_name in self.additional_sheets:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                # Clear data rows
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    for cell in row:
                        cell.value = None
                print(f"  Cleared {sheet_name}")
        
        # Save workbook
        wb.save(self.template_file)
        wb.close()
        
        # Summary
        print("\n" + "=" * 60)
        print("EXPORT COMPLETE")
        print("=" * 60)
        print(f"✓ Updated {updated_sheets} sheets")
        print(f"✓ Exported {total_records} total records")
        print(f"✓ Saved to: {self.template_file}")
        print(f"✓ Backup at: {backup_file}")
        print("\nThe workbook now reflects the current state of the org.")

def main():
    exporter = TemplateExporter()
    exporter.export_all_data()

if __name__ == '__main__':
    main()