#!/usr/bin/env python3
"""
Export data from Salesforce to template preserving original formatting.
Uses a simpler approach by copying the template and updating values in place.
"""

import subprocess
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
import io
import csv

class SimpleFormattedExporter:
    def __init__(self):
        self.template_file = Path('data/Revenue_Cloud_Complete_Upload_Template.xlsx')
        self.output_file = Path(f'data/Revenue_Cloud_Export_Formatted_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
        self.target_org = 'fortradp2'
        
        # Core queries
        self.core_queries = {
            'ProductCatalog': "SELECT Id, Name, Code FROM ProductCatalog",
            'ProductCategory': "SELECT Id, Name, Code, CatalogId, ParentCategoryId FROM ProductCategory",
            'Product2': "SELECT Id, Name, ProductCode, Description, Family, Type, BasedOnId, IsActive FROM Product2",
            'AttributeDefinition': "SELECT Id, Name, Code, Label FROM AttributeDefinition",
            'ProductClassification': "SELECT Id, Name, Code, Status FROM ProductClassification",
            'ProductClassificationAttr': "SELECT Id, Name, ProductClassificationId, AttributeDefinitionId, Status FROM ProductClassificationAttr",
            'Pricebook2': "SELECT Id, Name, Description, IsActive, IsStandard FROM Pricebook2",
            'PricebookEntry': "SELECT Id, Product2Id, Pricebook2Id, UnitPrice, IsActive FROM PricebookEntry WHERE IsActive = true",
            'ProductCategoryProduct': "SELECT Id, ProductCategoryId, ProductId FROM ProductCategoryProduct",
            'ProductAttributeDefinition': "SELECT Id, Name, Product2Id, ProductClassificationAttributeId, AttributeDefinitionId, Status FROM ProductAttributeDefinition",
            'ProductSellingModel': "SELECT Id, Name, SellingModelType, Status, PricingTermUnit, PricingTerm FROM ProductSellingModel",
            'ProductRelatedComponent': "SELECT Id, ParentProductId, ChildProductId, ProductRelationshipTypeId, Quantity, IsDefaultComponent, Sequence FROM ProductRelatedComponent"
        }
        
        # Sheet mappings
        self.sheet_mappings = {
            '11_ProductCatalog': 'ProductCatalog',
            '12_ProductCategory': 'ProductCategory',
            '13_Product2': 'Product2',
            '09_AttributeDefinition': 'AttributeDefinition',
            '08_ProductClassification': 'ProductClassification',
            '19_Pricebook2': 'Pricebook2',
            '20_PricebookEntry': 'PricebookEntry',
            '26_ProductCategoryProduct': 'ProductCategoryProduct',
            '17_ProductAttributeDef': 'ProductAttributeDefinition',
            '15_ProductSellingModel': 'ProductSellingModel',
            '25_ProductRelatedComponent': 'ProductRelatedComponent'
        }
    
    def query_salesforce(self, query):
        """Query Salesforce and return results."""
        try:
            cmd = [
                'sf', 'data', 'query',
                '--query', query,
                '--target-org', self.target_org,
                '--result-format', 'csv'
            ]
            
            result = subprocess.run(cmd, capture_output=True, text=True)
            
            if result.returncode == 0:
                lines = result.stdout.strip().split('\n')
                if len(lines) > 1:
                    data_lines = [line for line in lines if not line.startswith('Total number of records') and line.strip()]
                    if data_lines:
                        csv_data = '\n'.join(data_lines)
                        reader = csv.DictReader(io.StringIO(csv_data))
                        return list(reader)
            else:
                print(f"Error: {result.stderr}")
        except Exception as e:
            print(f"Exception: {str(e)}")
        return []
    
    def export_formatted(self):
        """Export data while preserving template formatting."""
        print(f"Exporting from: {self.target_org}")
        print(f"Using template: {self.template_file}")
        print(f"Output file: {self.output_file}\n")
        
        # Copy template to output
        import shutil
        shutil.copy2(self.template_file, self.output_file)
        
        # Load workbook
        wb = load_workbook(self.output_file)
        
        # Query all data
        print("Querying Salesforce data...")
        all_data = {}
        for obj_name, query in self.core_queries.items():
            print(f"  {obj_name}...", end='')
            data = self.query_salesforce(query)
            all_data[obj_name] = data
            print(f" {len(data)} records")
        
        print("\nUpdating sheets...")
        
        # Process each sheet
        for sheet_name in wb.sheetnames:
            if sheet_name in ['Instructions', 'Picklist Values']:
                continue
                
            if sheet_name in self.sheet_mappings:
                object_name = self.sheet_mappings[sheet_name]
                if object_name in all_data and all_data[object_name]:
                    ws = wb[sheet_name]
                    data = all_data[object_name]
                    
                    print(f"  {sheet_name}: {len(data)} records")
                    
                    # Get headers from row 1
                    headers = []
                    header_map = {}  # Map column index to field name
                    for col_idx, cell in enumerate(ws[1], 1):
                        if cell.value:
                            headers.append(cell.value)
                            # Clean header for mapping
                            clean_header = str(cell.value).replace('*', '').strip()
                            header_map[col_idx] = clean_header
                    
                    # Clear data rows but keep formatting
                    # Start from row 2, clear values only
                    max_row = ws.max_row
                    for row_num in range(2, max_row + 1):
                        for col_num in range(1, len(headers) + 1):
                            cell = ws.cell(row=row_num, column=col_num)
                            cell.value = None
                    
                    # Populate with new data
                    for row_idx, record in enumerate(data, 2):
                        if row_idx > max_row:
                            # Add new row if needed (will copy formatting from row above)
                            ws.insert_rows(row_idx)
                        
                        for col_idx in range(1, len(headers) + 1):
                            cell = ws.cell(row=row_idx, column=col_idx)
                            header = header_map.get(col_idx, '')
                            
                            # Map data to cell
                            if header in record:
                                value = record[header]
                                # Handle boolean values
                                if value.lower() in ['true', 'false']:
                                    cell.value = value.lower() == 'true'
                                # Handle numeric values
                                elif header in ['UnitPrice', 'Quantity', 'Sequence']:
                                    try:
                                        cell.value = float(value) if '.' in value else int(value)
                                    except:
                                        cell.value = value
                                else:
                                    cell.value = value
                            elif headers[col_idx-1] == 'External_ID__c':
                                cell.value = f"{sheet_name}_{row_idx-1}"
                            elif headers[col_idx-1] == 'Name' and 'Id' in record:
                                # Use ID if no name
                                cell.value = record.get('Name', record['Id'])
        
        # Save workbook
        wb.save(self.output_file)
        wb.close()
        
        print(f"\nExport completed!")
        print(f"File saved: {self.output_file}")
        print("\nThe file is ready for editing and upsert operations.")

def main():
    exporter = SimpleFormattedExporter()
    exporter.export_formatted()

if __name__ == '__main__':
    main()