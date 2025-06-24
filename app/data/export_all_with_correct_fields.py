#!/usr/bin/env python3
"""
Export all Revenue Cloud objects with correct field names.
"""

import subprocess
import json
import pandas as pd
from pathlib import Path
import openpyxl
from datetime import datetime

class CompleteOrgExporter:
    def __init__(self):
        self.workbook_path = Path('data/Revenue_Cloud_Complete_Upload_Template.xlsx')
        self.target_org = 'fortradp2'
        
    def get_available_fields(self, object_name):
        """Get list of available fields for an object."""
        cmd = [
            'sf', 'sobject', 'describe',
            '--sobject', object_name,
            '--target-org', self.target_org,
            '--json'
        ]
        
        result = subprocess.run(cmd, capture_output=True, text=True)
        
        if result.returncode == 0:
            data = json.loads(result.stdout)
            if 'result' in data and 'fields' in data['result']:
                fields = data['result']['fields']
                # Return field names that are accessible
                return [f['name'] for f in fields if not f['name'].startswith('System')]
        return []
    
    def query_all_fields(self, object_name, sheet_headers):
        """Query object with fields that match sheet headers."""
        # Get available fields
        available_fields = self.get_available_fields(object_name)
        
        # Clean sheet headers (remove asterisks)
        clean_headers = [h.replace('*', '').strip() for h in sheet_headers if h]
        
        # Find fields that exist in both sheet and object
        fields_to_query = []
        for header in clean_headers:
            if header in available_fields:
                fields_to_query.append(header)
            # Also check for common variations
            elif header == 'Product2.Name' and object_name == 'PricebookEntry':
                # Special handling for relationship fields
                continue
            elif header == 'Product2.ProductCode' and object_name == 'PricebookEntry':
                continue
        
        if not fields_to_query:
            return []
        
        # Always include Id if available
        if 'Id' in available_fields and 'Id' not in fields_to_query:
            fields_to_query.insert(0, 'Id')
        
        # Build query
        query = f"SELECT {', '.join(fields_to_query)} FROM {object_name}"
        
        # Add ORDER BY
        if 'Name' in fields_to_query:
            query += " ORDER BY Name"
        elif 'Code' in fields_to_query:
            query += " ORDER BY Code"
        else:
            query += " ORDER BY Id"
        
        # Execute query
        cmd = [
            'sf', 'data', 'query',
            '--query', query,
            '--target-org', self.target_org,
            '--json'
        ]
        
        result = subprocess.run(cmd, capture_output=True, text=True)
        
        if result.returncode == 0:
            data = json.loads(result.stdout)
            if 'result' in data and 'records' in data['result']:
                records = data['result']['records']
                # Remove attributes field
                for record in records:
                    if 'attributes' in record:
                        del record['attributes']
                return records
        
        return []
    
    def export_to_excel(self):
        """Export all data to Excel workbook."""
        print("=" * 70)
        print("EXPORTING ALL REVENUE CLOUD DATA TO EXCEL")
        print("=" * 70)
        print(f"Target workbook: {self.workbook_path}")
        print(f"Source org: {self.target_org}")
        print(f"Export time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print()
        
        # Load workbook
        wb = openpyxl.load_workbook(self.workbook_path)
        
        # Track statistics
        total_records = 0
        sheets_updated = 0
        
        # Object name mappings
        object_mappings = {
            '11_ProductCatalog': 'ProductCatalog',
            '12_ProductCategory': 'ProductCategory',
            '08_ProductClassification': 'ProductClassification',
            '09_AttributeDefinition': 'AttributeDefinition',
            '10_AttributeCategory': 'AttributeCategory',
            '14_AttributePicklist': 'AttributePicklist',
            '15_ProductSellingModel': 'ProductSellingModel',
            '13_Product2': 'Product2',
            '17_ProductAttributeDef': 'ProductAttributeDefinition',
            '18_AttributePicklistValue': 'AttributePicklistValue',
            '19_Pricebook2': 'Pricebook2',
            '20_PricebookEntry': 'PricebookEntry',
            '26_ProductCategoryProduct': 'ProductCategoryProduct',
            '25_ProductRelatedComponent': 'ProductRelatedComponent'
        }
        
        # Process each sheet
        for sheet_name, object_name in object_mappings.items():
            if sheet_name not in wb.sheetnames:
                continue
                
            print(f"\nProcessing {sheet_name} ({object_name})...")
            
            ws = wb[sheet_name]
            
            # Get headers from sheet
            headers = []
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=1, column=col).value
                if cell_value:
                    headers.append(cell_value)
            
            if not headers:
                print(f"  ⚠️  No headers found")
                continue
            
            # Query data using actual fields
            records = self.query_all_fields(object_name, headers)
            
            if records:
                print(f"  Found {len(records)} records")
                
                # Clear existing data (preserve headers)
                for row in range(2, ws.max_row + 1):
                    for col in range(1, len(headers) + 1):
                        ws.cell(row=row, column=col).value = None
                
                # Create header mapping
                header_to_col = {}
                for col_idx, header in enumerate(headers, 1):
                    clean_header = header.replace('*', '').strip()
                    header_to_col[clean_header] = col_idx
                
                # Write data
                for row_idx, record in enumerate(records, 2):
                    for field_name, value in record.items():
                        if field_name in header_to_col:
                            col_idx = header_to_col[field_name]
                            ws.cell(row=row_idx, column=col_idx).value = value
                
                print(f"  ✓ Exported {len(records)} records")
                total_records += len(records)
                sheets_updated += 1
            else:
                print(f"  No records found")
                
                # Clear any existing data
                for row in range(2, ws.max_row + 1):
                    for col in range(1, len(headers) + 1):
                        ws.cell(row=row, column=col).value = None
        
        # Update Product references
        self.update_product_references(wb)
        
        # Save workbook
        wb.save(self.workbook_path)
        wb.close()
        
        print("\n" + "=" * 70)
        print("EXPORT COMPLETE")
        print("=" * 70)
        print(f"✓ Sheets updated: {sheets_updated}")
        print(f"✓ Total records exported: {total_records}")
        print(f"✓ Workbook saved: {self.workbook_path}")
        
    def update_product_references(self, wb):
        """Update sheets with Product2 references."""
        print("\n" + "-" * 50)
        print("Updating Product references...")
        
        # Get Product2 data
        product_map = {}
        if '13_Product2' in wb.sheetnames:
            ws = wb['13_Product2']
            
            # Find columns
            id_col = name_col = code_col = None
            for col in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col).value
                if header:
                    clean = header.replace('*', '').strip()
                    if clean == 'Id':
                        id_col = col
                    elif clean == 'Name':
                        name_col = col
                    elif clean == 'ProductCode':
                        code_col = col
            
            # Build product map
            if id_col:
                for row in range(2, ws.max_row + 1):
                    prod_id = ws.cell(row=row, column=id_col).value
                    if prod_id:
                        product_map[prod_id] = {
                            'Name': ws.cell(row=row, column=name_col).value if name_col else None,
                            'ProductCode': ws.cell(row=row, column=code_col).value if code_col else None
                        }
        
        # Update PricebookEntry
        if '20_PricebookEntry' in wb.sheetnames and product_map:
            ws = wb['20_PricebookEntry']
            self.add_product_data_to_sheet(ws, product_map, 'Product2Id')
            print("  ✓ Updated PricebookEntry")
        
        # Update ProductAttributeDef
        if '17_ProductAttributeDef' in wb.sheetnames and product_map:
            ws = wb['17_ProductAttributeDef']
            self.add_product_data_to_sheet(ws, product_map, 'ProductId')
            print("  ✓ Updated ProductAttributeDef")
    
    def add_product_data_to_sheet(self, ws, product_map, id_field):
        """Add product name and code to a sheet."""
        # Find columns
        cols = {}
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header:
                clean = header.replace('*', '').strip()
                if clean == id_field:
                    cols['id'] = col
                elif clean == 'Product2.Name':
                    cols['name'] = col
                elif clean == 'Product2.ProductCode':
                    cols['code'] = col
        
        # Update rows
        if 'id' in cols:
            for row in range(2, ws.max_row + 1):
                prod_id = ws.cell(row=row, column=cols['id']).value
                if prod_id and prod_id in product_map:
                    if 'name' in cols and product_map[prod_id]['Name']:
                        ws.cell(row=row, column=cols['name']).value = product_map[prod_id]['Name']
                    if 'code' in cols and product_map[prod_id]['ProductCode']:
                        ws.cell(row=row, column=cols['code']).value = product_map[prod_id]['ProductCode']

def main():
    exporter = CompleteOrgExporter()
    exporter.export_to_excel()

if __name__ == '__main__':
    main()