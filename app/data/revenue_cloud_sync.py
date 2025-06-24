#!/usr/bin/env python3
"""
Revenue Cloud Sync Tool
Synchronizes data from Salesforce to local Excel workbooks
"""

import argparse
import json
import os
import subprocess
import sys
from datetime import datetime
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
import time

# Revenue Cloud object mappings
OBJECT_MAPPINGS = {
    # Core Product Objects
    'ProductCatalog': {
        'api_name': 'ProductCatalog',
        'sheet_name': '11_ProductCatalog',
        'fields': ['Id', 'Name', 'Code', 'Description', 'EffectiveStartDate', 'EffectiveEndDate', 'CatalogType']
    },
    'ProductCategory': {
        'api_name': 'ProductCategory',
        'sheet_name': '12_ProductCategory',
        'fields': ['Id', 'Name', 'Code', 'Description', 'CatalogId', 'ParentCategoryId', 'NumberOfProducts']
    },
    'Product2': {
        'api_name': 'Product2',
        'sheet_name': '13_Product2',
        'fields': ['Id', 'Name', 'ProductCode', 'Description', 'IsActive', 'Family']
    },
    'ProductCategoryProduct': {
        'api_name': 'ProductCategoryProduct',
        'sheet_name': '26_ProductCategoryProduct',
        'fields': ['Id', 'Name', 'ProductCategoryId', 'ProductId', 'CatalogId']
    },
    'ProductSellingModel': {
        'api_name': 'ProductSellingModel',
        'sheet_name': '15_ProductSellingModel',
        'fields': ['Id', 'Name', 'Status']
    },
    'ProductComponentGroup': {
        'api_name': 'ProductComponentGroup',
        'sheet_name': '14_ProductComponentGroup',
        'fields': ['Id', 'Name', 'Code', 'Description', 'ParentProductId', 'ParentGroupId']
    },
    'ProductRelatedComponent': {
        'api_name': 'ProductRelatedComponent',
        'sheet_name': '25_ProductRelatedComponent',
        'fields': ['Id', 'Name', 'ChildProductId', 'ParentProductId', 'ProductComponentGroupId']
    },
    'ProductClassification': {
        'api_name': 'ProductClassification',
        'sheet_name': '08_ProductClassification',
        'fields': ['Id', 'Name', 'Code', 'Status']
    },
    
    # Attribute Objects
    'AttributeDefinition': {
        'api_name': 'AttributeDefinition',
        'sheet_name': '09_AttributeDefinition',
        'fields': ['Id', 'Name', 'Code', 'Description', 'IsActive', 'DeveloperName']
    },
    'AttributePicklist': {
        'api_name': 'AttributePicklist',
        'sheet_name': '14_AttributePicklist',
        'fields': ['Id', 'Name', 'Description', 'Status']
    },
    'AttributePicklistValue': {
        'api_name': 'AttributePicklistValue',
        'sheet_name': '18_AttributePicklistValue',
        'fields': ['Id', 'Name', 'Code', 'PicklistId', 'Status']
    },
    'ProductAttributeDefinition': {
        'api_name': 'ProductAttributeDefinition',
        'sheet_name': '17_ProductAttributeDef',
        'fields': ['Id', 'Name', 'Description', 'Product2Id', 'AttributeDefinitionId', 'Status']
    },
    'AttributeCategory': {
        'api_name': 'AttributeCategory',
        'sheet_name': '10_AttributeCategory',
        'fields': ['Id', 'Name', 'Code', 'Description']
    },
    
    # Pricing Objects
    'Pricebook2': {
        'api_name': 'Pricebook2',
        'sheet_name': '19_Pricebook2',
        'fields': ['Id', 'Name', 'Description', 'IsActive', 'IsStandard']
    },
    'PricebookEntry': {
        'api_name': 'PricebookEntry',
        'sheet_name': '20_PricebookEntry',
        'fields': ['Id', 'Product2Id', 'Pricebook2Id', 'UnitPrice', 'IsActive']
    },
    'CostBook': {
        'api_name': 'CostBook',
        'sheet_name': '01_CostBook',
        'fields': ['Id', 'Name']
    },
    'CostBookEntry': {
        'api_name': 'CostBookEntry',
        'sheet_name': '15_CostBookEntry',
        'fields': ['Id', 'Name', 'CostBookId', 'ProductId', 'Description']
    },
    'PriceAdjustmentSchedule': {
        'api_name': 'PriceAdjustmentSchedule',
        'sheet_name': '21_PriceAdjustmentSchedule',
        'fields': ['Id', 'Name', 'Description', 'IsActive', 'Pricebook2Id']
    },
    'PriceAdjustmentTier': {
        'api_name': 'PriceAdjustmentTier',
        'sheet_name': '22_PriceAdjustmentTier',
        'fields': ['Id', 'Name', 'PriceAdjustmentScheduleId', 'Product2Id']
    },
    
    # Tax & Legal Objects
    'LegalEntity': {
        'api_name': 'LegalEntity',
        'sheet_name': '02_LegalEntity',
        'fields': ['Id', 'Name', 'CompanyName', 'Description', 'Status']
    },
    'TaxTreatment': {
        'api_name': 'TaxTreatment',
        'sheet_name': '05_TaxTreatment',
        'fields': ['Id', 'Name', 'Description', 'Status', 'TaxCode', 'ProductCode']
    },
    'TaxPolicy': {
        'api_name': 'TaxPolicy',
        'sheet_name': '04_TaxPolicy',
        'fields': ['Id', 'Name', 'Description', 'Status']
    },
    'TaxEngine': {
        'api_name': 'TaxEngine',
        'sheet_name': '03_TaxEngine',
        'fields': ['Id', 'TaxEngineName', 'Description', 'Status', 'SellerCode']
    }
}

def query_salesforce_data(org, object_name, fields):
    """Query Salesforce for object data"""
    try:
        # Build SOQL query
        field_list = ', '.join(fields)
        query = f"SELECT {field_list} FROM {object_name}"
        
        # Use Salesforce CLI to execute query
        cmd = [
            'sf', 'data', 'query',
            '--query', query,
            '--target-org', org,
            '--json'
        ]
        
        print(f"  Querying {object_name}...")
        # Run command, redirecting stderr to devnull to avoid warning messages
        with open(os.devnull, 'w') as devnull:
            result = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=devnull, text=True)
        
        if result.returncode != 0:
            # Re-run to get error message if failed
            result2 = subprocess.run(cmd, capture_output=True, text=True)
            print(f"  ⚠️  Error querying {object_name}: {result2.stderr}")
            return None
            
        # Parse JSON response - handle potential warnings in stderr
        try:
            # The actual JSON output should be in stdout
            response = json.loads(result.stdout)
        except json.JSONDecodeError as e:
            print(f"  ⚠️  Error parsing response for {object_name}: {str(e)}")
            print(f"  Response: {result.stdout[:200]}...")
            return None
        
        if response.get('status') == 0:
            records = response.get('result', {}).get('records', [])
            print(f"  ✓ Retrieved {len(records)} records from {object_name}")
            return records
        else:
            print(f"  ⚠️  Query failed: {response.get('message', 'Unknown error')}")
            return None
            
    except Exception as e:
        print(f"  ⚠️  Exception querying {object_name}: {str(e)}")
        return None

def update_excel_sheet(workbook_path, sheet_name, records, field_mapping):
    """Update Excel sheet with Salesforce data"""
    try:
        # Load workbook
        wb = load_workbook(workbook_path)
        
        if sheet_name not in wb.sheetnames:
            print(f"  ⚠️  Sheet {sheet_name} not found in workbook")
            return False
            
        # Get the sheet
        ws = wb[sheet_name]
        
        # Get existing headers from row 1
        headers = []
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header:
                headers.append(header)
        
        # Create DataFrame from records
        if records:
            df = pd.DataFrame(records)
            
            # Clear existing data (keep headers)
            for row in range(2, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).value = None
            
            # Write new data
            row_num = 2
            for _, record in df.iterrows():
                for col_num, header in enumerate(headers, 1):
                    # Map Salesforce field names to Excel headers
                    sf_field = None
                    for field in field_mapping:
                        if field.lower() == header.lower() or field.replace('_', '').lower() == header.replace(' ', '').lower():
                            sf_field = field
                            break
                    
                    if sf_field and sf_field in record:
                        ws.cell(row=row_num, column=col_num).value = record[sf_field]
                
                row_num += 1
            
            # Save workbook
            wb.save(workbook_path)
            print(f"  ✓ Updated {sheet_name} with {len(records)} records")
            return True
        else:
            print(f"  ℹ️  No records to update for {sheet_name}")
            return True
            
    except Exception as e:
        print(f"  ⚠️  Error updating sheet {sheet_name}: {str(e)}")
        return False

def write_progress(progress_file, status):
    """Write progress to a file for real-time monitoring"""
    with open(progress_file, 'w') as f:
        json.dump(status, f)

def sync_all_objects(org, workbook_path, objects_to_sync=None, progress_file=None):
    """Sync all or specified objects from Salesforce to Excel"""
    
    print(f"\n🔄 Starting Salesforce sync...")
    print(f"Org: {org}")
    print(f"Workbook: {workbook_path}")
    print(f"Timestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
    
    # Initial progress
    if progress_file:
        write_progress(progress_file, {
            'status': 'initializing',
            'current_object': None,
            'completed': 0,
            'total': 0,
            'percent': 0,
            'message': 'Creating backup...'
        })
    
    # Create backup
    backup_path = create_backup(workbook_path)
    print(f"✓ Created backup: {backup_path}\n")
    
    # Determine which objects to sync
    if objects_to_sync:
        sync_list = {k: v for k, v in OBJECT_MAPPINGS.items() if k in objects_to_sync}
    else:
        sync_list = OBJECT_MAPPINGS
    
    total_objects = len(sync_list)
    success_count = 0
    error_count = 0
    total_records = 0
    completed_objects = 0
    
    # Update progress with total count
    if progress_file:
        write_progress(progress_file, {
            'status': 'syncing',
            'current_object': None,
            'completed': 0,
            'total': total_objects,
            'percent': 0,
            'message': f'Starting sync of {total_objects} objects...'
        })
    
    # Process each object
    for object_key, mapping in sync_list.items():
        print(f"📊 Syncing {object_key}...")
        
        # Update progress
        if progress_file:
            write_progress(progress_file, {
                'status': 'syncing',
                'current_object': object_key,
                'completed': completed_objects,
                'total': total_objects,
                'percent': int((completed_objects / total_objects) * 100),
                'message': f'Querying {object_key} from Salesforce...'
            })
        
        # Query Salesforce
        records = query_salesforce_data(org, mapping['api_name'], mapping['fields'])
        
        if records is not None:
            # Update progress - writing to Excel
            if progress_file:
                write_progress(progress_file, {
                    'status': 'syncing',
                    'current_object': object_key,
                    'completed': completed_objects,
                    'total': total_objects,
                    'percent': int((completed_objects / total_objects) * 100),
                    'message': f'Writing {len(records)} records for {object_key}...'
                })
            
            # Update Excel
            if update_excel_sheet(workbook_path, mapping['sheet_name'], records, mapping['fields']):
                success_count += 1
                total_records += len(records)
            else:
                error_count += 1
        else:
            error_count += 1
        
        completed_objects += 1
        
        # Small delay to avoid rate limits
        time.sleep(0.5)
    
    # Final progress
    if progress_file:
        write_progress(progress_file, {
            'status': 'completed',
            'current_object': None,
            'completed': total_objects,
            'total': total_objects,
            'percent': 100,
            'message': f'Sync completed! {success_count} objects synced, {total_records} total records.'
        })
    
    # Summary
    print(f"\n📈 Sync Summary:")
    print(f"  ✓ Successfully synced: {success_count} objects")
    print(f"  ⚠️  Errors: {error_count} objects")
    print(f"  📊 Total records: {total_records}")
    print(f"  🕒 Completed: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    return {
        'success': error_count == 0,
        'success_count': success_count,
        'error_count': error_count,
        'total_records': total_records,
        'backup_path': str(backup_path)
    }

def create_backup(workbook_path):
    """Create a timestamped backup of the workbook"""
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    backup_dir = Path(workbook_path).parent / 'backups'
    backup_dir.mkdir(exist_ok=True)
    
    backup_name = f"{Path(workbook_path).stem}_sync_backup_{timestamp}.xlsx"
    backup_path = backup_dir / backup_name
    
    # Copy file
    import shutil
    shutil.copy2(workbook_path, backup_path)
    
    return backup_path

def main():
    parser = argparse.ArgumentParser(description='Sync Revenue Cloud data from Salesforce')
    parser.add_argument('--org', required=True, help='Salesforce org alias')
    parser.add_argument('--workbook', required=True, help='Path to Excel workbook')
    parser.add_argument('--objects', nargs='+', help='Specific objects to sync (default: all)')
    parser.add_argument('--output-json', help='Output results as JSON to file')
    parser.add_argument('--progress-file', help='Write progress updates to this file')
    
    args = parser.parse_args()
    
    # Validate workbook exists
    if not os.path.exists(args.workbook):
        print(f"Error: Workbook not found: {args.workbook}")
        sys.exit(1)
    
    # Execute sync
    result = sync_all_objects(args.org, args.workbook, args.objects, args.progress_file)
    
    # Output JSON if requested
    if args.output_json:
        with open(args.output_json, 'w') as f:
            json.dump(result, f, indent=2)
    
    # Exit with appropriate code
    sys.exit(0 if result['success'] else 1)

if __name__ == "__main__":
    main()