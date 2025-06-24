#!/usr/bin/env python3
"""
Export data from Salesforce org to the existing Excel template while preserving formatting.
"""

import subprocess
import json
import pandas as pd
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

class FormattedExporter:
    def __init__(self):
        self.target_org = 'fortradp2'
        self.template_file = Path('data/Revenue_Cloud_Complete_Upload_Template.xlsx')
        self.output_file = self.template_file  # Same file name as requested
        
        # Define queries for each sheet
        self.sheet_queries = {
            '11_ProductCatalog': {
                'object': 'ProductCatalog',
                'query': "SELECT Id, Name, Code, Description FROM ProductCatalog ORDER BY Name"
            },
            '12_ProductCategory': {
                'object': 'ProductCategory', 
                'query': "SELECT Id, Name, Code, CatalogId, ParentCategoryId, Description, SortOrder FROM ProductCategory ORDER BY Name"
            },
            '08_ProductClassification': {
                'object': 'ProductClassification',
                'query': "SELECT Id, Name, Code, Type, Status, Description FROM ProductClassification ORDER BY Name"
            },
            '09_AttributeDefinition': {
                'object': 'AttributeDefinition',
                'query': "SELECT Id, Name, Code, Label, Description, DataType, IsActive, IsRequired, DefaultValue, DeveloperName, DefaultHelpText, ValueDescription, SourceSystemIdentifier FROM AttributeDefinition ORDER BY Name"
            },
            '10_AttributeCategory': {
                'object': 'AttributeCategory',
                'query': "SELECT Id, Name, Code, DisplaySequence, Description FROM AttributeCategory ORDER BY Name"
            },
            '15_ProductSellingModel': {
                'object': 'ProductSellingModel',
                'query': "SELECT Id, Name, SellingModelType, Status, Description, PricingTermUnit, PricingTerm FROM ProductSellingModel ORDER BY Name"
            },
            '13_Product2': {
                'object': 'Product2',
                'query': "SELECT Id, Name, ProductCode, StockKeepingUnit, Description, IsActive, QuantityUnitOfMeasure, Family, Type, AvailabilityDate, BasedOnId, ProductClassId, TaxPolicyId, BillingPolicyId, RecordTypeId, UnitOfMeasureId, VatId, LegalEntityId, ProductSellingModelId FROM Product2 ORDER BY Name"
            },
            '19_Pricebook2': {
                'object': 'Pricebook2',
                'query': "SELECT Id, Name, Description, IsActive, IsStandard, IsArchived, ValidFrom, ValidTo FROM Pricebook2 WHERE IsStandard = false ORDER BY Name"
            },
            '17_ProductAttributeDef': {
                'object': 'ProductAttributeDefinition',
                'query': "SELECT Id, Name, Product2Id, AttributeDefinitionId, AttributeCategoryId, Sequence, IsRequired, IsHidden, IsReadOnly, IsPriceImpacting, DefaultValue, HelpText, MinimumValue, MaximumValue, DisplayType, Status, AttributeNameOverride, Description FROM ProductAttributeDefinition ORDER BY Product2Id, Sequence"
            },
            '26_ProductCategoryProduct': {
                'object': 'ProductCategoryProduct',
                'query': "SELECT Id, ProductCategoryId, ProductId, IsPrimaryCategory FROM ProductCategoryProduct ORDER BY ProductCategoryId"
            },
            '20_PricebookEntry': {
                'object': 'PricebookEntry',
                'query': "SELECT Id, Pricebook2Id, Product2Id, UnitPrice, IsActive, ProductSellingModelId, UseStandardPrice FROM PricebookEntry WHERE Pricebook2.IsStandard = false ORDER BY Product2Id"
            },
            '25_ProductRelatedComponent': {
                'object': 'ProductRelatedComponent',
                'query': "SELECT Id, ParentProductId, ChildProductId, ProductRelationshipTypeId, MinQuantity, MaxQuantity, Quantity, IsComponentRequired, Sequence FROM ProductRelatedComponent ORDER BY ParentProductId, Sequence"
            }
        }
    
    def query_salesforce(self, query):
        """Execute a SOQL query and return results as DataFrame."""
        cmd = [
            'sf', 'data', 'query',
            '--query', query,
            '--target-org', self.target_org,
            '--json'
        ]
        
        result = subprocess.run(cmd, capture_output=True, text=True)
        
        if result.returncode == 0:
            data = json.loads(result.stdout)
            if 'result' in data and 'records' in data['result']:
                records = data['result']['records']
                # Remove attributes field from each record
                for record in records:
                    if 'attributes' in record:
                        del record['attributes']
                
                if records:
                    return pd.DataFrame(records)
        
        return pd.DataFrame()
    
    def copy_formatting(self, source_ws, target_ws, df):
        """Copy formatting from source worksheet to target."""
        # Copy column widths
        for idx, col in enumerate(source_ws.iter_cols(min_row=1, max_row=1), 1):
            col_letter = openpyxl.utils.get_column_letter(idx)
            if idx <= len(df.columns) + 1:  # +1 for potential index column
                target_ws.column_dimensions[col_letter].width = source_ws.column_dimensions[col_letter].width
        
        # Copy row heights
        for idx, row in enumerate(source_ws.iter_rows(min_row=1, max_row=min(100, len(df) + 2)), 1):
            target_ws.row_dimensions[idx].height = source_ws.row_dimensions[idx].height
        
        # Copy header formatting
        for idx, cell in enumerate(source_ws[1], 1):
            target_cell = target_ws.cell(row=1, column=idx)
            if cell.fill and cell.fill.fgColor:
                target_cell.fill = PatternFill(
                    start_color=cell.fill.fgColor.rgb if cell.fill.fgColor.rgb else "00000000",
                    end_color=cell.fill.fgColor.rgb if cell.fill.fgColor.rgb else "00000000",
                    fill_type="solid"
                )
            if cell.font:
                target_cell.font = Font(
                    name=cell.font.name,
                    size=cell.font.size,
                    bold=cell.font.bold,
                    italic=cell.font.italic,
                    color=cell.font.color.rgb if cell.font.color and cell.font.color.rgb else "00000000"
                )
            if cell.alignment:
                target_cell.alignment = Alignment(
                    horizontal=cell.alignment.horizontal,
                    vertical=cell.alignment.vertical,
                    wrap_text=cell.alignment.wrap_text
                )
            if cell.border:
                target_cell.border = cell.border
    
    def export_all_sheets(self):
        """Export all sheets while preserving formatting."""
        print("=" * 60)
        print("EXPORTING DATA TO FORMATTED EXCEL TEMPLATE")
        print("=" * 60)
        print(f"Template: {self.template_file}")
        print(f"Output: {self.output_file}\n")
        
        # Load the existing workbook
        wb = openpyxl.load_workbook(self.template_file)
        
        # Process each sheet
        for sheet_name, config in self.sheet_queries.items():
            if sheet_name in wb.sheetnames:
                print(f"\nProcessing {sheet_name}...")
                
                # Query data
                df = self.query_salesforce(config['query'])
                
                if len(df) > 0:
                    print(f"  Found {len(df)} records")
                    
                    # Get the worksheet
                    ws = wb[sheet_name]
                    
                    # Store formatting from first few rows
                    original_formatting = []
                    for row in ws.iter_rows(min_row=1, max_row=min(10, ws.max_row)):
                        row_format = []
                        for cell in row:
                            row_format.append({
                                'fill': cell.fill,
                                'font': cell.font,
                                'border': cell.border,
                                'alignment': cell.alignment,
                                'number_format': cell.number_format
                            })
                        original_formatting.append(row_format)
                    
                    # Clear existing data but keep formatting
                    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                        for cell in row:
                            cell.value = None
                    
                    # Get headers from existing sheet
                    headers = []
                    for cell in ws[1]:
                        if cell.value:
                            headers.append(cell.value)
                        else:
                            break
                    
                    # Map dataframe columns to sheet headers
                    # Remove asterisks from headers for matching
                    header_mapping = {}
                    for header in headers:
                        clean_header = header.replace('*', '')
                        if clean_header in df.columns:
                            header_mapping[clean_header] = header
                    
                    # Reorder dataframe columns to match sheet
                    ordered_columns = []
                    for header in headers:
                        clean_header = header.replace('*', '')
                        if clean_header in df.columns:
                            ordered_columns.append(clean_header)
                    
                    # Add any remaining columns not in headers
                    for col in df.columns:
                        if col not in ordered_columns:
                            ordered_columns.append(col)
                    
                    df = df[ordered_columns]
                    
                    # Write data starting from row 2
                    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
                        for c_idx, value in enumerate(row, 1):
                            cell = ws.cell(row=r_idx, column=c_idx)
                            cell.value = value
                            
                            # Apply formatting from original if available
                            if r_idx <= len(original_formatting) + 1 and c_idx <= len(original_formatting[0]):
                                fmt = original_formatting[min(r_idx-1, len(original_formatting)-1)][c_idx-1]
                                if fmt['fill']:
                                    cell.fill = fmt['fill']
                                if fmt['font']:
                                    cell.font = fmt['font']
                                if fmt['border']:
                                    cell.border = fmt['border']
                                if fmt['alignment']:
                                    cell.alignment = fmt['alignment']
                                if fmt['number_format']:
                                    cell.number_format = fmt['number_format']
                    
                    print(f"  ✓ Updated {sheet_name}")
                else:
                    print(f"  No records found")
            else:
                print(f"\n⚠️  Sheet {sheet_name} not found in template")
        
        # Save the workbook
        wb.save(self.output_file)
        print(f"\n✓ Export completed: {self.output_file}")
        
        # Summary
        print("\n" + "=" * 60)
        print("EXPORT SUMMARY")
        print("=" * 60)
        
        total_records = 0
        for sheet_name, config in self.sheet_queries.items():
            df = self.query_salesforce(config['query'])
            count = len(df)
            total_records += count
            if count > 0:
                print(f"{sheet_name}: {count} records")
        
        print(f"\nTotal records exported: {total_records}")

def main():
    exporter = FormattedExporter()
    exporter.export_all_sheets()

if __name__ == '__main__':
    main()