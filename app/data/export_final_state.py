#!/usr/bin/env python3
"""
Export the final state of all Revenue Cloud objects back to the Excel template.
This will show what was successfully imported and what failed.
"""

import subprocess
import json
import pandas as pd
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

class FinalStateExporter:
    def __init__(self):
        self.workbook_path = Path('data/Revenue_Cloud_Complete_Upload_Template.xlsx')
        self.target_org = 'fortradp2'
        
        # Object configurations for export
        self.export_configs = [
            {
                'sheet_name': '11_ProductCatalog',
                'object_name': 'ProductCatalog',
                'fields': 'Id, Name, Description'
            },
            {
                'sheet_name': '12_ProductCategory', 
                'object_name': 'ProductCategory',
                'fields': 'Id, CatalogId, ParentCategoryId, Name, Code, ProductCount'
            },
            {
                'sheet_name': '08_ProductClassification',
                'object_name': 'ProductClassification',
                'fields': 'Id, Type, SalesforceProductGrouping, Name, Code'
            },
            {
                'sheet_name': '09_AttributeDefinition',
                'object_name': 'AttributeDefinition',
                'fields': 'Id, Name, Code, DataType, DisplaySequence, AttributeCategoryId, HelpText, FacetedSearchDisplayType, IsActive, IsReadOnly, IsRequired, ValueSet, MinValue, MaxValue, PicklistId'
            },
            {
                'sheet_name': '10_AttributeCategory',
                'object_name': 'AttributeCategory',
                'fields': 'Id, Name, Code, DisplaySequence'
            },
            {
                'sheet_name': '14_AttributePicklist',
                'object_name': 'AttributePicklist',
                'fields': 'Id, Name, Description, Status, DataType, UnitOfMeasureId, OwnerId'
            },
            {
                'sheet_name': '15_ProductSellingModel',
                'object_name': 'ProductSellingModel',
                'fields': 'Id, Name, Status, SellingModelType, PricingTermUnit, BillingFrequency, StartingUnit'
            },
            {
                'sheet_name': '13_Product2',
                'object_name': 'Product2',
                'fields': 'Id, Name, ProductCode, Description, Family, IsActive, IsIncludedInAma, QuantityUnitOfMeasure, BasedOnId, RevenueTreatmentId, StockKeepingUnit, Type'
            },
            {
                'sheet_name': '17_ProductAttributeDef',
                'object_name': 'ProductAttributeDefinition',
                'fields': 'Id, ProductId, AttributeDefinitionId, Product2.Name, AttributeDefinition.Name'
            },
            {
                'sheet_name': '18_AttributePicklistValue',
                'object_name': 'AttributePicklistValue',
                'fields': 'Id, Name, PicklistId, Value, DisplayValue, Code, Sequence, IsDefault, Status, Abbreviation'
            },
            {
                'sheet_name': '19_Pricebook2',
                'object_name': 'Pricebook2',
                'fields': 'Id, Name, Description, IsActive, IsStandard'
            },
            {
                'sheet_name': '20_PricebookEntry',
                'object_name': 'PricebookEntry',
                'fields': 'Id, Product2Id, Pricebook2Id, UnitPrice, IsActive, ProductSellingModelId, Product2.Name, Product2.ProductCode'
            },
            {
                'sheet_name': '26_ProductCategoryProduct',
                'object_name': 'ProductCategoryProduct',
                'fields': 'Id, ProductCategoryId, ProductId, IsPrimaryCategory'
            },
            {
                'sheet_name': '25_ProductRelatedComponent',
                'object_name': 'ProductRelatedComponent',
                'fields': 'Id, ParentProductId, ChildProductId, ChildProductRole, MinQuantity, MaxQuantity, DefaultQuantity'
            }
        ]
    
    def query_and_export(self):
        """Query each object and export to Excel."""
        print("=" * 70)
        print("EXPORTING FINAL STATE TO EXCEL")
        print("=" * 70)
        print(f"Target workbook: {self.workbook_path}")
        print()
        
        # Load workbook
        wb = openpyxl.load_workbook(self.workbook_path)
        
        # Track results
        total_records = 0
        successful_exports = 0
        
        for config in self.export_configs:
            sheet_name = config['sheet_name']
            object_name = config['object_name']
            fields = config['fields']
            
            print(f"\nExporting {object_name} to {sheet_name}...")
            
            # Build query
            query = f"SELECT {fields} FROM {object_name} ORDER BY Name"
            
            # Run query
            cmd = [
                'sf', 'data', 'query',
                '--query', query,
                '--target-org', self.target_org,
                '--json'
            ]
            
            result = subprocess.run(cmd, capture_output=True, text=True)
            
            if result.returncode == 0:
                data = json.loads(result.stdout)
                if 'result' in data and 'records' in data['result']:
                    records = data['result']['records']
                    
                    if records:
                        # Get the sheet
                        if sheet_name in wb.sheetnames:
                            ws = wb[sheet_name]
                            
                            # Clear existing data (keep headers)
                            for row in ws.iter_rows(min_row=2):
                                for cell in row:
                                    cell.value = None
                            
                            # Get headers
                            headers = [cell.value for cell in ws[1] if cell.value]
                            
                            # Process records
                            for row_idx, record in enumerate(records, 2):
                                for col_idx, header in enumerate(headers, 1):
                                    # Remove asterisk from header for field matching
                                    field_name = header.replace('*', '')
                                    
                                    # Handle nested fields (e.g., Product2.Name)
                                    if '.' in field_name and field_name in record:
                                        # Already flattened by SOQL
                                        value = record.get(field_name)
                                    elif field_name in record:
                                        value = record.get(field_name)
                                    else:
                                        # Try without special characters
                                        clean_field = field_name.replace('_', '').replace(' ', '')
                                        value = record.get(clean_field)
                                    
                                    if value is not None:
                                        ws.cell(row=row_idx, column=col_idx).value = value
                            
                            print(f"  ✓ Exported {len(records)} records")
                            total_records += len(records)
                            successful_exports += 1
                        else:
                            print(f"  ⚠️  Sheet {sheet_name} not found")
                    else:
                        print(f"  ⚠️  No records found")
                else:
                    print(f"  ⚠️  No data in response")
            else:
                print(f"  ✗ Query failed: {result.stderr[:100]}")
        
        # Save workbook
        wb.save(self.workbook_path)
        wb.close()
        
        # Summary
        print("\n" + "=" * 70)
        print("EXPORT SUMMARY")
        print("=" * 70)
        print(f"✓ Successfully exported: {successful_exports} objects")
        print(f"✓ Total records exported: {total_records}")
        print(f"✓ Saved to: {self.workbook_path}")
        
        # Also create a summary report
        self.create_summary_report()
    
    def create_summary_report(self):
        """Create a summary report of what was imported."""
        print("\n" + "=" * 70)
        print("IMPORT SUMMARY REPORT")
        print("=" * 70)
        
        # Query counts for each object
        object_counts = {}
        
        for config in self.export_configs:
            object_name = config['object_name']
            
            # Count query
            query = f"SELECT COUNT() FROM {object_name}"
            cmd = [
                'sf', 'data', 'query',
                '--query', query,
                '--target-org', self.target_org,
                '--json'
            ]
            
            result = subprocess.run(cmd, capture_output=True, text=True)
            
            if result.returncode == 0:
                data = json.loads(result.stdout)
                if 'result' in data and 'totalSize' in data['result']:
                    count = data['result']['totalSize']
                    object_counts[object_name] = count
                    print(f"{object_name}: {count} records")
        
        # Save summary to file
        summary_file = Path('data/import_summary_report.txt')
        with open(summary_file, 'w') as f:
            f.write("REVENUE CLOUD IMPORT SUMMARY\n")
            f.write("=" * 50 + "\n\n")
            
            for obj, count in object_counts.items():
                f.write(f"{obj}: {count} records\n")
            
            f.write(f"\nTotal objects with data: {len([c for c in object_counts.values() if c > 0])}\n")
            f.write(f"Total records imported: {sum(object_counts.values())}\n")
        
        print(f"\n✓ Summary report saved to: {summary_file}")

def main():
    exporter = FinalStateExporter()
    exporter.query_and_export()

if __name__ == '__main__':
    main()