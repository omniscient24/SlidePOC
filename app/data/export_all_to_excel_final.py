#!/usr/bin/env python3
"""
Export all Revenue Cloud objects from the target org to the Excel template,
maintaining the same formatting and sheet names.
"""

import subprocess
import json
import pandas as pd
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime

class CompleteOrgExporter:
    def __init__(self):
        self.workbook_path = Path('data/Revenue_Cloud_Complete_Upload_Template.xlsx')
        self.target_org = 'fortradp2'
        
        # Define all sheets and their configurations
        self.sheet_configs = [
            {
                'sheet_name': '11_ProductCatalog',
                'object_name': 'ProductCatalog',
                'fields': ['Id', 'Name', 'Description']
            },
            {
                'sheet_name': '12_ProductCategory',
                'object_name': 'ProductCategory',
                'fields': ['Id', 'CatalogId', 'ParentCategoryId', 'Name', 'Code', 'ProductCount']
            },
            {
                'sheet_name': '08_ProductClassification',
                'object_name': 'ProductClassification',
                'fields': ['Id', 'Type', 'SalesforceProductGrouping', 'Name', 'Code']
            },
            {
                'sheet_name': '09_AttributeDefinition',
                'object_name': 'AttributeDefinition',
                'fields': ['Id', 'Name', 'Code', 'DataType', 'DisplaySequence', 'AttributeCategoryId', 
                          'HelpText', 'FacetedSearchDisplayType', 'IsActive', 'IsReadOnly', 'IsRequired', 
                          'ValueSet', 'MinValue', 'MaxValue', 'PicklistId']
            },
            {
                'sheet_name': '10_AttributeCategory',
                'object_name': 'AttributeCategory',
                'fields': ['Id', 'Name', 'Code', 'DisplaySequence']
            },
            {
                'sheet_name': '14_AttributePicklist',
                'object_name': 'AttributePicklist',
                'fields': ['Id', 'Name', 'Description', 'Status', 'DataType', 'UnitOfMeasureId', 'OwnerId']
            },
            {
                'sheet_name': '15_ProductSellingModel',
                'object_name': 'ProductSellingModel',
                'fields': ['Id', 'Name', 'Status', 'SellingModelType', 'PricingTermUnit', 
                          'BillingFrequency', 'StartingUnit']
            },
            {
                'sheet_name': '13_Product2',
                'object_name': 'Product2',
                'fields': ['Id', 'Name', 'ProductCode', 'Description', 'Family', 'IsActive', 
                          'IsIncludedInAma', 'QuantityUnitOfMeasure', 'BasedOnId', 'RevenueTreatmentId', 
                          'StockKeepingUnit', 'Type']
            },
            {
                'sheet_name': '17_ProductAttributeDef',
                'object_name': 'ProductAttributeDefinition',
                'fields': ['Id', 'ProductId', 'AttributeDefinitionId']
            },
            {
                'sheet_name': '18_AttributePicklistValue',
                'object_name': 'AttributePicklistValue',
                'fields': ['Id', 'Name', 'PicklistId', 'Value', 'DisplayValue', 'Code', 
                          'Sequence', 'IsDefault', 'Status', 'Abbreviation']
            },
            {
                'sheet_name': '19_Pricebook2',
                'object_name': 'Pricebook2',
                'fields': ['Id', 'Name', 'Description', 'IsActive', 'IsStandard']
            },
            {
                'sheet_name': '20_PricebookEntry',
                'object_name': 'PricebookEntry',
                'fields': ['Id', 'Product2Id', 'Pricebook2Id', 'UnitPrice', 'IsActive', 
                          'ProductSellingModelId']
            },
            {
                'sheet_name': '26_ProductCategoryProduct',
                'object_name': 'ProductCategoryProduct',
                'fields': ['Id', 'ProductCategoryId', 'ProductId', 'IsPrimaryCategory']
            },
            {
                'sheet_name': '25_ProductRelatedComponent',
                'object_name': 'ProductRelatedComponent',
                'fields': ['Id', 'ParentProductId', 'ChildProductId', 'ChildProductRole', 
                          'MinQuantity', 'MaxQuantity', 'DefaultQuantity']
            }
        ]
        
    def query_object_data(self, object_name, fields):
        """Query data from Salesforce object."""
        # Build SOQL query
        field_list = ', '.join(fields)
        query = f"SELECT {field_list} FROM {object_name}"
        
        # Add ORDER BY if Name field exists
        if 'Name' in fields:
            query += " ORDER BY Name"
        elif 'Code' in fields:
            query += " ORDER BY Code"
        else:
            query += " ORDER BY Id"
        
        # Execute query
        cmd = [
            'sf', 'data', 'query',
            '--query', query,
            '--target-org', self.target_org,
            '--json'
        ]
        
        result = subprocess.run(cmd, capture_output=True, text=True)
        
        if result.returncode == 0:
            data = json.loads(result.stdout)
            if 'result' in data and 'records' in data['result']:
                records = data['result']['records']
                # Remove attributes field from each record
                for record in records:
                    if 'attributes' in record:
                        del record['attributes']
                return records
        return []
    
    def export_to_excel(self):
        """Export all data to Excel workbook."""
        print("=" * 70)
        print("EXPORTING ALL REVENUE CLOUD DATA TO EXCEL")
        print("=" * 70)
        print(f"Target workbook: {self.workbook_path}")
        print(f"Source org: {self.target_org}")
        print(f"Export time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print()
        
        # Load existing workbook
        wb = openpyxl.load_workbook(self.workbook_path)
        
        # Track statistics
        total_records = 0
        sheets_updated = 0
        
        # Process each sheet
        for config in self.sheet_configs:
            sheet_name = config['sheet_name']
            object_name = config['object_name']
            fields = config['fields']
            
            print(f"\nProcessing {sheet_name} ({object_name})...")
            
            if sheet_name not in wb.sheetnames:
                print(f"  ⚠️  Sheet {sheet_name} not found, skipping")
                continue
            
            # Get the sheet
            ws = wb[sheet_name]
            
            # Get existing headers from row 1
            existing_headers = []
            header_row = 1
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=header_row, column=col).value
                if cell_value:
                    existing_headers.append(cell_value)
            
            if not existing_headers:
                print(f"  ⚠️  No headers found in sheet, skipping")
                continue
            
            # Query data
            records = self.query_object_data(object_name, fields)
            
            if records:
                print(f"  Found {len(records)} records")
                
                # Clear existing data (keep headers and formatting)
                # Start from row 2 to preserve headers
                for row in range(2, ws.max_row + 1):
                    for col in range(1, len(existing_headers) + 1):
                        ws.cell(row=row, column=col).value = None
                
                # Map field names to column positions
                field_to_col = {}
                for col_idx, header in enumerate(existing_headers, 1):
                    # Remove asterisk from header for matching
                    clean_header = header.replace('*', '').strip()
                    field_to_col[clean_header] = col_idx
                
                # Write data
                for row_idx, record in enumerate(records, 2):
                    for field_name, value in record.items():
                        if field_name in field_to_col:
                            col_idx = field_to_col[field_name]
                            ws.cell(row=row_idx, column=col_idx).value = value
                        else:
                            # Try to match with different variations
                            # Handle Product2.Name type fields
                            if '.' in field_name:
                                base_field = field_name.split('.')[1]
                                if base_field in field_to_col:
                                    col_idx = field_to_col[base_field]
                                    ws.cell(row=row_idx, column=col_idx).value = value
                
                print(f"  ✓ Exported {len(records)} records")
                total_records += len(records)
                sheets_updated += 1
            else:
                print(f"  No records found")
        
        # Add Product2.Name and Product2.ProductCode to sheets that need them
        self.add_product_references(wb)
        
        # Save workbook
        wb.save(self.workbook_path)
        wb.close()
        
        print("\n" + "=" * 70)
        print("EXPORT COMPLETE")
        print("=" * 70)
        print(f"✓ Sheets updated: {sheets_updated}")
        print(f"✓ Total records exported: {total_records}")
        print(f"✓ Workbook saved: {self.workbook_path}")
        
    def add_product_references(self, wb):
        """Add Product2.Name and ProductCode to relevant sheets."""
        print("\n" + "-" * 50)
        print("Adding Product references...")
        
        # First, get Product2 data
        product_data = {}
        if '13_Product2' in wb.sheetnames:
            ws_product = wb['13_Product2']
            
            # Find column indices
            id_col = None
            name_col = None
            code_col = None
            
            for col in range(1, ws_product.max_column + 1):
                header = ws_product.cell(row=1, column=col).value
                if header:
                    if header.replace('*', '').strip() == 'Id':
                        id_col = col
                    elif header.replace('*', '').strip() == 'Name':
                        name_col = col
                    elif header.replace('*', '').strip() == 'ProductCode':
                        code_col = col
            
            if id_col and name_col and code_col:
                for row in range(2, ws_product.max_row + 1):
                    product_id = ws_product.cell(row=row, column=id_col).value
                    if product_id:
                        product_data[product_id] = {
                            'Name': ws_product.cell(row=row, column=name_col).value,
                            'ProductCode': ws_product.cell(row=row, column=code_col).value
                        }
        
        # Update PricebookEntry sheet
        if '20_PricebookEntry' in wb.sheetnames:
            ws_pbe = wb['20_PricebookEntry']
            self.update_sheet_with_product_refs(ws_pbe, product_data, 'Product2Id')
            print("  ✓ Updated PricebookEntry with Product names")
        
        # Update ProductAttributeDef sheet
        if '17_ProductAttributeDef' in wb.sheetnames:
            ws_pad = wb['17_ProductAttributeDef']
            self.update_sheet_with_product_refs(ws_pad, product_data, 'ProductId')
            print("  ✓ Updated ProductAttributeDef with Product names")
    
    def update_sheet_with_product_refs(self, ws, product_data, id_field_name):
        """Update a sheet with Product2 Name and ProductCode."""
        # Find column indices
        id_col = None
        name_col = None
        code_col = None
        
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header:
                clean_header = header.replace('*', '').strip()
                if clean_header == id_field_name:
                    id_col = col
                elif clean_header == 'Product2.Name':
                    name_col = col
                elif clean_header == 'Product2.ProductCode':
                    code_col = col
        
        if id_col and (name_col or code_col):
            for row in range(2, ws.max_row + 1):
                product_id = ws.cell(row=row, column=id_col).value
                if product_id and product_id in product_data:
                    if name_col:
                        ws.cell(row=row, column=name_col).value = product_data[product_id]['Name']
                    if code_col:
                        ws.cell(row=row, column=code_col).value = product_data[product_id]['ProductCode']

def main():
    exporter = CompleteOrgExporter()
    exporter.export_to_excel()

if __name__ == '__main__':
    main()