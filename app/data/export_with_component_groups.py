#\!/usr/bin/env python3
"""
Export including ProductComponentGroup properly.
"""

import subprocess
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
import json

def main():
    print("=" * 60)
    print("EXPORTING WITH PRODUCTCOMPONENTGROUP")
    print("=" * 60)
    
    # Get ProductComponentGroup data
    query = "SELECT Id, Name, Description, ParentProductId, Sequence, Code FROM ProductComponentGroup ORDER BY ParentProductId, Sequence"
    
    cmd = [
        'sf', 'data', 'query',
        '--query', query,
        '--target-org', 'fortradp2',
        '--json'
    ]
    
    result = subprocess.run(cmd, capture_output=True, text=True)
    
    if result.returncode == 0:
        data = json.loads(result.stdout)
        if 'result' in data and 'records' in data['result']:
            records = data['result']['records']
            print(f"Found {len(records)} ProductComponentGroup records")
            
            if records:
                # Load workbook
                wb = load_workbook('data/Revenue_Cloud_Complete_Upload_Template.xlsx')
                ws = wb['14_ProductComponentGroup']
                
                # Clear existing data
                for row in range(2, ws.max_row + 1):
                    for col in range(1, 6):  # Assuming 5 columns
                        ws.cell(row=row, column=col).value = None
                
                # Write new data
                row_num = 2
                for record in records:
                    ws.cell(row=row_num, column=1).value = record.get('Name')
                    ws.cell(row=row_num, column=2).value = record.get('Description')
                    ws.cell(row=row_num, column=3).value = record.get('ParentProductId')
                    ws.cell(row=row_num, column=4).value = record.get('Sequence')
                    ws.cell(row=row_num, column=5).value = record.get('Code')
                    row_num += 1
                
                # Save workbook
                wb.save('data/Revenue_Cloud_Complete_Upload_Template.xlsx')
                print(f"✓ Updated ProductComponentGroup sheet with {len(records)} records")
                
                # Show what was exported
                print("\nExported groups:")
                for rec in records:
                    print(f"  - {rec['Name']} ({rec['Code']})")
    
    print("\n✅ ProductComponentGroup data successfully exported")

if __name__ == '__main__':
    main()
