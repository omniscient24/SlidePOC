#!/usr/bin/env python3
"""
Export data from Salesforce and populate the template while preserving formatting.
"""

import pandas as pd
import subprocess
import json
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import io
import csv

class FormattedExporter:
    def __init__(self):
        self.template_file = Path('data/Revenue_Cloud_Clean_Template.xlsx')
        self.output_file = Path(f'data/Revenue_Cloud_Export_Formatted_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
        self.target_org = 'fortradp2'
        
        # Core objects with known fields
        self.core_queries = {
            'ProductCatalog': "SELECT Id, Name, Code FROM ProductCatalog",
            'ProductCategory': "SELECT Id, Name, Code, CatalogId, ParentCategoryId FROM ProductCategory",
            'Product2': "SELECT Id, Name, ProductCode, Description, Family, Type, BasedOnId, IsActive FROM Product2",
            'AttributeDefinition': "SELECT Id, Name, Code, Label FROM AttributeDefinition",
            'ProductClassification': "SELECT Id, Name, Code, Status FROM ProductClassification",
            'ProductClassificationAttr': "SELECT Id, Name, ProductClassificationId, AttributeDefinitionId, Status FROM ProductClassificationAttr",
            'Pricebook2': "SELECT Id, Name, Description, IsActive, IsStandard FROM Pricebook2",
            'PricebookEntry': "SELECT Id, Product2Id, Pricebook2Id, UnitPrice, IsActive FROM PricebookEntry WHERE IsActive = true",
            'ProductCategoryProduct': "SELECT Id, ProductCategoryId, ProductId FROM ProductCategoryProduct",
            'ProductAttributeDefinition': "SELECT Id, Name, Product2Id, ProductClassificationAttributeId, AttributeDefinitionId, Status FROM ProductAttributeDefinition",
            'ProductSellingModel': "SELECT Id, Name, SellingModelType, Status, PricingTermUnit, PricingTerm FROM ProductSellingModel",
            'ProductRelatedComponent': "SELECT Id, ParentProductId, ChildProductId, ProductRelationshipTypeId, Quantity, IsDefaultComponent, Sequence FROM ProductRelatedComponent",
            'ProductRelationshipType': "SELECT Id, Name FROM ProductRelationshipType"
        }
        
        # Sheet to object mapping
        self.sheet_mappings = {
            '11_ProductCatalog': 'ProductCatalog',
            '12_ProductCategory': 'ProductCategory',
            '13_Product2': 'Product2',
            '09_AttributeDefinition': 'AttributeDefinition',
            '08_ProductClassification': 'ProductClassification',
            '19_Pricebook2': 'Pricebook2',
            '20_PricebookEntry': 'PricebookEntry',
            '26_ProductCategoryProduct': 'ProductCategoryProduct',
            '17_ProductAttributeDef': 'ProductAttributeDefinition',
            '15_ProductSellingModel': 'ProductSellingModel',
            '25_ProductRelatedComponent': 'ProductRelatedComponent'
        }
    
    def query_salesforce_simple(self, query):
        """Execute a simple query and return results."""
        try:
            cmd = [
                'sf', 'data', 'query',
                '--query', query,
                '--target-org', self.target_org,
                '--result-format', 'csv'
            ]
            
            result = subprocess.run(cmd, capture_output=True, text=True)
            
            if result.returncode == 0:
                lines = result.stdout.strip().split('\n')
                if len(lines) > 1:
                    data_lines = [line for line in lines if not line.startswith('Total number of records') and line.strip()]
                    if data_lines:
                        csv_data = '\n'.join(data_lines)
                        reader = csv.DictReader(io.StringIO(csv_data))
                        return list(reader)
                return []
            else:
                print(f"Error: {result.stderr}")
                return []
        except Exception as e:
            print(f"Exception: {str(e)}")
            return []
    
    def copy_cell_formatting(self, source_cell, target_cell):
        """Copy formatting from source cell to target cell."""
        if source_cell.has_style:
            # Copy font
            if source_cell.font:
                target_cell.font = Font(
                    name=source_cell.font.name,
                    size=source_cell.font.size,
                    bold=source_cell.font.bold,
                    italic=source_cell.font.italic,
                    vertAlign=source_cell.font.vertAlign,
                    underline=source_cell.font.underline,
                    strike=source_cell.font.strike,
                    color=source_cell.font.color
                )
            
            # Copy fill
            if source_cell.fill:
                target_cell.fill = PatternFill(
                    fill_type=source_cell.fill.fill_type,
                    start_color=source_cell.fill.start_color,
                    end_color=source_cell.fill.end_color
                )
            
            # Copy alignment
            if source_cell.alignment:
                target_cell.alignment = Alignment(
                    horizontal=source_cell.alignment.horizontal,
                    vertical=source_cell.alignment.vertical,
                    text_rotation=source_cell.alignment.text_rotation,
                    wrap_text=source_cell.alignment.wrap_text,
                    shrink_to_fit=source_cell.alignment.shrink_to_fit,
                    indent=source_cell.alignment.indent
                )
            
            # Copy border
            if source_cell.border:
                target_cell.border = Border(
                    left=source_cell.border.left,
                    right=source_cell.border.right,
                    top=source_cell.border.top,
                    bottom=source_cell.border.bottom,
                    diagonal=source_cell.border.diagonal,
                    diagonal_direction=source_cell.border.diagonal_direction,
                    outline=source_cell.border.outline,
                    vertical=source_cell.border.vertical,
                    horizontal=source_cell.border.horizontal
                )
            
            # Copy number format
            target_cell.number_format = source_cell.number_format
            
            # Copy protection
            target_cell.protection = source_cell.protection
    
    def export_with_formatting(self):
        """Export data and preserve template formatting."""
        print(f"Starting formatted export from org: {self.target_org}")
        print(f"Template: {self.template_file}")
        print(f"Output: {self.output_file}\n")
        
        # First, copy the template file to preserve all formatting
        import shutil
        shutil.copy2(self.template_file, self.output_file)
        
        # Load the workbook with formatting
        wb = load_workbook(self.output_file)
        
        # Query all data first
        exported_data = {}
        for obj_name, query in self.core_queries.items():
            print(f"Querying {obj_name}...")
            data = self.query_salesforce_simple(query)
            if data:
                print(f"  Found {len(data)} records")
                exported_data[obj_name] = data
            else:
                print(f"  No records found")
                exported_data[obj_name] = []
        
        print("\nPopulating sheets with formatting...")
        
        # Process each sheet
        for sheet_name in wb.sheetnames:
            if sheet_name in ['Instructions', 'Picklist Values']:
                continue
                
            if sheet_name in self.sheet_mappings and self.sheet_mappings[sheet_name] in exported_data:
                ws = wb[sheet_name]
                object_name = self.sheet_mappings[sheet_name]
                data = exported_data[object_name]
                
                if data:
                    print(f"Populating {sheet_name} with {len(data)} records")
                    
                    # Get column headers from first row
                    headers = []
                    for cell in ws[1]:
                        if cell.value:
                            headers.append(cell.value)
                    
                    # Store formatting from row 2 (first data row)
                    row_2_formatting = {}
                    for col_idx, cell in enumerate(ws[2], 1):
                        row_2_formatting[col_idx] = {
                            'font': cell.font,
                            'fill': cell.fill,
                            'alignment': cell.alignment,
                            'border': cell.border,
                            'number_format': cell.number_format,
                            'protection': cell.protection
                        }
                    
                    # Clear existing data (keep headers)
                    for row in ws.iter_rows(min_row=2):
                        for cell in row:
                            cell.value = None
                    
                    # Populate with new data
                    for row_idx, record in enumerate(data, 2):
                        for col_idx, header in enumerate(headers, 1):
                            cell = ws.cell(row=row_idx, column=col_idx)
                            
                            # Map data to cell
                            clean_header = header.replace('*', '')
                            if clean_header in record:
                                cell.value = record[clean_header]
                            elif header == 'External_ID__c':
                                cell.value = f"{sheet_name}_{row_idx-1}"
                            elif header == 'Name' and 'Name' not in record and 'Id' in record:
                                cell.value = record['Id']
                            elif 'Id' in header and header != 'Id':
                                cell.value = 'ID_PLACEHOLDER'
                            elif header in ['IsActive', 'IsRequired', 'IsHidden']:
                                cell.value = True
                            elif header in ['Sequence', 'MinQuantity', 'MaxQuantity']:
                                cell.value = 0
                            elif header == 'Quantity':
                                cell.value = 1
                            
                            # Apply formatting from row 2
                            if col_idx in row_2_formatting:
                                fmt = row_2_formatting[col_idx]
                                if fmt['font']:
                                    cell.font = fmt['font']
                                if fmt['fill']:
                                    cell.fill = fmt['fill']
                                if fmt['alignment']:
                                    cell.alignment = fmt['alignment']
                                if fmt['border']:
                                    cell.border = fmt['border']
                                if fmt['number_format']:
                                    cell.number_format = fmt['number_format']
                                if fmt['protection']:
                                    cell.protection = fmt['protection']
                    
                    # Copy column widths
                    for col in ws.columns:
                        col_letter = col[0].column_letter
                        if ws.column_dimensions[col_letter].width:
                            ws.column_dimensions[col_letter].width = ws.column_dimensions[col_letter].width
        
        # Save the workbook
        wb.save(self.output_file)
        wb.close()
        
        print(f"\nExport completed with formatting preserved!")
        print(f"Output file: {self.output_file}")
        
        return self.output_file

def main():
    exporter = FormattedExporter()
    exporter.export_with_formatting()

if __name__ == '__main__':
    main()