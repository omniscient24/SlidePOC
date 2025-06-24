#!/usr/bin/env python3
"""
Create phase-specific Excel templates from the master template.
Splits the master workbook into smaller workbooks based on implementation phases.
"""

import pandas as pd
import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Define phase mappings
PHASE_MAPPINGS = {
    'phase1-foundation': {
        'name': 'Phase 1 - Foundation',
        'sheets': ['02_LegalEntity', '05_TaxTreatment', '04_TaxPolicy', '03_TaxEngine'],
        'description': 'Core foundation objects that must be configured first'
    },
    'phase2-products': {
        'name': 'Phase 2 - Products & Pricing',
        'sheets': [
            '11_ProductCatalog', '12_ProductCategory', '13_Product2', 
            '26_ProductCategoryProduct', '15_ProductSellingModel',
            '14_ProductComponentGroup', '25_ProductRelatedComponent',
            '08_ProductClassification', '09_AttributeDefinition',
            '14_AttributePicklist', '18_AttributePicklistValue',
            '17_ProductAttributeDef', '10_AttributeCategory',
            '19_Pricebook2', '20_PricebookEntry'
        ],
        'description': 'Product catalog, attributes, and pricing configuration'
    },
    'phase3-operations': {
        'name': 'Phase 3 - Operations',
        'sheets': [
            '01_CostBook', '15_CostBookEntry',
            '21_PriceAdjustmentSchedule', '22_PriceAdjustmentTier',
            '23_AttributeBasedAdjRule', '24_AttributeBasedAdj',
            '06_BillingPolicy', '07_BillingTreatment'
        ],
        'description': 'Operational configurations for cost, pricing adjustments, and billing'
    }
}

def create_phase_template(master_file, phase_key, phase_config, output_dir):
    """Create a phase-specific template from the master file."""
    
    # Create output directory if it doesn't exist
    phase_dir = os.path.join(output_dir, phase_key)
    os.makedirs(phase_dir, exist_ok=True)
    
    # Load the master workbook
    master_wb = load_workbook(master_file)
    
    # Create a new workbook for this phase
    phase_wb = Workbook()
    phase_wb.remove(phase_wb.active)  # Remove default sheet
    
    # Always copy the Instructions sheet first
    if 'Instructions' in master_wb.sheetnames:
        instructions_sheet = master_wb['Instructions']
        new_sheet = phase_wb.create_sheet('Instructions')
        
        # Copy all cells
        for row in instructions_sheet.iter_rows():
            for cell in row:
                new_cell = new_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    new_cell.font = cell.font.copy()
                    new_cell.fill = cell.fill.copy()
                    new_cell.border = cell.border.copy()
                    new_cell.alignment = cell.alignment.copy()
    
    # Create a summary sheet for this phase
    summary_sheet = phase_wb.create_sheet('Phase_Summary')
    summary_sheet['A1'] = phase_config['name']
    summary_sheet['A2'] = phase_config['description']
    summary_sheet['A4'] = 'Included Objects:'
    
    # Copy relevant sheets for this phase
    row_num = 5
    for sheet_name in phase_config['sheets']:
        if sheet_name in master_wb.sheetnames:
            # Add to summary
            summary_sheet[f'A{row_num}'] = sheet_name.split('_')[1] if '_' in sheet_name else sheet_name
            row_num += 1
            
            # Copy the sheet
            source_sheet = master_wb[sheet_name]
            new_sheet = phase_wb.create_sheet(sheet_name)
            
            # Copy all cells with their values and styles
            for row in source_sheet.iter_rows():
                for cell in row:
                    new_cell = new_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
                    if cell.has_style:
                        new_cell.font = cell.font.copy()
                        new_cell.fill = cell.fill.copy()
                        new_cell.border = cell.border.copy()
                        new_cell.alignment = cell.alignment.copy()
            
            # Copy column dimensions
            for col_letter, col_dim in source_sheet.column_dimensions.items():
                new_sheet.column_dimensions[col_letter].width = col_dim.width
    
    # Save the phase template
    output_file = os.path.join(phase_dir, f'{phase_key}_template.xlsx')
    phase_wb.save(output_file)
    print(f"Created {output_file}")
    
    # Also create individual object templates
    for sheet_name in phase_config['sheets']:
        if sheet_name in master_wb.sheetnames:
            object_wb = Workbook()
            object_wb.remove(object_wb.active)
            
            # Copy the instructions sheet
            if 'Instructions' in master_wb.sheetnames:
                instructions_sheet = master_wb['Instructions']
                new_sheet = object_wb.create_sheet('Instructions')
                
                for row in instructions_sheet.iter_rows():
                    for cell in row:
                        new_cell = new_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
                        if cell.has_style:
                            new_cell.font = cell.font.copy()
                            new_cell.fill = cell.fill.copy()
                            new_cell.border = cell.border.copy()
                            new_cell.alignment = cell.alignment.copy()
            
            # Copy the object sheet
            source_sheet = master_wb[sheet_name]
            new_sheet = object_wb.create_sheet(sheet_name)
            
            for row in source_sheet.iter_rows():
                for cell in row:
                    new_cell = new_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
                    if cell.has_style:
                        new_cell.font = cell.font.copy()
                        new_cell.fill = cell.fill.copy()
                        new_cell.border = cell.border.copy()
                        new_cell.alignment = cell.alignment.copy()
            
            # Copy column dimensions
            for col_letter, col_dim in source_sheet.column_dimensions.items():
                new_sheet.column_dimensions[col_letter].width = col_dim.width
            
            # Save individual object template
            object_name = sheet_name.split('_')[1] if '_' in sheet_name else sheet_name
            object_file = os.path.join(phase_dir, f'{object_name}.xlsx')
            object_wb.save(object_file)
            print(f"  Created individual template: {object_file}")

def main():
    """Main function to create all phase templates."""
    
    master_file = 'data/templates/master/Revenue_Cloud_Complete_Upload_Template.xlsx'
    output_dir = 'data/templates/phase-templates'
    
    print("Creating phase-specific templates...")
    
    for phase_key, phase_config in PHASE_MAPPINGS.items():
        print(f"\nProcessing {phase_config['name']}...")
        create_phase_template(master_file, phase_key, phase_config, output_dir)
    
    print("\nPhase templates created successfully!")
    
    # Create a phase 4 validation template
    phase4_dir = os.path.join(output_dir, 'phase4-finalization')
    os.makedirs(phase4_dir, exist_ok=True)
    
    # Create validation checklist
    val_wb = Workbook()
    val_sheet = val_wb.active
    val_sheet.title = 'Validation_Checklist'
    
    val_sheet['A1'] = 'Revenue Cloud Implementation - Validation Checklist'
    val_sheet['A3'] = 'Object'
    val_sheet['B3'] = 'Status'
    val_sheet['C3'] = 'Records Loaded'
    val_sheet['D3'] = 'Validation Notes'
    
    # Add all objects to checklist
    row = 4
    all_objects = []
    for phase_config in PHASE_MAPPINGS.values():
        for sheet in phase_config['sheets']:
            object_name = sheet.split('_')[1] if '_' in sheet else sheet
            all_objects.append(object_name)
    
    for obj in sorted(set(all_objects)):
        val_sheet[f'A{row}'] = obj
        val_sheet[f'B{row}'] = 'Pending'
        row += 1
    
    val_file = os.path.join(phase4_dir, 'validation_checklist.xlsx')
    val_wb.save(val_file)
    print(f"\nCreated validation checklist: {val_file}")

if __name__ == "__main__":
    main()