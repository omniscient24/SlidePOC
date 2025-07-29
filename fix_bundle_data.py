#!/usr/bin/env python3
"""
Fix Bundle Data in Excel Workbook
Adds the missing ProductRelatedComponent data for bundle products
"""

import pandas as pd
import os
from openpyxl import load_workbook

def add_bundle_data():
    """Add bundle component relationships to the Excel workbook"""
    
    # Define the bundle mappings (from server.py)
    bundle_mappings = {
        # DCS Bundles
        '01tdp000006HfphAAC': [  # DCS Essentials Bundle
            {'id': '01tdp000006HfpoAAC', 'name': 'DCS for Windows', 'seq': 10, 'qty': 1},
            {'id': '01tdp000006HfpkAAC', 'name': 'Data Detection Engine', 'seq': 20, 'qty': 1},
            {'id': '01tdp000006HfprAAC', 'name': 'DCS Getting Started Package', 'seq': 30, 'qty': 1}
        ],
        '01tdp000006HfpiAAC': [  # DCS Advanced Bundle
            {'id': '01tdp000006HfpoAAC', 'name': 'DCS for Windows', 'seq': 40, 'qty': 1},
            {'id': '01tdp000006HfpkAAC', 'name': 'Data Detection Engine', 'seq': 50, 'qty': 1},
            {'id': '01tdp000006HfpnAAC', 'name': 'DCS Admin Console', 'seq': 60, 'qty': 1},
            {'id': '01tdp000006HfpmAAC', 'name': 'DCS Analysis Collector', 'seq': 70, 'qty': 1},
            {'id': '01tdp000006HfppAAC', 'name': 'DCS for OWA', 'seq': 80, 'qty': 1}
        ],
        '01tdp000006HfpjAAC': [  # DCS Elite Bundle
            {'id': '01tdp000006HfpoAAC', 'name': 'DCS for Windows', 'seq': 90, 'qty': 1},
            {'id': '01tdp000006HfpkAAC', 'name': 'Data Detection Engine', 'seq': 100, 'qty': 1},
            {'id': '01tdp000006HfpnAAC', 'name': 'DCS Admin Console', 'seq': 110, 'qty': 1},
            {'id': '01tdp000006HfpmAAC', 'name': 'DCS Analysis Collector', 'seq': 120, 'qty': 1},
            {'id': '01tdp000006HfppAAC', 'name': 'DCS for OWA', 'seq': 130, 'qty': 1},
            {'id': '01tdp000006HfplAAC', 'name': 'Unlimited Classification', 'seq': 140, 'qty': 1},
            {'id': '01tdp000006HfpqAAC', 'name': 'Software Development Kit', 'seq': 150, 'qty': 1}
        ],
        # HRM Bundles
        '01tdp000006iLGbAAM': [  # HRM Essentials Bundle
            {'id': '01tdp000006iLGcAAM', 'name': 'HRM Core Module', 'seq': 10, 'qty': 1},
            {'id': '01tdp000006iLGdAAM', 'name': 'HRM Basic Training', 'seq': 20, 'qty': 1}
        ],
        '01tdp000006m0jpAAA': [  # HRM Advanced Bundle
            {'id': '01tdp000006iLGcAAM', 'name': 'HRM Core Module', 'seq': 10, 'qty': 1},
            {'id': '01tdp000006iLGdAAM', 'name': 'HRM Basic Training', 'seq': 20, 'qty': 1},
            {'id': '01tdp000006iLGeAAM', 'name': 'HRM Phishing Simulation', 'seq': 30, 'qty': 1},
            {'id': '01tdp000006iLGfAAM', 'name': 'HRM Advanced Analytics', 'seq': 40, 'qty': 1}
        ],
        '01tdp000006m14nAAA': [  # HRM Elite Bundle
            {'id': '01tdp000006iLGcAAM', 'name': 'HRM Core Module', 'seq': 10, 'qty': 1},
            {'id': '01tdp000006iLGdAAM', 'name': 'HRM Basic Training', 'seq': 20, 'qty': 1},
            {'id': '01tdp000006iLGeAAM', 'name': 'HRM Phishing Simulation', 'seq': 30, 'qty': 1},
            {'id': '01tdp000006iLGfAAM', 'name': 'HRM Advanced Analytics', 'seq': 40, 'qty': 1},
            {'id': '01tdp000006iLGgAAM', 'name': 'HRM Executive Dashboard', 'seq': 50, 'qty': 1},
            {'id': '01tdp000006iLGhAAM', 'name': 'HRM Custom Campaigns', 'seq': 60, 'qty': 1}
        ]
    }
    
    # Paths to check
    paths = [
        'data/Revenue_Cloud_Complete_Upload_Template.xlsx',
        'data/templates/master/Revenue_Cloud_Complete_Upload_Template.xlsx'
    ]
    
    workbook_path = None
    for path in paths:
        if os.path.exists(path):
            workbook_path = path
            break
    
    if not workbook_path:
        print("❌ Could not find Excel workbook")
        return
    
    print(f"📚 Loading workbook from: {workbook_path}")
    
    # Load the workbook
    wb = load_workbook(workbook_path)
    
    # Create new data for ProductRelatedComponent sheet
    prc_data = []
    
    for bundle_id, components in bundle_mappings.items():
        for comp in components:
            prc_data.append({
                'ProductId': bundle_id,
                'ChildProductId': comp['id'],
                'ProductRelationshipType': 'Component',
                'IsActive': True,
                'Quantity': comp['qty'],
                'SequenceNumber': comp['seq']
            })
    
    # Convert to DataFrame
    prc_df = pd.DataFrame(prc_data)
    
    # Load existing data if sheet exists
    sheet_name = '25_ProductRelatedComponent'
    if sheet_name in wb.sheetnames:
        existing_df = pd.read_excel(workbook_path, sheet_name=sheet_name)
        print(f"✓ Found existing {len(existing_df)} rows in {sheet_name}")
        
        # Check if bundle data already exists
        existing_bundles = set(existing_df['ProductId'].unique()) if 'ProductId' in existing_df.columns else set()
        new_bundles = set(bundle_mappings.keys())
        
        if existing_bundles.intersection(new_bundles):
            print("⚠️  Some bundle data already exists. Skipping to avoid duplicates.")
            return
        
        # Append new data
        combined_df = pd.concat([existing_df, prc_df], ignore_index=True)
    else:
        combined_df = prc_df
        print(f"Creating new sheet: {sheet_name}")
    
    # Save back to Excel
    with pd.ExcelWriter(workbook_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        combined_df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    print(f"✅ Added {len(prc_df)} bundle component relationships")
    print(f"📝 Total rows in {sheet_name}: {len(combined_df)}")
    
    # Also update Product2 sheet to mark bundles
    products_df = pd.read_excel(workbook_path, sheet_name='13_Product2')
    bundle_ids = list(bundle_mappings.keys())
    
    # Add IsBundle column if it doesn't exist
    if 'IsBundle' not in products_df.columns:
        products_df['IsBundle'] = products_df['Id'].isin(bundle_ids)
        
        with pd.ExcelWriter(workbook_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            products_df.to_excel(writer, sheet_name='13_Product2', index=False)
        
        print(f"✅ Marked {products_df['IsBundle'].sum()} products as bundles")

if __name__ == '__main__':
    add_bundle_data()