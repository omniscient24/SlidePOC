#!/usr/bin/env python3
"""Create a new workbook for Fulfillment Plan configuration objects."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path
from datetime import datetime

# Define Fulfillment objects with their fields
FULFILLMENT_OBJECTS = {
    # Core Fulfillment Configuration
    'Location': {
        'sheet_name': '01_Location',
        'fields': [
            'Id', 'Name', 'LocationType', 'TimeZone', 'Description',
            'DrivingDirections', 'IsInventoryLocation', 'IsMobile',
            'Latitude', 'Longitude', 'OpenDate', 'CloseDate',
            'ConstructionStartDate', 'ConstructionCompleteDate',
            'RemodelDate', 'PossessionDate', 'ExternalReference'
        ]
    },
    'AssociatedLocation': {
        'sheet_name': '02_AssociatedLocation',
        'fields': [
            'Id', 'ParentLocationId', 'AssociatedLocationId', 'Type',
            'ActiveFrom', 'ActiveTo'
        ]
    },
    'OrderDeliveryMethod': {
        'sheet_name': '03_OrderDeliveryMethod',
        'fields': [
            'Id', 'Name', 'Description', 'IsActive', 'ProductId',
            'ClassOfService', 'Carrier', 'ReferenceNumber'
        ]
    },
    
    # Fulfillment Plan Configuration
    'WorkPlanTemplate': {
        'sheet_name': '04_WorkPlanTemplate',
        'fields': [
            'Id', 'Name', 'Description', 'IsActive', 'MasterLabel'
        ]
    },
    'WorkPlanTemplateEntry': {
        'sheet_name': '05_WorkPlanTemplateEntry',
        'fields': [
            'Id', 'WorkPlanTemplateId', 'WorkStepTemplateId', 
            'ExecutionOrder', 'MasterLabel'
        ]
    },
    'WorkPlan': {
        'sheet_name': '06_WorkPlan',
        'fields': [
            'Id', 'Name', 'Description', 'WorkPlanTemplateId',
            'ParentRecordId', 'ParentRecordType'
        ]
    },
    
    # Fulfillment Orchestration
    'FulfillmentOrder': {
        'sheet_name': '07_FulfillmentOrder',
        'fields': [
            'Id', 'FulfillmentOrderNumber', 'OrderSummaryId', 'AccountId',
            'Status', 'StatusCategory', 'Type', 'TypeCategory',
            'FulfilledFromLocationId', 'FulfilledToName', 'FulfilledToStreet',
            'FulfilledToCity', 'FulfilledToState', 'FulfilledToPostalCode',
            'FulfilledToCountry', 'FulfilledToLatitude', 'FulfilledToLongitude',
            'FulfilledToGeocodeAccuracy', 'FulfilledToPhone', 'FulfilledToEmail',
            'ItemCount', 'TotalProductAmount', 'TotalProductTaxAmount',
            'TotalAdjustmentAmount', 'TotalAdjustmentTaxAmount', 'TotalAmount',
            'TotalTaxAmount', 'GrandTotalAmount'
        ]
    },
    'FulfillmentOrderLineItem': {
        'sheet_name': '08_FulfillmentOrderLineItem',
        'fields': [
            'Id', 'FulfillmentOrderId', 'OrderItemId', 'OrderItemSummaryId',
            'Product2Id', 'Description', 'Quantity', 'OriginalQuantity',
            'AvailableQuantity', 'RejectedQuantity', 'ReasonForRejection',
            'UnitPrice', 'TotalPrice', 'TotalLineAmount', 'TotalLineTaxAmount',
            'TotalAdjustmentAmount', 'TotalAdjustmentTaxAmount', 'GrossUnitPrice',
            'TotalAmount'
        ]
    },
    
    # Order Configuration
    'OrderDeliveryGroup': {
        'sheet_name': '09_OrderDeliveryGroup',
        'fields': [
            'Id', 'OrderId', 'OrderDeliveryMethodId', 'DeliverToName',
            'DeliverToStreet', 'DeliverToCity', 'DeliverToState',
            'DeliverToPostalCode', 'DeliverToCountry', 'DeliverToLatitude',
            'DeliverToLongitude', 'Instructions', 'DesiredDeliveryDate',
            'EmailAddress', 'PhoneNumber', 'TotalAmount', 'TotalTaxAmount'
        ]
    },
    'OrderItemGroup': {
        'sheet_name': '10_OrderItemGroup',
        'fields': [
            'Id', 'Name', 'Description', 'OrderId'
        ]
    },
    
    # Shipment Management
    'Shipment': {
        'sheet_name': '11_Shipment',
        'fields': [
            'Id', 'ShipmentNumber', 'OrderSummaryId', 'FulfillmentOrderId',
            'ShipFromLocationId', 'ShipToName', 'ShipToStreet', 'ShipToCity',
            'ShipToState', 'ShipToPostalCode', 'ShipToCountry',
            'TrackingNumber', 'TrackingUrl', 'Carrier', 'Status',
            'ShipDate', 'DeliveryDate', 'ActualDeliveryDate'
        ]
    },
    'ShipmentItem': {
        'sheet_name': '12_ShipmentItem',
        'fields': [
            'Id', 'ShipmentId', 'OrderItemSummaryId', 'FulfillmentOrderLineItemId',
            'Product2Id', 'ProductName', 'Quantity', 'Description'
        ]
    },
    
    # Product Fulfillment Configuration
    'ProductSellingModel': {
        'sheet_name': '13_ProductSellingModel',
        'fields': [
            'Id', 'Name', 'SellingModelType', 'Status', 'PricingTerm',
            'PricingTermUnit', 'IsActive', 'Description'
        ]
    },
    'ProductSellingModelOption': {
        'sheet_name': '14_ProductSellingModelOption',
        'fields': [
            'Id', 'ProductId', 'ProductSellingModelId', 'ProrationPolicyId',
            'LifecycleMgmtEnabledFlag'
        ]
    },
    
    # Action Plans
    'ActionPlan': {
        'sheet_name': '15_ActionPlan',
        'fields': [
            'Id', 'Name', 'Description', 'StartDate', 'TargetDate',
            'OwnerId', 'Status', 'Type', 'Category', 'ParentId'
        ]
    },
    'ActionPlanItemDependency': {
        'sheet_name': '16_ActionPlanItemDependency',
        'fields': [
            'Id', 'PreviousItemId', 'ActionPlanItemId'
        ]
    },
    
    # Asset Management  
    'AssetRelationship': {
        'sheet_name': '17_AssetRelationship',
        'fields': [
            'Id', 'AssetId', 'RelatedAssetId', 'RelationshipType',
            'FromDate', 'ToDate'
        ]
    },
    'AssetStatePeriod': {
        'sheet_name': '18_AssetStatePeriod',
        'fields': [
            'Id', 'AssetId', 'StartDate', 'EndDate', 'Quantity',
            'Amount', 'MonthlyRecurringRevenue', 'AssetStateId'
        ]
    }
}

def create_fulfillment_workbook():
    """Create a new workbook for fulfillment configuration."""
    # Create new workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Create Instructions sheet first
    instructions_ws = wb.create_sheet("Instructions", 0)
    instructions_ws['A1'] = "Revenue Cloud Fulfillment Configuration Workbook"
    instructions_ws['A1'].font = Font(bold=True, size=16)
    
    instructions_ws['A3'] = "Purpose:"
    instructions_ws['A3'].font = Font(bold=True)
    instructions_ws['A4'] = "This workbook contains sheets for configuring Fulfillment Plan and related objects in Revenue Cloud."
    
    instructions_ws['A6'] = "Instructions:"
    instructions_ws['A6'].font = Font(bold=True)
    instructions_ws['A7'] = "1. Fill in data for each object in the corresponding sheet"
    instructions_ws['A8'] = "2. Required fields are marked in the field headers"
    instructions_ws['A9'] = "3. Upload objects in the order they appear (dependencies are ordered)"
    instructions_ws['A10'] = "4. Start with Location and OrderDeliveryMethod configuration"
    instructions_ws['A11'] = "5. Then configure WorkPlan templates"
    instructions_ws['A12'] = "6. Finally, configure fulfillment orders and shipments"
    
    instructions_ws['A14'] = "Object Categories:"
    instructions_ws['A14'].font = Font(bold=True)
    instructions_ws['A15'] = "• Core Configuration: Location, AssociatedLocation, OrderDeliveryMethod"
    instructions_ws['A16'] = "• Fulfillment Plans: WorkPlanTemplate, WorkPlanTemplateEntry, WorkPlan"
    instructions_ws['A17'] = "• Order Fulfillment: FulfillmentOrder, FulfillmentOrderLineItem, OrderDeliveryGroup"
    instructions_ws['A18'] = "• Shipment Management: Shipment, ShipmentItem"
    instructions_ws['A19'] = "• Product Configuration: ProductSellingModel, ProductSellingModelOption"
    instructions_ws['A20'] = "• Action Plans: ActionPlan, ActionPlanItemDependency"
    instructions_ws['A21'] = "• Asset Management: AssetRelationship, AssetStatePeriod"
    
    instructions_ws['A23'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    instructions_ws['A23'].font = Font(italic=True)
    
    # Adjust column width
    instructions_ws.column_dimensions['A'].width = 100
    
    # Create sheets for each object
    for obj_name, obj_info in FULFILLMENT_OBJECTS.items():
        sheet_name = obj_info['sheet_name']
        fields = obj_info['fields']
        
        # Create sheet
        ws = wb.create_sheet(sheet_name)
        
        # Add headers
        for col_idx, field in enumerate(fields, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = field
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            
            # Adjust column width
            ws.column_dimensions[cell.column_letter].width = max(len(field) + 2, 15)
        
        # Add sample/placeholder rows (2-3 rows)
        for row_idx in range(2, 4):
            for col_idx in range(1, len(fields) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                # Add placeholder text for specific fields
                if fields[col_idx-1] == 'Id' and row_idx == 2:
                    cell.value = f"// Leave blank for new records"
                elif fields[col_idx-1] == 'Name' and row_idx == 2:
                    cell.value = f"// Enter {obj_name} name"
    
    # Save workbook
    output_path = Path('/Users/marcdebrey/cpq-revenue-cloud-migration/POC/data/Revenue_Cloud_Fulfillment_Configuration.xlsx')
    wb.save(output_path)
    print(f"Created fulfillment configuration workbook: {output_path}")
    print(f"Total sheets: {len(wb.sheetnames)}")
    print(f"Objects included: {len(FULFILLMENT_OBJECTS)}")
    
    return str(output_path)

if __name__ == "__main__":
    create_fulfillment_workbook()