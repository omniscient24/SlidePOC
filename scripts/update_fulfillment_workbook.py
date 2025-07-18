#!/usr/bin/env python3
"""Update the fulfillment workbook with advanced fulfillment orchestration objects."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path

# Define additional advanced fulfillment objects
ADVANCED_FULFILLMENT_OBJECTS = {
    # Fulfillment Decomposition
    'AssetFulfillmentDecomp': {
        'sheet_name': '19_AssetFulfillmentDecomp',
        'fields': [
            'Id', 'Name', 'AssetId', 'FulfillmentOrderId', 'DecompositionDate',
            'Status', 'Description'
        ]
    },
    'ProductFulfillmentDecompRule': {
        'sheet_name': '20_ProductFulfillmentDecompRule',
        'fields': [
            'Id', 'Name', 'Product2Id', 'DecompositionType', 'Priority',
            'IsActive', 'Description', 'RuleExpression'
        ]
    },
    'ProductFulfillmentScenario': {
        'sheet_name': '21_ProductFulfillmentScenario',
        'fields': [
            'Id', 'Name', 'Product2Id', 'ScenarioType', 'Priority',
            'IsActive', 'Description'
        ]
    },
    
    # Fulfillment Assets
    'FulfillmentAsset': {
        'sheet_name': '22_FulfillmentAsset',
        'fields': [
            'Id', 'Name', 'FulfillmentOrderId', 'AssetId', 'Product2Id',
            'Status', 'Quantity', 'SerialNumber'
        ]
    },
    'FulfillmentAssetAttribute': {
        'sheet_name': '23_FulfillmentAssetAttribute',
        'fields': [
            'Id', 'FulfillmentAssetId', 'AttributeName', 'AttributeValue',
            'AttributeType'
        ]
    },
    'FulfillmentAssetRelationship': {
        'sheet_name': '24_FulfillmentAssetRelationship',
        'fields': [
            'Id', 'ParentFulfillmentAssetId', 'ChildFulfillmentAssetId',
            'RelationshipType', 'StartDate', 'EndDate'
        ]
    },
    
    # Fulfillment Steps and Workflow
    'FulfillmentStepDefinitionGroup': {
        'sheet_name': '25_FulfillmentStepDefinitionGroup',
        'fields': [
            'Id', 'Name', 'Description', 'GroupType', 'IsActive',
            'DisplayOrder'
        ]
    },
    'FulfillmentStepDefinition': {
        'sheet_name': '26_FulfillmentStepDefinition',
        'fields': [
            'Id', 'Name', 'FulfillmentStepDefinitionGroupId', 'StepType',
            'IsRequired', 'IsActive', 'ExecutionOrder', 'Description',
            'EstimatedDuration', 'EstimatedDurationUnit'
        ]
    },
    'FulfillmentStep': {
        'sheet_name': '27_FulfillmentStep',
        'fields': [
            'Id', 'Name', 'FulfillmentOrderId', 'FulfillmentStepDefinitionId',
            'Status', 'StartDate', 'EndDate', 'ActualDuration',
            'AssignedToId', 'CompletedById'
        ]
    },
    'FulfillmentStepDependencyDef': {
        'sheet_name': '28_FulfillmentStepDependencyDef',
        'fields': [
            'Id', 'PredecessorStepDefinitionId', 'SuccessorStepDefinitionId',
            'DependencyType', 'IsRequired'
        ]
    },
    'FulfillmentStepDependency': {
        'sheet_name': '29_FulfillmentStepDependency',
        'fields': [
            'Id', 'PredecessorStepId', 'SuccessorStepId', 'Status'
        ]
    },
    'FulfillmentStepSource': {
        'sheet_name': '30_FulfillmentStepSource',
        'fields': [
            'Id', 'FulfillmentStepId', 'SourceObjectId', 'SourceObjectType'
        ]
    },
    
    # Fulfillment Rules and Monitoring
    'FulfillmentFalloutRule': {
        'sheet_name': '31_FulfillmentFalloutRule',
        'fields': [
            'Id', 'Name', 'RuleType', 'Priority', 'IsActive',
            'ConditionLogic', 'ActionType', 'Description'
        ]
    },
    'FulfillmentStepJeopardyRule': {
        'sheet_name': '32_FulfillmentStepJeopardyRule',
        'fields': [
            'Id', 'Name', 'FulfillmentStepDefinitionId', 'ThresholdValue',
            'ThresholdUnit', 'AlertType', 'IsActive', 'Description'
        ]
    },
    'FulfillmentTaskAssignmentRule': {
        'sheet_name': '33_FulfillmentTaskAssignmentRule',
        'fields': [
            'Id', 'Name', 'AssignmentType', 'AssignToType', 'AssignToId',
            'Priority', 'IsActive', 'ConditionLogic'
        ]
    },
    
    # Fulfillment Line Items and Relationships
    'FulfillmentLineAttribute': {
        'sheet_name': '34_FulfillmentLineAttribute',
        'fields': [
            'Id', 'FulfillmentOrderLineItemId', 'AttributeName', 
            'AttributeValue', 'AttributeType'
        ]
    },
    'FulfillmentLineSourceRel': {
        'sheet_name': '35_FulfillmentLineSourceRel',
        'fields': [
            'Id', 'FulfillmentOrderLineItemId', 'SourceObjectId',
            'SourceObjectType', 'RelationshipType'
        ]
    },
    'FulfillmentLineRel': {
        'sheet_name': '36_FulfillmentLineRel',
        'fields': [
            'Id', 'ParentLineItemId', 'ChildLineItemId', 'RelationshipType'
        ]
    },
    
    # Fulfillment Tax and Adjustments
    'FulfillmentOrderItemAdjustment': {
        'sheet_name': '37_FulfillmentOrderItemAdjustment',
        'fields': [
            'Id', 'FulfillmentOrderLineItemId', 'AdjustmentType',
            'AdjustmentAmount', 'AdjustmentPercentage', 'Reason'
        ]
    },
    'FulfillmentOrderItemTax': {
        'sheet_name': '38_FulfillmentOrderItemTax',
        'fields': [
            'Id', 'FulfillmentOrderLineItemId', 'TaxType', 'TaxRate',
            'TaxAmount', 'TaxableAmount', 'JurisdictionName'
        ]
    },
    
    # Fulfillment Workspace
    'FulfillmentWorkspace': {
        'sheet_name': '39_FulfillmentWorkspace',
        'fields': [
            'Id', 'Name', 'WorkspaceType', 'Status', 'OwnerId',
            'Description', 'IsActive'
        ]
    },
    'FulfillmentWorkspaceItem': {
        'sheet_name': '40_FulfillmentWorkspaceItem',
        'fields': [
            'Id', 'FulfillmentWorkspaceId', 'ItemType', 'ItemId',
            'Status', 'Priority', 'AssignedToId'
        ]
    }
}

def update_fulfillment_workbook():
    """Update the existing fulfillment workbook with advanced objects."""
    workbook_path = Path('/Users/marcdebrey/cpq-revenue-cloud-migration/POC/data/Revenue_Cloud_Fulfillment_Configuration.xlsx')
    
    # Load existing workbook
    wb = openpyxl.load_workbook(workbook_path)
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Update Instructions sheet
    instructions_ws = wb['Instructions']
    
    # Add new section for advanced objects
    instructions_ws['A25'] = "Advanced Fulfillment Objects Added:"
    instructions_ws['A25'].font = Font(bold=True)
    instructions_ws['A26'] = "• Decomposition: AssetFulfillmentDecomp, ProductFulfillmentDecompRule, ProductFulfillmentScenario"
    instructions_ws['A27'] = "• Fulfillment Assets: FulfillmentAsset, FulfillmentAssetAttribute, FulfillmentAssetRelationship"
    instructions_ws['A28'] = "• Step Management: FulfillmentStepDefinitionGroup, FulfillmentStepDefinition, FulfillmentStep"
    instructions_ws['A29'] = "• Dependencies: FulfillmentStepDependencyDef, FulfillmentStepDependency, FulfillmentStepSource"
    instructions_ws['A30'] = "• Rules: FulfillmentFalloutRule, FulfillmentStepJeopardyRule, FulfillmentTaskAssignmentRule"
    instructions_ws['A31'] = "• Line Items: FulfillmentLineAttribute, FulfillmentLineSourceRel, FulfillmentLineRel"
    instructions_ws['A32'] = "• Adjustments: FulfillmentOrderItemAdjustment, FulfillmentOrderItemTax"
    instructions_ws['A33'] = "• Workspace: FulfillmentWorkspace, FulfillmentWorkspaceItem"
    
    instructions_ws['A35'] = "Upload Order:"
    instructions_ws['A35'].font = Font(bold=True)
    instructions_ws['A36'] = "1. Start with Definition Groups and Definitions (sheets 25-26)"
    instructions_ws['A37'] = "2. Configure Rules (sheets 31-33)"
    instructions_ws['A38'] = "3. Set up Decomposition Rules (sheets 19-21)"
    instructions_ws['A39'] = "4. Then proceed with runtime objects (Steps, Assets, etc.)"
    
    # Create sheets for each advanced object
    for obj_name, obj_info in ADVANCED_FULFILLMENT_OBJECTS.items():
        sheet_name = obj_info['sheet_name']
        fields = obj_info['fields']
        
        # Create sheet
        ws = wb.create_sheet(sheet_name)
        
        # Add headers
        for col_idx, field in enumerate(fields, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = field
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            
            # Adjust column width
            ws.column_dimensions[cell.column_letter].width = max(len(field) + 2, 15)
        
        # Add sample/placeholder rows
        for row_idx in range(2, 4):
            for col_idx in range(1, len(fields) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                # Add placeholder text for specific fields
                if fields[col_idx-1] == 'Id' and row_idx == 2:
                    cell.value = f"// Leave blank for new records"
                elif fields[col_idx-1] == 'Name' and row_idx == 2:
                    cell.value = f"// Enter {obj_name} name"
    
    # Save updated workbook
    wb.save(workbook_path)
    print(f"Updated fulfillment workbook: {workbook_path}")
    print(f"Total sheets now: {len(wb.sheetnames)}")
    print(f"Advanced objects added: {len(ADVANCED_FULFILLMENT_OBJECTS)}")
    
    # List all sheets
    print("\nComplete sheet list:")
    for idx, sheet_name in enumerate(wb.sheetnames):
        print(f"  {idx+1:2d}. {sheet_name}")

if __name__ == "__main__":
    update_fulfillment_workbook()