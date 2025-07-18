#!/usr/bin/env python3
"""
Update ProductCode and StockKeepingUnit for rows 27-35 in Product2 sheet
Based on Fortra Product Naming Convention Guide
"""

import pandas as pd
from openpyxl import load_workbook
import os

# Define the updates based on Fortra naming convention
product_updates = {
    'DCS Cloud': {
        'ProductCode': 'CYB-DCS-CLOUD-BUN',
        'StockKeepingUnit': 'SKU-CYB-DCS-CLOUD-BUN-SAAS-2024',
        'Description': 'Data Classification Suite Cloud Bundle'
    },
    'HRM Essentials Bundle': {
        'ProductCode': 'CYB-HRM-ESS-BUN',
        'StockKeepingUnit': 'SKU-CYB-HRM-ESS-BUN-SAAS-2024',
        'Description': 'Human Risk Management Essentials Bundle'
    },
    'HRM Managed Service Starter Pack': {
        'ProductCode': 'CYB-HRM-MSSP-BUN',
        'StockKeepingUnit': 'SKU-CYB-HRM-MSSP-BUN-SAAS-2024',
        'Description': 'HRM Managed Service Starter Pack'
    },
    'HRM Managed Service -Awareness Campaigns': {
        'ProductCode': 'CYB-HRM-AWC-SVC',
        'StockKeepingUnit': 'SKU-CYB-HRM-AWC-SVC-SAAS-2024',
        'Description': 'HRM Managed Service - Awareness Campaigns'
    },
    'HRM Managed Service-Phishing Campaigns': {
        'ProductCode': 'CYB-HRM-PHC-SVC',
        'StockKeepingUnit': 'SKU-CYB-HRM-PHC-SVC-SAAS-2024',
        'Description': 'HRM Managed Service - Phishing Campaigns'
    },
    'HRM Managed Service-Post Quiz': {
        'ProductCode': 'CYB-HRM-PQZ-SVC',
        'StockKeepingUnit': 'SKU-CYB-HRM-PQZ-SVC-SAAS-2024',
        'Description': 'HRM Managed Service - Post Quiz'
    },
    'HRM Security Advisory Services': {
        'ProductCode': 'CYB-HRM-SAS-SVC',
        'StockKeepingUnit': 'SKU-CYB-HRM-SAS-SVC-SAAS-2024',
        'Description': 'HRM Security Advisory Services'
    },
    'HRM Hourly Services': {
        'ProductCode': 'CYB-HRM-PS1-SVC',
        'StockKeepingUnit': 'SKU-CYB-HRM-PS1-SVC-2024',
        'Description': 'HRM Professional Services - Hourly'
    }
}

def update_product_codes():
    """Update ProductCode and SKU for specific products"""
    
    # File path
    file_path = 'data/Revenue_Cloud_Complete_Upload_Template_FINAL.xlsx'
    
    # Load the workbook
    print(f"Loading workbook: {file_path}")
    wb = load_workbook(file_path)
    sheet = wb['13_Product2']
    
    # Find column indices
    headers = [cell.value for cell in sheet[1]]
    name_col = headers.index('Name') + 1
    product_code_col = headers.index('ProductCode') + 1
    sku_col = headers.index('StockKeepingUnit') + 1
    
    print(f"\nColumn indices: Name={name_col}, ProductCode={product_code_col}, SKU={sku_col}")
    
    # Update rows 27-35 (Excel rows, so we start from row 27)
    updates_made = 0
    for row_num in range(27, 36):  # 27-35 inclusive
        name_cell = sheet.cell(row=row_num, column=name_col)
        if name_cell.value:
            # Clean the name (remove extra spaces)
            name = name_cell.value.strip()
            
            # Check if we have updates for this product
            if name in product_updates:
                updates = product_updates[name]
                
                # Update ProductCode
                product_code_cell = sheet.cell(row=row_num, column=product_code_col)
                product_code_cell.value = updates['ProductCode']
                
                # Update SKU
                sku_cell = sheet.cell(row=row_num, column=sku_col)
                sku_cell.value = updates['StockKeepingUnit']
                
                print(f"Row {row_num}: Updated '{name}'")
                print(f"  ProductCode: {updates['ProductCode']}")
                print(f"  SKU: {updates['StockKeepingUnit']}")
                
                updates_made += 1
    
    # Save the workbook
    print(f"\nSaving workbook with {updates_made} updates...")
    wb.save(file_path)
    print("Updates completed successfully!")
    
    # Verify the updates
    print("\nVerifying updates...")
    df = pd.read_excel(file_path, sheet_name='13_Product2')
    for idx in range(26, min(35, len(df))):
        row = df.iloc[idx]
        print(f"Row {idx+2}: {row['Name']} | ProductCode: {row.get('ProductCode', 'None')} | SKU: {row.get('StockKeepingUnit', 'None')}")

if __name__ == "__main__":
    update_product_codes()