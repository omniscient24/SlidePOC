#!/usr/bin/env python3
"""
Add ID columns as the first column to all sheets in the Revenue Cloud workbook - automated version
"""
import pandas as pd
import os
import shutil
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Mapping of sheet names to their corresponding object API names and ID field names
SHEET_TO_OBJECT_MAPPING = {
    '01_CostBook': ('CostBook', 'Id'),
    '02_LegalEntity': ('LegalEntity', 'Id'),
    '03_TaxEngine': ('TaxEngine', 'Id'),
    '04_TaxPolicy': ('TaxPolicy', 'Id'),
    '05_TaxTreatment': ('TaxTreatment', 'Id'),
    '06_BillingPolicy': ('BillingPolicy', 'Id'),
    '07_BillingTreatment': ('BillingTreatment', 'Id'),
    '08_ProductClassification': ('ProductClassification', 'Id'),
    '09_AttributeDefinition': ('AttributeDefinition', 'Id'),
    '10_AttributeCategory': ('AttributeCategory', 'Id'),
    '11_ProductCatalog': ('ProductCatalog', 'Id'),
    '12_ProductCategory': ('ProductCategory', 'Id'),
    '13_Product2': ('Product2', 'Id'),
    '14_AttributePicklist': ('AttributePicklist', 'Id'),
    '14_ProductComponentGroup': ('ProductComponentGroup', 'Id'),
    '15_CostBookEntry': ('CostBookEntry', 'Id'),
    '15_ProductSellingModel': ('ProductSellingModel', 'Id'),
    '17_ProductAttributeDef': ('ProductAttributeDefinition', 'Id'),
    '18_AttributePicklistValue': ('AttributePicklistValue', 'Id'),
    '19_Pricebook2': ('Pricebook2', 'Id'),
    '20_PricebookEntry': ('PricebookEntry', 'Id'),
    '21_PriceAdjustmentSchedule': ('PriceAdjustmentSchedule', 'Id'),
    '22_PriceAdjustmentTier': ('PriceAdjustmentTier', 'Id'),
    '23_AttributeBasedAdjRule': ('AttributeBasedAdjRule', 'Id'),
    '24_AttributeBasedAdj': ('AttributeBasedAdjustment', 'Id'),
    '25_ProductRelatedComponent': ('ProductRelatedComponent', 'Id'),
    '26_ProductCategoryProduct': ('ProductCategoryProduct', 'Id')
}

def add_id_columns_to_workbook(input_file, output_file):
    """Add ID columns as first column to all sheets"""
    
    print(f"Reading workbook: {input_file}")
    
    # Load the workbook with openpyxl to preserve formatting
    wb = load_workbook(input_file)
    
    # Also load with pandas for easier data manipulation
    xl_file = pd.ExcelFile(input_file)
    
    # Process each sheet
    for sheet_name in xl_file.sheet_names:
        print(f"\nProcessing sheet: {sheet_name}")
        
        # Read the sheet data
        df = pd.read_excel(xl_file, sheet_name=sheet_name)
        
        # Skip if sheet is empty
        if df.empty:
            print(f"  - Sheet is empty, skipping")
            continue
        
        # Determine the ID column name
        id_column_name = 'Id'  # Default
        if sheet_name in SHEET_TO_OBJECT_MAPPING:
            object_name, id_field = SHEET_TO_OBJECT_MAPPING[sheet_name]
            id_column_name = id_field
            print(f"  - Object: {object_name}, ID Field: {id_field}")
        else:
            print(f"  - Using default ID field name: Id")
        
        # Check if ID column already exists
        if id_column_name in df.columns:
            # If it exists but is not the first column, reorder
            if df.columns[0] != id_column_name:
                print(f"  - Moving existing '{id_column_name}' column to first position")
                cols = [id_column_name] + [col for col in df.columns if col != id_column_name]
                df = df[cols]
            else:
                print(f"  - '{id_column_name}' column already exists as first column")
        else:
            # Add ID column as first column
            print(f"  - Adding '{id_column_name}' column as first column")
            df.insert(0, id_column_name, '')
        
        # Clear the existing sheet
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                cell.value = None
        
        # Write the updated data back
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        print(f"  - Updated sheet with {len(df)} rows and {len(df.columns)} columns")
    
    # Save the updated workbook
    print(f"\nSaving updated workbook to: {output_file}")
    wb.save(output_file)
    print("Done!")

def main():
    """Main function"""
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    data_dir = os.path.join(base_dir, 'data')
    
    input_file = os.path.join(data_dir, 'Revenue_Cloud_Complete_Upload_Template_FINAL.xlsx')
    
    if not os.path.exists(input_file):
        print(f"Error: Input file not found: {input_file}")
        return
    
    # Create backup
    backup_file = input_file.replace('.xlsx', '_backup_before_id_columns.xlsx')
    print(f"Creating backup: {backup_file}")
    shutil.copy2(input_file, backup_file)
    
    # Process the file in-place
    temp_file = input_file.replace('.xlsx', '_temp.xlsx')
    add_id_columns_to_workbook(input_file, temp_file)
    
    # Replace original with updated file
    print(f"\nReplacing original file: {input_file}")
    shutil.move(temp_file, input_file)
    print("Original file updated successfully!")
    print(f"Backup saved as: {backup_file}")

if __name__ == "__main__":
    main()