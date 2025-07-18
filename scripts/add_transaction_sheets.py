#!/usr/bin/env python3
"""
Add Transaction Object Sheets to Revenue Cloud Workbook
Adds sheets for Order, OrderItem, Asset, AssetAction, AssetActionSource, and Contract
"""

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

# Define workbook path
WORKBOOK_PATH = Path('/Users/marcdebrey/cpq-revenue-cloud-migration/POC/data/Revenue_Cloud_Complete_Upload_Template_FINAL.xlsx')

# Define Transaction object structures
TRANSACTION_OBJECTS = {
    'Order': {
        'sheet_name': '27_Order',
        'fields': [
            'Id', 'Name', 'OrderNumber', 'AccountId', 'Status', 'EffectiveDate',
            'EndDate', 'ActivatedDate', 'ActivatedById', 'BillingStreet', 'BillingCity',
            'BillingState', 'BillingPostalCode', 'BillingCountry', 'ShippingStreet',
            'ShippingCity', 'ShippingState', 'ShippingPostalCode', 'ShippingCountry',
            'TotalAmount', 'Description', 'Pricebook2Id', 'OpportunityId', 'QuoteId',
            'ContractId', 'CustomerAuthorizedById', 'CompanyAuthorizedById',
            'Type', 'BillingFrequency', 'OrderReferenceNumber'
        ]
    },
    'OrderItem': {
        'sheet_name': '28_OrderItem',
        'fields': [
            'Id', 'Product2Id', 'IsDeleted', 'OrderId', 'PricebookEntryId',
            'Quantity', 'UnitPrice', 'ListPrice', 'TotalPrice', 'ServiceDate',
            'EndDate', 'Description', 'OrderItemNumber'
        ]
    },
    'Asset': {
        'sheet_name': '29_Asset',
        'fields': [
            'Id', 'Name', 'AccountId', 'Product2Id', 'SerialNumber', 'InstallDate',
            'PurchaseDate', 'UsageEndDate', 'LifecycleStartDate', 'LifecycleEndDate',
            'Status', 'Price', 'Quantity', 'Description', 'OwnerId', 'AssetProvidedById',
            'AssetServicedById', 'IsInternal', 'AssetLevel', 'StockKeepingUnit',
            'CurrentMrr', 'CurrentLifecycleEndDate', 'CurrentQuantity', 'CurrentAmount',
            'TotalLifecycleAmount'
        ]
    },
    'AssetAction': {
        'sheet_name': '30_AssetAction',
        'fields': [
            'Id', 'AssetId', 'Type', 'CategoryEnum', 'ActionDate', 'ProductAmendmentBehavior',
            'Amount', 'Quantity', 'StartDate', 'EndDate', 'QuantityChange', 'MrrChange',
            'AssetActionNumber'
        ]
    },
    'AssetActionSource': {
        'sheet_name': '31_AssetActionSource', 
        'fields': [
            'Id', 'AssetActionId', 'ReferenceEntityId', 'ExternalReference',
            'ExternalReferenceDataSource'
        ]
    },
    'Contract': {
        'sheet_name': '32_Contract',
        'fields': [
            'Id', 'AccountId', 'Pricebook2Id', 'OwnerId', 'Status', 'ContractNumber',
            'StartDate', 'EndDate', 'ContractTerm', 'BillingStreet', 'BillingCity',
            'BillingState', 'BillingPostalCode', 'BillingCountry', 'ShippingStreet',
            'ShippingCity', 'ShippingState', 'ShippingPostalCode', 'ShippingCountry',
            'CompanySignedId', 'CompanySignedDate', 'CustomerSignedId', 'CustomerSignedTitle',
            'CustomerSignedDate', 'SpecialTerms', 'Description', 'ActivatedById', 'ActivatedDate'
        ]
    }
}

def add_transaction_sheets():
    """Add Transaction object sheets to the workbook"""
    
    print(f"Loading workbook from {WORKBOOK_PATH}")
    
    # Load the workbook with openpyxl
    wb = load_workbook(WORKBOOK_PATH)
    
    # Define formatting styles to match existing sheets
    black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    left_align = Alignment(horizontal='left', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add each transaction object sheet
    for obj_name, obj_info in TRANSACTION_OBJECTS.items():
        sheet_name = obj_info['sheet_name']
        fields = obj_info['fields']
        
        print(f"\nAdding sheet: {sheet_name} for {obj_name}")
        
        # Check if sheet already exists
        if sheet_name in wb.sheetnames:
            print(f"  Sheet {sheet_name} already exists, skipping...")
            continue
        
        # Create new sheet
        ws = wb.create_sheet(title=sheet_name)
        
        # Turn off gridlines to match other sheets
        ws.sheet_view.showGridLines = False
        
        # Add headers
        for col_num, field_name in enumerate(fields, 1):
            cell = ws.cell(row=1, column=col_num, value=field_name)
            
            # Apply header formatting
            cell.fill = black_fill
            cell.font = white_font
            cell.alignment = left_align
            cell.border = thin_border
            
            # Set column width based on header length + padding
            column_letter = get_column_letter(col_num)
            ws.column_dimensions[column_letter].width = max(len(field_name) + 2, 12)
        
        # Add a sample data row (empty but formatted)
        for col_num in range(1, len(fields) + 1):
            cell = ws.cell(row=2, column=col_num, value="")
            cell.alignment = left_align
            cell.border = thin_border
        
        print(f"  Added {len(fields)} fields")
    
    # Save the workbook
    print(f"\nSaving workbook...")
    wb.save(WORKBOOK_PATH)
    print("✅ Transaction sheets added successfully!")
    
    # Verify sheets were added
    print("\nVerifying sheets...")
    wb_verify = load_workbook(WORKBOOK_PATH, read_only=True)
    for obj_name, obj_info in TRANSACTION_OBJECTS.items():
        sheet_name = obj_info['sheet_name']
        if sheet_name in wb_verify.sheetnames:
            print(f"  ✓ {sheet_name} verified")
        else:
            print(f"  ✗ {sheet_name} NOT FOUND")
    
    print("\nDone!")

if __name__ == "__main__":
    add_transaction_sheets()