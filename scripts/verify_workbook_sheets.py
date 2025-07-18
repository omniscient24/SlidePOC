#!/usr/bin/env python3
"""Verify sheets in the workbook and their properties."""

import openpyxl
from pathlib import Path
import sys

def verify_workbook_sheets():
    """Verify all sheets in the workbook."""
    workbook_path = Path('/Users/marcdebrey/cpq-revenue-cloud-migration/POC/data/Revenue_Cloud_Complete_Upload_Template_FINAL.xlsx')
    
    if not workbook_path.exists():
        print(f"ERROR: Workbook not found at {workbook_path}")
        return
    
    print(f"Verifying workbook: {workbook_path}")
    print(f"File size: {workbook_path.stat().st_size:,} bytes")
    print(f"Last modified: {workbook_path.stat().st_mtime}")
    print("\n" + "="*80 + "\n")
    
    try:
        # Load workbook
        wb = openpyxl.load_workbook(workbook_path, read_only=True)
        
        print(f"Total sheets in workbook: {len(wb.sheetnames)}")
        print("\nAll sheets:")
        print("-" * 40)
        
        # List all sheets with their index
        for idx, sheet_name in enumerate(wb.sheetnames):
            sheet = wb[sheet_name]
            # Get dimensions
            max_row = sheet.max_row
            max_col = sheet.max_column
            print(f"{idx+1:3d}. {sheet_name:<30} [{max_row} rows x {max_col} cols]")
        
        print("\n" + "="*80 + "\n")
        
        # Check specifically for Transaction object sheets
        print("Transaction Object Sheets Check:")
        print("-" * 40)
        
        expected_sheets = [
            ('27_Order', 'Order'),
            ('28_OrderItem', 'OrderItem'),
            ('29_Asset', 'Asset'),
            ('30_AssetAction', 'AssetAction'),
            ('31_AssetActionSource', 'AssetActionSource'),
            ('32_Contract', 'Contract')
        ]
        
        for sheet_name, object_name in expected_sheets:
            if sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                print(f"✓ {sheet_name:<25} FOUND - {sheet.max_row} rows, {sheet.max_column} columns")
                # Show first row (headers) if present
                if sheet.max_row > 0:
                    headers = []
                    for col in range(1, min(sheet.max_column + 1, 6)):  # Show first 5 headers
                        cell_value = sheet.cell(row=1, column=col).value
                        if cell_value:
                            headers.append(str(cell_value))
                    if headers:
                        print(f"  Headers: {', '.join(headers)}...")
            else:
                print(f"✗ {sheet_name:<25} NOT FOUND")
        
        wb.close()
        
    except Exception as e:
        print(f"ERROR reading workbook: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    verify_workbook_sheets()