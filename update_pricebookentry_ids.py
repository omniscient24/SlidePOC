#!/usr/bin/env python3
"""Update PricebookEntry IDs in workbook from success records"""

import pandas as pd
from openpyxl import load_workbook
import sys
import os

# Add parent directory to path
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from app.services.sync_service import SyncService

def main():
    workbook_path = "/Users/marcdebrey/cpq-revenue-cloud-migration/POC/data/Revenue_Cloud_Complete_Upload_Template_FINAL.xlsx"
    success_file = "/Users/marcdebrey/cpq-revenue-cloud-migration/POC/750dp00000C0JJtAAN-success-records.csv"
    
    print("Reading success records...")
    success_df = pd.read_csv(success_file)
    
    # Get records that were created (sf__Created = true)
    # The sf__Created field might be a boolean or string
    created_records = success_df[
        (success_df['sf__Created'] == True) | 
        (success_df['sf__Created'] == 'true') |
        (success_df['sf__Created'] == 'True')
    ].copy()
    print(f"Found {len(created_records)} newly created records")
    
    if len(created_records) == 0:
        print("No new records to update")
        return
    
    print("\nLoading workbook...")
    wb = load_workbook(workbook_path)
    ws = wb['20_PricebookEntry']
    
    # Create a mapping of Product2Id to row number in the sheet
    product_to_row = {}
    for row in range(2, ws.max_row + 1):  # Skip header row
        product_id = ws.cell(row=row, column=3).value  # Column C is Product2Id
        if product_id:
            product_to_row[product_id] = row
    
    print("\nUpdating IDs in workbook...")
    updates_made = 0
    
    for _, record in created_records.iterrows():
        product_id = record['Product2Id']
        new_id = record['sf__Id']
        
        if product_id in product_to_row:
            row_num = product_to_row[product_id]
            current_id = ws.cell(row=row_num, column=1).value
            
            if not current_id or current_id == '':
                ws.cell(row=row_num, column=1).value = new_id
                print(f"Updated row {row_num}: Product {product_id} -> ID {new_id}")
                updates_made += 1
            else:
                print(f"Row {row_num} already has ID: {current_id}")
    
    if updates_made > 0:
        print(f"\nSaving workbook with {updates_made} updates...")
        wb.save(workbook_path)
        
        # Apply formatting cleanup
        print("Applying formatting cleanup...")
        sync_service = SyncService()
        sync_service.workbook_path = workbook_path
        sync_service._cleanup_workbook_formatting()
        
        print("Done!")
    else:
        print("\nNo updates needed")

if __name__ == "__main__":
    main()